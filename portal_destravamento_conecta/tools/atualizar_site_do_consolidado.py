import argparse
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Instale a dependência: pip install -r requirements_portal.txt") from exc

ABAS_RESPONSAVEIS = {
    "WALCERI": "Walceir",
    "WALCEIR": "Walceir",
    "CAMILA": "Camila",
    "LUCASRAMOS": "LucasRamos",
    "LUCAS RAMOS": "LucasRamos",
}

STATUS_COLS = [
    "STATUS", "STATUS_EXECUCAO", "STATUS_DA_EXECUCAO", "SITUACAO", "SITUACAO_EXECUCAO",
    "STATUS_CONSOLIDACAO", "ESTADO", "DS_STATUS", "STATUS_CENARIO"
]

CENARIO_COLS = [
    "CENARIO", "CENARIO_TESTE", "NOME_CENARIO", "DESCRICAO_CENARIO", "DESCRICAO",
    "NOME", "TITULO", "ITEM", "ARQUIVO_ORIGEM"
]

FRENTE_COLS = [
    "FRENTE", "MODULO", "AREA", "GRUPO", "PROCESSO", "RELATORIO_ORIGEM"
]

OCORRENCIA_COLS = [
    "OCORRENCIA", "OCORRENCIAS", "STATUS_OCORRENCIA", "SITUACAO_OCORRENCIA", "DESCRICAO_OCORRENCIA",
    "OCORRENCIA_ABERTA", "QTD_OCORRENCIAS", "QUANTIDADE_OCORRENCIAS"
]

RESP_COLS = ["RESPONSAVEL", "RESPONSAVEL_TRATATIVA", "DONO_ACAO", "OWNER"]


def normalizar(texto):
    texto = str(texto or "")
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = texto.strip().upper()
    texto = re.sub(r"[^A-Z0-9]+", "_", texto)
    texto = re.sub(r"_+", "_", texto).strip("_")
    return texto


def valor_para_json(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d %H:%M:%S") if isinstance(valor, datetime) else valor.strftime("%Y-%m-%d")
    return valor


def linha_vazia(valores):
    return all(v is None or str(v).strip() == "" for v in valores)


def encontrar_primeira_coluna(row, headers_norm, candidatos):
    for candidato in candidatos:
        if candidato in headers_norm:
            return row.get(candidato)
    return None


def texto_concatenado(row):
    return " | ".join(str(v) for v in row.values() if v is not None and str(v).strip())


def contem(texto, *termos):
    base = normalizar(texto).replace("_", " ")
    return any(normalizar(t).replace("_", " ") in base for t in termos)


def to_bool_ocorrencia(valor, texto_geral):
    if valor is None or str(valor).strip() == "":
        return contem(texto_geral, "OCORRENCIA ABERTA", "OCORRENCIAS ABERTAS", "COM OCORRENCIA")
    if isinstance(valor, (int, float)):
        return valor > 0
    txt = normalizar(valor).replace("_", " ")
    if txt in {"S", "SIM", "TRUE", "VERDADEIRO", "1", "ABERTA", "ABERTO"}:
        return True
    return "ABERTA" in txt or "ABERTO" in txt or "OCORRENCIA" in txt


def classificar(row, headers_norm, responsavel_aba):
    status = encontrar_primeira_coluna(row, headers_norm, STATUS_COLS)
    cenario = encontrar_primeira_coluna(row, headers_norm, CENARIO_COLS)
    frente = encontrar_primeira_coluna(row, headers_norm, FRENTE_COLS)
    ocorrencia = encontrar_primeira_coluna(row, headers_norm, OCORRENCIA_COLS)
    resp_linha = encontrar_primeira_coluna(row, headers_norm, RESP_COLS)

    arquivo = row.get("ARQUIVO_ORIGEM") or ""
    relatorio = row.get("RELATORIO_ORIGEM") or ""
    status_consolidacao = row.get("STATUS_CONSOLIDACAO") or ""
    tipo_origem = row.get("TIPO_ORIGEM") or ""
    texto = texto_concatenado(row)

    status_txt = str(status or status_consolidacao or "Não classificado")
    cenario_txt = str(cenario or arquivo or "Item sem descrição")
    frente_txt = str(frente or relatorio or responsavel_aba or "Sem frente")
    responsavel = str(resp_linha or responsavel_aba or "Sem responsável")

    nao_iniciado = (
        contem(status_txt, "NAO INICIADO", "SEM DADOS")
        or contem(status_consolidacao, "SEM DADOS")
        or contem(arquivo, "NAO INICIADO", "NAO_INICIADO", "CENARIO E EXECUCAO NAO INICIADO")
        or contem(tipo_origem, "PLACEHOLDER")
    )

    bloqueado = contem(status_txt, "BLOQUEADO", "BLOCKED", "IMPEDIMENTO", "TRAVADO") or contem(texto, "BLOQUEADO", "IMPEDIMENTO ATIVO")
    ocorrencia_aberta = to_bool_ocorrencia(ocorrencia, texto)

    indice = (5 if bloqueado else 0) + (3 if ocorrencia_aberta else 0) + (1 if nao_iniciado else 0)

    return {
        "responsavel": responsavel,
        "frente": frente_txt,
        "relatorio_origem": str(relatorio or ""),
        "cenario": cenario_txt,
        "status": status_txt,
        "ocorrencia": str(ocorrencia or ""),
        "arquivo_origem": str(arquivo or ""),
        "tipo_origem": str(tipo_origem or ""),
        "bloqueado": bool(bloqueado),
        "ocorrencia_aberta": bool(ocorrencia_aberta),
        "nao_iniciado": bool(nao_iniciado),
        "indice_apoio": int(indice),
    }


def ler_resumo_execucao(wb):
    resumo = {
        "ultima_atualizacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "tempo_geracao": "-",
    }

    if "RESUMO_EXECUCAO" not in wb.sheetnames:
        return resumo

    ws = wb["RESUMO_EXECUCAO"]
    pares = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        chave = normalizar(row[0])
        valor = valor_para_json(row[1])
        if chave:
            pares[chave] = valor

    for chave_data in ["FIM_EXECUCAO", "FIM_DA_EXECUCAO", "DATA_FIM", "DATA_EXECUCAO", "INICIO_EXECUCAO"]:
        if chave_data in pares:
            resumo["ultima_atualizacao"] = str(pares[chave_data])
            break

    for chave_tempo in ["DURACAO_TOTAL", "TEMPO_TOTAL", "TEMPO_GERACAO", "DURACAO"]:
        if chave_tempo in pares:
            resumo["tempo_geracao"] = str(pares[chave_tempo])
            break

    return resumo


def converter(consolidado_path, saida_json):
    consolidado_path = Path(consolidado_path)
    saida_json = Path(saida_json)

    if not consolidado_path.exists():
        raise FileNotFoundError(f"Consolidado não encontrado: {consolidado_path}")

    wb = load_workbook(consolidado_path, read_only=True, data_only=True)
    itens = []

    for nome_aba in wb.sheetnames:
        responsavel_aba = ABAS_RESPONSAVEIS.get(normalizar(nome_aba).replace("_", " "))
        if not responsavel_aba:
            continue

        ws = wb[nome_aba]
        rows = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows)
        except StopIteration:
            continue

        if not header_raw:
            continue

        headers_norm = []
        usados = defaultdict(int)
        for h in header_raw:
            base = normalizar(h) or "COLUNA"
            usados[base] += 1
            nome_final = base if usados[base] == 1 else f"{base}_{usados[base]}"
            headers_norm.append(nome_final)

        for row_values in rows:
            if linha_vazia(row_values):
                continue
            row = {headers_norm[i]: valor_para_json(row_values[i]) if i < len(row_values) else None for i in range(len(headers_norm))}
            item = classificar(row, headers_norm, responsavel_aba)
            item["_aba_origem"] = nome_aba
            itens.append(item)

    resumo = ler_resumo_execucao(wb)
    resumo["arquivo_origem"] = str(consolidado_path.name)
    resumo["total_itens"] = len(itens)
    resumo["gerado_em"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    payload = {
        "resumo": resumo,
        "itens": itens,
    }

    saida_json.parent.mkdir(parents=True, exist_ok=True)
    with open(saida_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print("=" * 90)
    print("JSON do dashboard gerado com sucesso.")
    print(f"Consolidado origem: {consolidado_path}")
    print(f"JSON destino:       {saida_json}")
    print(f"Itens exportados:   {len(itens)}")
    print("=" * 90)


def main():
    parser = argparse.ArgumentParser(description="Atualiza o Portal de Destravamento usando o consolidado do GTN.")
    parser.add_argument("consolidado", help="Caminho do arquivo consolidado_HH-MM-SS.xlsx")
    parser.add_argument(
        "--saida",
        default=str(Path(__file__).resolve().parents[1] / "data" / "conecta_dashboard.json"),
        help="Caminho de saída do JSON consumido pelo site.",
    )
    args = parser.parse_args()
    converter(args.consolidado, args.saida)


if __name__ == "__main__":
    main()
