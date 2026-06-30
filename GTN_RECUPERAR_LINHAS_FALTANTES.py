# -*- coding: utf-8 -*-
"""
GTN_RECUPERAR_LINHAS_FALTANTES.py

Objetivo:
    Reprocessar somente as linhas TR que ficaram sem tratativa completa na planilha
    linhas_faltantes_por_relatorio.xlsx.

Uso recomendado, dentro do repositório SIT_PMO:

    python GTN_RECUPERAR_LINHAS_FALTANTES.py ^
        --planilha linhas_faltantes_por_relatorio.xlsx ^
        --aba Linhas_Faltantes ^
        --cred-dir credenciais ^
        --output-dir output/recuperacao_linhas_faltantes

Para testar sem clicar/baixar:

    python GTN_RECUPERAR_LINHAS_FALTANTES.py --dry-run

Observações importantes:
    1. O script procura arquivos .env no padrão já usado no projeto:
       credenciais/env_WALCEIR_P1.env, credenciais/env_CAMILA_P3.env etc.

    2. Cada .env precisa ter, no mínimo, usuário, senha, URL e o VALUE do relatório salvo.
       Chaves aceitas:
           Usuário: GTN_USUARIO, USUARIO, USERNAME, APEX_USERNAME
           Senha:   GTN_SENHA, SENHA, PASSWORD, APEX_PASSWORD
           URL:     GTN_URL, URL_LOGIN, LOGIN_URL, URL_EXECUCAO_TESTES, APEX_URL
           Value:   SAVED_REPORT_VALUE, RELATORIO_VALUE, VALUE_RELATORIO,
                    APEX_SAVED_REPORT_VALUE, RELATORIO_SAVED_REPORT_VALUE
           Região:  APEX_REGION_ID, REGION_ID, IR_REGION_ID

    3. O seletor padrão do relatório salvo é:
           #R35932200234408468_saved_reports
       Caso mude, informe APEX_REGION_ID no .env.

    4. Este script não usa credenciais fixas. Tudo vem do .env.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Faltou instalar openpyxl. Execute: pip install openpyxl"
    ) from exc

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import Page, sync_playwright
except ImportError as exc:
    raise SystemExit(
        "Faltou instalar Playwright. Execute:\n"
        "  pip install playwright\n"
        "  playwright install chromium"
    ) from exc


DEFAULT_APEX_REGION_ID = "35932200234408468"
DEFAULT_SHEET = "Linhas_Faltantes"
DEFAULT_PLANILHA = "linhas_faltantes_por_relatorio.xlsx"

STATUS_REPROCESSAR_PADRAO = {
    "INCONCLUSIVA",
    "SEM_LINHA_NAO_ENCONTRADA",
    "BURACO_SEQUENCIA",
}


@dataclass(frozen=True)
class LinhaRecuperacao:
    tipo_faltante: str
    relatorio: str
    execucao_id: str
    lider: str
    prioridade: str
    linha_tr: int
    primeira_tr: Optional[int]
    ultima_tr: Optional[int]
    observacao: str
    arquivo_origem: str


@dataclass
class RelatorioConfig:
    relatorio: str
    lider: str
    prioridade: str
    env_path: Path
    usuario: str
    senha: str
    url: str
    saved_report_value: str
    apex_region_id: str = DEFAULT_APEX_REGION_ID


def remover_acentos(texto: str) -> str:
    texto = texto or ""
    return "".join(
        ch for ch in unicodedata.normalize("NFD", texto)
        if unicodedata.category(ch) != "Mn"
    )


def normalizar_nome(texto: str) -> str:
    texto = remover_acentos(texto or "")
    texto = re.sub(r"[^A-Za-z0-9]+", "", texto)
    return texto.upper()


def parse_relatorio_nome(relatorio: str) -> Tuple[str, str]:
    """Extrai lider/prioridade de RELATORIO_13_RenatoMezzalira_P1."""
    partes = relatorio.split("_")
    if len(partes) >= 4:
        return partes[2], partes[3]
    return "", ""


def carregar_env(path: Path) -> Dict[str, str]:
    env: Dict[str, str] = {}
    if not path.exists():
        raise FileNotFoundError(f"Arquivo .env não encontrado: {path}")

    for linha in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        linha = linha.strip()
        if not linha or linha.startswith("#") or "=" not in linha:
            continue
        chave, valor = linha.split("=", 1)
        chave = chave.strip()
        valor = valor.strip().strip('"').strip("'")
        env[chave] = valor
    return env


def primeiro_valor(env: Dict[str, str], chaves: Iterable[str]) -> str:
    for chave in chaves:
        valor = env.get(chave)
        if valor not in (None, ""):
            return valor
    return ""


def localizar_env(cred_dir: Path, lider: str, prioridade: str, relatorio: str) -> Optional[Path]:
    lider_norm = normalizar_nome(lider)
    prio_norm = normalizar_nome(prioridade)
    rel_norm = normalizar_nome(relatorio)

    candidatos = [
        cred_dir / f"env_{lider_norm}_{prio_norm}.env",
        cred_dir / f".env_{lider_norm}_{prio_norm}.env",
        cred_dir / f".env_{lider_norm}_{prio_norm}",
        cred_dir / f"env_{lider}_{prioridade}.env",
        cred_dir / f".env_{lider}_{prioridade}.env",
        cred_dir / f"env_{relatorio}.env",
        cred_dir / f".env_{relatorio}.env",
    ]

    for c in candidatos:
        if c.exists():
            return c

    if cred_dir.exists():
        arquivos = list(cred_dir.glob("*.env")) + list(cred_dir.glob(".env*"))
        for arq in arquivos:
            nome = normalizar_nome(arq.name)
            if lider_norm in nome and prio_norm in nome:
                return arq
            if rel_norm in nome:
                return arq

    return None


def montar_config(relatorio: str, lider: str, prioridade: str, cred_dir: Path) -> RelatorioConfig:
    env_path = localizar_env(cred_dir, lider, prioridade, relatorio)
    if not env_path:
        raise FileNotFoundError(
            f"Não achei .env para {relatorio}. Esperado algo como: "
            f"{cred_dir / ('env_' + normalizar_nome(lider) + '_' + normalizar_nome(prioridade) + '.env')}"
        )

    env = carregar_env(env_path)
    usuario = primeiro_valor(env, ["GTN_USUARIO", "USUARIO", "USERNAME", "APEX_USERNAME", "LOGIN"])
    senha = primeiro_valor(env, ["GTN_SENHA", "SENHA", "PASSWORD", "APEX_PASSWORD"])
    url = primeiro_valor(env, ["URL_EXECUCAO_TESTES", "GTN_URL", "URL_LOGIN", "LOGIN_URL", "APEX_URL"])
    saved_report_value = primeiro_valor(env, [
        "SAVED_REPORT_VALUE",
        "RELATORIO_VALUE",
        "VALUE_RELATORIO",
        "APEX_SAVED_REPORT_VALUE",
        "RELATORIO_SAVED_REPORT_VALUE",
        "SAVED_REPORT",
        "VALUE",
    ])
    apex_region_id = primeiro_valor(env, ["APEX_REGION_ID", "REGION_ID", "IR_REGION_ID"]) or DEFAULT_APEX_REGION_ID

    faltando = []
    if not usuario:
        faltando.append("usuário")
    if not senha:
        faltando.append("senha")
    if not url:
        faltando.append("url")
    if not saved_report_value:
        faltando.append("saved_report_value")
    if faltando:
        raise ValueError(
            f".env incompleto para {relatorio}: {env_path}\n"
            f"Campos faltando: {', '.join(faltando)}"
        )

    return RelatorioConfig(
        relatorio=relatorio,
        lider=lider,
        prioridade=prioridade,
        env_path=env_path,
        usuario=usuario,
        senha=senha,
        url=url,
        saved_report_value=saved_report_value,
        apex_region_id=apex_region_id,
    )


def ler_planilha(planilha: Path, aba: str, tipos: Optional[set[str]] = None) -> List[LinhaRecuperacao]:
    if not planilha.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {planilha}")

    wb = load_workbook(planilha, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        raise ValueError(f"Aba '{aba}' não existe. Abas disponíveis: {', '.join(wb.sheetnames)}")

    ws = wb[aba]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []

    headers = [str(x or "").strip() for x in linhas[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, col, default=""):
        pos = idx.get(col)
        if pos is None or pos >= len(row):
            return default
        valor = row[pos]
        return "" if valor is None else str(valor).strip()

    saida: List[LinhaRecuperacao] = []
    for row in linhas[1:]:
        tipo = get(row, "TIPO_FALTANTE") or get(row, "STATUS_LINHA") or ""
        if tipos and tipo not in tipos:
            continue

        relatorio = get(row, "RELATORIO")
        linha_tr_txt = get(row, "LINHA_TR")
        if not relatorio or not linha_tr_txt:
            continue
        try:
            linha_tr = int(float(linha_tr_txt))
        except ValueError:
            continue

        lider = get(row, "LIDER")
        prioridade = get(row, "PRIORIDADE")
        if not lider or not prioridade:
            lider2, prio2 = parse_relatorio_nome(relatorio)
            lider = lider or lider2
            prioridade = prioridade or prio2

        def to_int_or_none(txt: str) -> Optional[int]:
            try:
                return int(float(txt)) if txt else None
            except ValueError:
                return None

        saida.append(
            LinhaRecuperacao(
                tipo_faltante=tipo,
                relatorio=relatorio,
                execucao_id=get(row, "EXECUCAO_ID"),
                lider=lider,
                prioridade=prioridade,
                linha_tr=linha_tr,
                primeira_tr=to_int_or_none(get(row, "PRIMEIRA_TR")),
                ultima_tr=to_int_or_none(get(row, "ULTIMA_TR")),
                observacao=get(row, "OBSERVACAO"),
                arquivo_origem=get(row, "ARQUIVO_ORIGEM"),
            )
        )

    return saida


def agrupar_por_relatorio(linhas: List[LinhaRecuperacao]) -> Dict[str, List[LinhaRecuperacao]]:
    grupos: Dict[str, List[LinhaRecuperacao]] = {}
    for item in linhas:
        grupos.setdefault(item.relatorio, []).append(item)
    for rel in grupos:
        grupos[rel].sort(key=lambda x: x.linha_tr)
    return grupos


def esperar_apex(page: Page, timeout_ms: int = 1500) -> None:
    """Espera curta por overlays comuns do Oracle APEX; se não existir, segue."""
    seletores = [
        ".u-Processing",
        ".a-Processing",
        ".apex_wait_overlay",
        ".a-IRR-progress",
    ]
    for sel in seletores:
        try:
            page.locator(sel).first.wait_for(state="hidden", timeout=timeout_ms)
        except Exception:
            pass


def login_gtn(page: Page, cfg: RelatorioConfig) -> None:
    page.goto(cfg.url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(1200)

    # Se já estiver logado, os campos podem nem aparecer.
    usuario_box = page.get_by_role("textbox", name=re.compile("Usuário|Usuario|User", re.I))
    senha_box = page.get_by_role("textbox", name=re.compile("Senha|Password", re.I))

    if usuario_box.count() > 0:
        try:
            usuario_box.first.fill(cfg.usuario, timeout=5000)
            senha_box.first.fill(cfg.senha, timeout=5000)
            page.get_by_role("button", name=re.compile("Acessar|Entrar|Login|Sign in", re.I)).first.click(timeout=10_000)
            page.wait_for_load_state("domcontentloaded", timeout=60_000)
            esperar_apex(page)
        except Exception as exc:
            print(f"[AVISO] Falha tentando logar automaticamente: {exc}")


def abrir_execucao_testes(page: Page, cfg: RelatorioConfig) -> None:
    """Garante que a página do relatório está aberta."""
    # Se URL já é a página de execução, o login já deixou no lugar correto.
    # Caso contrário, tenta caminho pela navegação principal.
    try:
        if page.locator(f"#R{cfg.apex_region_id}_saved_reports").count() > 0:
            return
    except Exception:
        pass

    try:
        page.get_by_role("button", name=re.compile("Navegação Principal|Navegacao Principal", re.I)).click(timeout=5000)
        page.get_by_role("treeitem", name=re.compile("Execução de Testes|Execucao de Testes", re.I)).click(timeout=10_000)
        page.wait_for_load_state("domcontentloaded", timeout=60_000)
        esperar_apex(page)
    except Exception:
        # Último recurso: fica na URL já carregada.
        pass


def aplicar_relatorio_salvo(page: Page, cfg: RelatorioConfig) -> None:
    seletor = f"#R{cfg.apex_region_id}_saved_reports"
    combo = page.locator(seletor)
    combo.wait_for(state="visible", timeout=30_000)
    combo.select_option(cfg.saved_report_value)
    page.wait_for_timeout(800)

    # Em alguns APEX existe botão GO/Ir após escolher relatório salvo.
    possiveis_go = [
        page.get_by_role("button", name=re.compile(r"^(Ir|Go|Aplicar)$", re.I)),
        page.locator(f"#R{cfg.apex_region_id}_saved_reports_go"),
        page.locator("button:has-text('Ir')"),
        page.locator("button:has-text('Go')"),
    ]
    for loc in possiveis_go:
        try:
            if loc.count() > 0 and loc.first.is_visible(timeout=1000):
                loc.first.click(timeout=3000)
                break
        except Exception:
            continue

    page.wait_for_load_state("domcontentloaded", timeout=60_000)
    esperar_apex(page)
    page.wait_for_timeout(1200)


def obter_contextos(page: Page):
    """Retorna page + frames, pois o GTN usa iframes em partes do fluxo."""
    contextos = [page]
    for frame in page.frames:
        if frame == page.main_frame:
            continue
        contextos.append(frame)
    return contextos


def click_com_download(page: Page, locator, destino_dir: Path, prefixo: str, timeout_ms: int = 10_000) -> Optional[Path]:
    destino_dir.mkdir(parents=True, exist_ok=True)
    try:
        with page.expect_download(timeout=timeout_ms) as download_info:
            locator.click(timeout=timeout_ms)
        download = download_info.value
        nome = download.suggested_filename or f"{prefixo}_{int(time.time())}.bin"
        nome = re.sub(r"[\\/:*?\"<>|]+", "_", nome)
        destino = destino_dir / nome
        if destino.exists():
            base = destino.stem
            ext = destino.suffix
            destino = destino_dir / f"{base}_{int(time.time())}{ext}"
        download.save_as(str(destino))
        return destino
    except PlaywrightTimeoutError:
        return None
    except Exception as exc:
        print(f"[AVISO] Clique não gerou download: {exc}")
        return None


def tentar_baixar_arquivos_da_pagina(page: Page, destino_dir: Path, prefixo: str, max_downloads: int) -> List[Path]:
    """Procura botões/links de download na página atual e nos iframes."""
    baixados: List[Path] = []
    seletores_css = [
        "a[download]",
        "a[href*='download' i]",
        "a[href*='wwv_flow_file_mgr' i]",
        "a[href*='get_blob' i]",
        "button[title*='Download' i]",
        "a[title*='Download' i]",
        "button[aria-label*='Download' i]",
        "a[aria-label*='Download' i]",
    ]
    textos = re.compile(r"Download|Baixar|Arquivo|Anexo|Exportar", re.I)

    for contexto in obter_contextos(page):
        # CSS primeiro.
        for sel in seletores_css:
            try:
                loc = contexto.locator(sel)
                count = min(loc.count(), 20)
                for i in range(count):
                    if len(baixados) >= max_downloads:
                        return baixados
                    item = loc.nth(i)
                    try:
                        if not item.is_visible(timeout=800):
                            continue
                    except Exception:
                        continue
                    destino = click_com_download(page, item, destino_dir, f"{prefixo}_{len(baixados)+1}")
                    if destino:
                        print(f"        [DOWNLOAD] {destino.name}")
                        baixados.append(destino)
            except Exception:
                continue

        # Texto depois.
        try:
            loc = contexto.get_by_text(textos)
            count = min(loc.count(), 20)
            for i in range(count):
                if len(baixados) >= max_downloads:
                    return baixados
                item = loc.nth(i)
                try:
                    if not item.is_visible(timeout=800):
                        continue
                except Exception:
                    continue
                destino = click_com_download(page, item, destino_dir, f"{prefixo}_{len(baixados)+1}")
                if destino:
                    print(f"        [DOWNLOAD] {destino.name}")
                    baixados.append(destino)
        except Exception:
            continue

    return baixados


def abrir_linha_tr(page: Page, linha_tr: int) -> bool:
    """Abre a linha do Interactive Report usando o padrão tr:nth-child(n)."""
    seletores = [
        f"tr:nth-child({linha_tr}) > .a-IRR-linkCol",
        f"tr:nth-child({linha_tr}) .a-IRR-linkCol",
        f"tr:nth-child({linha_tr}) a",
    ]
    for sel in seletores:
        loc = page.locator(sel).first
        try:
            loc.wait_for(state="visible", timeout=5000)
            loc.click(timeout=10_000)
            page.wait_for_load_state("domcontentloaded", timeout=60_000)
            esperar_apex(page)
            page.wait_for_timeout(1000)
            return True
        except Exception:
            continue
    return False


def processar_linha(
    page: Page,
    cfg: RelatorioConfig,
    linha: LinhaRecuperacao,
    output_dir: Path,
    max_downloads_per_line: int,
) -> Tuple[str, int, str]:
    """Retorna status, qtd_downloads, detalhe."""
    print(f"    TR {linha.linha_tr} | {linha.tipo_faltante}")

    # Reabre o relatório a cada linha. É mais lento, mas evita ficar perdido no detalhe.
    page.goto(cfg.url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(1000)
    abrir_execucao_testes(page, cfg)
    aplicar_relatorio_salvo(page, cfg)

    if not abrir_linha_tr(page, linha.linha_tr):
        return "SEM_LINHA", 0, f"Não abriu tr:nth-child({linha.linha_tr})"

    destino = output_dir / cfg.relatorio / f"TR_{linha.linha_tr:03d}"
    baixados = tentar_baixar_arquivos_da_pagina(
        page=page,
        destino_dir=destino,
        prefixo=f"{cfg.relatorio}_TR_{linha.linha_tr:03d}",
        max_downloads=max_downloads_per_line,
    )

    if baixados:
        return "DOWNLOAD_OK", len(baixados), "; ".join(p.name for p in baixados)

    return "SEM_DOWNLOAD_ENCONTRADO", 0, "Linha abriu, mas nenhum botão/link de download respondeu. Validar seletor específico do GTN."


def escrever_resultado_header(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([
            "DATA_HORA",
            "RELATORIO",
            "LIDER",
            "PRIORIDADE",
            "LINHA_TR",
            "TIPO_ORIGEM",
            "STATUS_RECUPERACAO",
            "QTD_DOWNLOADS",
            "DETALHE",
            "ENV_USADO",
        ])


def escrever_resultado(path: Path, cfg: RelatorioConfig, linha: LinhaRecuperacao, status: str, qtd: int, detalhe: str) -> None:
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            cfg.relatorio,
            cfg.lider,
            cfg.prioridade,
            linha.linha_tr,
            linha.tipo_faltante,
            status,
            qtd,
            detalhe,
            str(cfg.env_path),
        ])


def main() -> int:
    parser = argparse.ArgumentParser(description="Recupera downloads de linhas faltantes do GTN.")
    parser.add_argument("--planilha", default=DEFAULT_PLANILHA, help="Planilha linhas_faltantes_por_relatorio.xlsx")
    parser.add_argument("--aba", default=DEFAULT_SHEET, help="Aba de origem: Linhas_Faltantes ou Sem_Download")
    parser.add_argument("--cred-dir", default="credenciais", help="Pasta onde ficam os .env")
    parser.add_argument("--output-dir", default="output/recuperacao_linhas_faltantes", help="Pasta para salvar downloads e log")
    parser.add_argument("--headless", action="store_true", help="Executar Chromium sem abrir janela")
    parser.add_argument("--dry-run", action="store_true", help="Só mostra o que seria reprocessado")
    parser.add_argument("--max-downloads-per-line", type=int, default=10, help="Limite de arquivos por linha")
    parser.add_argument("--status", default="INCONCLUSIVA,SEM_LINHA_NAO_ENCONTRADA,BURACO_SEQUENCIA", help="Tipos/status a reprocessar separados por vírgula. Use ALL para tudo da aba.")
    args = parser.parse_args()

    planilha = Path(args.planilha).resolve()
    cred_dir = Path(args.cred_dir).resolve()
    output_dir = Path(args.output_dir).resolve()

    tipos = None if args.status.upper() == "ALL" else {x.strip() for x in args.status.split(",") if x.strip()}
    linhas = ler_planilha(planilha, args.aba, tipos=tipos)
    grupos = agrupar_por_relatorio(linhas)

    print("=" * 100)
    print("RECUPERAÇÃO DE LINHAS FALTANTES GTN")
    print("=" * 100)
    print(f"Planilha : {planilha}")
    print(f"Aba      : {args.aba}")
    print(f"Linhas   : {len(linhas)}")
    print(f"Relatórios: {len(grupos)}")
    print(f"Saída    : {output_dir}")
    print("-" * 100)

    if not linhas:
        print("Nenhuma linha para reprocessar.")
        return 0

    for relatorio, itens in grupos.items():
        print(f"{relatorio}: {', '.join('TR ' + str(i.linha_tr) for i in itens)}")

    if args.dry_run:
        print("\nDRY-RUN ativo: nada foi clicado/baixado.")
        return 0

    log_path = output_dir / f"resultado_recuperacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    escrever_resultado_header(log_path)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=0)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(20_000)

        try:
            for relatorio, itens in grupos.items():
                lider = itens[0].lider
                prioridade = itens[0].prioridade
                print("\n" + "=" * 100)
                print(f"RELATÓRIO: {relatorio} | {lider} {prioridade}")
                print("=" * 100)

                try:
                    cfg = montar_config(relatorio, lider, prioridade, cred_dir)
                    print(f"[ENV] {cfg.env_path}")
                except Exception as exc:
                    print(f"[ERRO CONFIG] {exc}")
                    fake_cfg = RelatorioConfig(relatorio, lider, prioridade, Path(""), "", "", "", "")
                    for linha in itens:
                        escrever_resultado(log_path, fake_cfg, linha, "ERRO_CONFIG", 0, str(exc))
                    continue

                try:
                    login_gtn(page, cfg)
                    abrir_execucao_testes(page, cfg)
                except Exception as exc:
                    print(f"[ERRO LOGIN/NAVEGAÇÃO] {exc}")
                    for linha in itens:
                        escrever_resultado(log_path, cfg, linha, "ERRO_LOGIN_NAVEGACAO", 0, str(exc))
                    continue

                for linha in itens:
                    try:
                        status, qtd, detalhe = processar_linha(
                            page=page,
                            cfg=cfg,
                            linha=linha,
                            output_dir=output_dir,
                            max_downloads_per_line=args.max_downloads_per_line,
                        )
                        print(f"        => {status} | downloads={qtd}")
                        escrever_resultado(log_path, cfg, linha, status, qtd, detalhe)
                    except Exception as exc:
                        print(f"        => ERRO | {exc}")
                        escrever_resultado(log_path, cfg, linha, "ERRO", 0, repr(exc))
        finally:
            context.close()
            browser.close()

    print("\n" + "=" * 100)
    print(f"Processo finalizado. Log gerado em: {log_path}")
    print("=" * 100)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
