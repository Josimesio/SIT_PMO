import csv
import os
import re
import threading
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

GTN_USUARIO = os.getenv("GTN_USUARIO") or os.getenv("GTN_USER")
GTN_SENHA = os.getenv("GTN_SENHA") or os.getenv("GTN_PASS")

URL_LOGIN = os.getenv(
    "URL_LOGIN",
    "https://gtn.ninecon.com.br/ords/r/gtn/gtn/login?session=601689329474445&tz=-3:00",
)

TIMEOUT_PADRAO = int(os.getenv("TIMEOUT_PADRAO", "60000"))
TIMEOUT_DOWNLOAD = int(os.getenv("TIMEOUT_DOWNLOAD", "180000"))
TIMEOUT_LINHA_CURTO = int(os.getenv("TIMEOUT_LINHA_CURTO", "5000"))
TIMEOUT_EDIT_EXECUCAO = int(os.getenv("TIMEOUT_EDIT_EXECUCAO", "20000"))
MAX_LINHAS_INEXISTENTES_SEGUIDAS = int(os.getenv("MAX_LINHAS_INEXISTENTES_SEGUIDAS", "10"))

DOWNLOAD_ROOT = BASE_DIR / os.getenv("DOWNLOAD_DIR", "downloads_gtn")

GERAR_PLACEHOLDER_SEM_EDIT = os.getenv("GERAR_PLACEHOLDER_SEM_EDIT", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}
TEMPLATE_NAO_INICIADO = os.getenv("TEMPLATE_NAO_INICIADO", "").strip()

GERAR_CONSOLIDADO_FINAL = os.getenv("GERAR_CONSOLIDADO_FINAL", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}
EXIBIR_TELA_TEMPO = os.getenv("EXIBIR_TELA_TEMPO", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}

# Arquivos gerados na execução atual.
# Usado para consolidar somente o que acabou de ser baixado/criado, sem misturar execuções antigas.
ARQUIVOS_DOWNLOAD_EXECUCAO = []

# Guarda o primeiro arquivo baixado com sucesso por relatório para usar como modelo
# quando algum cenário não tiver link Edit.
MODELOS_DOWNLOAD_POR_RELATORIO = {}


# Responsável que será concatenado no nome dos downloads por relatório.
# Regra oficial solicitada:
# RELATORIO_01 a RELATORIO_04 = Walceir
# RELATORIO_05 a RELATORIO_08 = Camila
# RELATORIO_09 a RELATORIO_12 = LucasRamos
RESPONSAVEL_POR_RELATORIO_NOME = {
    "RELATORIO_01": "Walceir",
    "RELATORIO_02": "Walceir",
    "RELATORIO_03": "Walceir",
    "RELATORIO_04": "Walceir",
    "RELATORIO_05": "Camila",
    "RELATORIO_06": "Camila",
    "RELATORIO_07": "Camila",
    "RELATORIO_08": "Camila",
    "RELATORIO_09": "LucasRamos",
    "RELATORIO_10": "LucasRamos",
    "RELATORIO_11": "LucasRamos",
    "RELATORIO_12": "LucasRamos",
}

# Mantém a regra mesmo se o nome do relatório mudar, usando o VALUE real salvo do APEX.
RESPONSAVEL_POR_SAVED_VALUE = {
    "96443782140269930": "Walceir",
    "96445783233272755": "Walceir",
    "96447734078275977": "Walceir",
    "96461147945312616": "Walceir",
    "96466495992333195": "Camila",
    "96468946665336064": "Camila",
    "96470998136338958": "Camila",
    "96473161397342454": "Camila",
    "96454201749302012": "LucasRamos",
    "96451968465288102": "LucasRamos",
    "96449983914285918": "LucasRamos",
    "96456284157307243": "LucasRamos",
}


def resolver_responsavel_relatorio(nome, saved_value):
    """Resolve o nome que será concatenado no download do relatório."""
    nome = (nome or "").strip().upper()
    saved_value = (saved_value or "").strip()

    return (
        RESPONSAVEL_POR_RELATORIO_NOME.get(nome)
        or RESPONSAVEL_POR_SAVED_VALUE.get(saved_value)
        or os.getenv("RELATORIO_RESPONSAVEL", "").strip()
    )


def prefixo_responsavel_download(relatorio):
    """Retorna o responsável já higienizado para compor o nome do arquivo."""
    responsavel = (relatorio.get("responsavel") or "").strip()
    return limpar_nome_arquivo(responsavel) if responsavel else ""


def aba_consolidado_responsavel(responsavel):
    """
    Define em qual aba o arquivo deve entrar no consolidado.

    Observação: a aba WALCERI foi mantida exatamente conforme solicitado.
    """
    chave = re.sub(r"[^A-Z0-9]", "", (responsavel or "").upper())

    if chave in {"WALCEIR", "WALCERI"}:
        return "WALCERI"

    if chave == "CAMILA":
        return "CAMILA"

    if chave in {"LUCASRAMOS", "LUCAS"}:
        return "LUCASRAMOS"

    return "SEM_RESPONSAVEL"


def registrar_arquivo_para_consolidacao(relatorio, caminho_arquivo, tipo_origem):
    """Registra arquivo gerado na execução atual para consolidação final."""
    if not caminho_arquivo:
        return

    caminho = Path(caminho_arquivo)

    if not caminho.exists():
        return

    responsavel = relatorio.get("responsavel") or ""
    aba_destino = aba_consolidado_responsavel(responsavel)

    if aba_destino == "SEM_RESPONSAVEL":
        print(
            f"[CONSOLIDADO] Responsável não mapeado para {caminho.name}. "
            "Arquivo não entrará nas abas WALCERI/CAMILA/LUCASRAMOS."
        )
        return

    ARQUIVOS_DOWNLOAD_EXECUCAO.append(
        {
            "caminho": caminho,
            "relatorio": relatorio.get("nome", ""),
            "saved_value": relatorio.get("saved_value", ""),
            "responsavel": responsavel,
            "aba": aba_destino,
            "tipo": tipo_origem,
        }
    )


# ============================================================
# CRONÔMETRO / AUDITORIA DE EXECUÇÃO
# ============================================================
def formatar_duracao(segundos):
    segundos = int(max(segundos or 0, 0))
    horas, resto = divmod(segundos, 3600)
    minutos, segundos = divmod(resto, 60)
    return f"{horas:02d}:{minutos:02d}:{segundos:02d}"


class CronometroExecucao:
    """
    Janela simples para acompanhar o tempo de execução.

    A tela é opcional e pode ser desativada no .env com:
    EXIBIR_TELA_TEMPO=N
    """

    def __init__(self, inicio_execucao):
        self.inicio_execucao = inicio_execucao
        self._inicio_perf = time.perf_counter()
        self._parar = False
        self._status = "Iniciando execução..."
        self._lock = threading.Lock()
        self._thread = None

    def iniciar(self):
        if not EXIBIR_TELA_TEMPO:
            return

        self._thread = threading.Thread(target=self._executar_janela, daemon=True)
        self._thread.start()

    def atualizar_status(self, status):
        with self._lock:
            self._status = status or "Executando..."

    def finalizar(self):
        with self._lock:
            self._status = "Finalizado. Consolidado gerado."
            self._parar = True

    def _executar_janela(self):
        try:
            import tkinter as tk
        except Exception as e:
            print(f"[TEMPO][AVISO] Não consegui abrir tela de tempo. Seguindo apenas pelo terminal. Detalhe: {e}")
            return

        try:
            root = tk.Tk()
            root.title("GTN - Tempo de Execução")
            root.geometry("460x185")
            root.resizable(False, False)

            try:
                root.attributes("-topmost", True)
            except Exception:
                pass

            frame = tk.Frame(root, padx=18, pady=16)
            frame.pack(fill="both", expand=True)

            titulo = tk.Label(frame, text="Automação GTN em execução", font=("Segoe UI", 13, "bold"))
            titulo.pack(anchor="w")

            lbl_inicio = tk.Label(
                frame,
                text=f"Início: {self.inicio_execucao.strftime('%d/%m/%Y %H:%M:%S')}",
                font=("Segoe UI", 10),
            )
            lbl_inicio.pack(anchor="w", pady=(8, 0))

            lbl_tempo = tk.Label(frame, text="Tempo decorrido: 00:00:00", font=("Segoe UI", 18, "bold"))
            lbl_tempo.pack(anchor="w", pady=(10, 4))

            lbl_status = tk.Label(frame, text="Status: Iniciando execução...", font=("Segoe UI", 10), wraplength=420, justify="left")
            lbl_status.pack(anchor="w")

            def atualizar():
                with self._lock:
                    parar = self._parar
                    status = self._status

                segundos = time.perf_counter() - self._inicio_perf
                lbl_tempo.config(text=f"Tempo decorrido: {formatar_duracao(segundos)}")
                lbl_status.config(text=f"Status: {status}")

                if parar:
                    root.after(1200, root.destroy)
                    return

                root.after(1000, atualizar)

            atualizar()
            root.mainloop()

        except Exception as e:
            print(f"[TEMPO][AVISO] A tela de tempo falhou, mas o script seguirá normalmente. Detalhe: {e}")


# ============================================================
# RELATÓRIOS PELO .ENV
#
# Formato da variável RELATORIOS:
# nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim
#
# Exemplo:
# RELATORIOS=PMO_EXECUCAO_TESTES|35932200234408468|96461147945312616|0|5|9|1002;OUTRO_RELATORIO|111|222|0|5|9|1002
# ============================================================
def carregar_relatorios_env():
    relatorios_raw = os.getenv("RELATORIOS", "").strip()

    # Compatibilidade com o modelo antigo, caso você ainda deixe as variáveis separadas no .env.
    if not relatorios_raw:
        apex_id = os.getenv("RELATORIO_APEX_ID", "").strip()
        saved_value = os.getenv("RELATORIO_SALVO_VALUE", "").strip()

        if apex_id and saved_value:
            return [
                {
                    "nome": os.getenv("RELATORIO_NOME", "PMO_EXECUCAO_TESTES").strip(),
                    "apex_id": apex_id,
                    "saved_value": saved_value,
                    "responsavel": resolver_responsavel_relatorio(
                        os.getenv("RELATORIO_NOME", "PMO_EXECUCAO_TESTES").strip(),
                        saved_value,
                    ),
                    "editar_inicio": int(os.getenv("EDITAR_INDICE_INICIO", "0")),
                    "editar_fim": int(os.getenv("EDITAR_INDICE_FIM", "5")),
                    "tr_inicio": int(os.getenv("TR_CHILD_INICIO", "9")),
                    "tr_fim": int(os.getenv("TR_CHILD_FIM", "1002")),
                }
            ]

        raise Exception(
            "Nenhum relatório configurado. Informe RELATORIOS no .env no formato: "
            "nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim"
        )

    relatorios = []

    for posicao, item in enumerate(relatorios_raw.split(";"), start=1):
        item = item.strip()

        if not item:
            continue

        partes = [parte.strip() for parte in item.split("|")]

        if len(partes) != 7:
            raise Exception(
                f"Formato inválido no relatório {posicao} da variável RELATORIOS. "
                "Use: nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim"
            )

        nome, apex_id, saved_value, editar_inicio, editar_fim, tr_inicio, tr_fim = partes

        if not nome:
            raise Exception(f"Nome vazio no relatório {posicao} da variável RELATORIOS.")

        if not apex_id:
            raise Exception(f"apex_id vazio no relatório {posicao} da variável RELATORIOS.")

        if not saved_value:
            raise Exception(f"saved_value vazio no relatório {posicao} da variável RELATORIOS.")

        relatorios.append(
            {
                "nome": nome,
                "apex_id": apex_id,
                "saved_value": saved_value,
                "responsavel": resolver_responsavel_relatorio(nome, saved_value),
                "editar_inicio": int(editar_inicio),
                "editar_fim": int(editar_fim),
                "tr_inicio": int(tr_inicio),
                "tr_fim": int(tr_fim),
            }
        )

    if not relatorios:
        raise Exception("A variável RELATORIOS está vazia ou inválida no .env.")

    return relatorios


RELATORIOS = carregar_relatorios_env()


# ============================================================
# UTILITÁRIOS
# ============================================================
def limpar_nome_arquivo(nome):
    nome = nome or "download"
    nome = re.sub(r'[\\/:*?"<>|]+', "_", nome)
    nome = re.sub(r"\s+", "_", nome).strip("._ ")
    return nome or "download"


def seletor_select_relatorio(relatorio):
    return f"#R{relatorio['apex_id']}_saved_reports"


def seletor_botao_ir_relatorio(relatorio):
    return f"#R{relatorio['apex_id']}_saved_reports_go"


def pasta_download_relatorio(relatorio):
    pasta = DOWNLOAD_ROOT / limpar_nome_arquivo(relatorio["nome"])
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta


def caminho_unico(caminho):
    if not caminho.exists():
        return caminho

    base = caminho.stem
    ext = caminho.suffix
    pasta = caminho.parent

    contador = 2
    while True:
        novo = pasta / f"{base}_{contador}{ext}"
        if not novo.exists():
            return novo
        contador += 1


def resolver_template_nao_iniciado(relatorio):
    """
    Retorna um arquivo modelo para criar o placeholder de cenário sem Edit.

    Prioridade:
    1) TEMPLATE_NAO_INICIADO informado no .env
    2) Primeiro download real já salvo para o relatório atual
    3) Nenhum modelo encontrado
    """
    if TEMPLATE_NAO_INICIADO:
        template = Path(TEMPLATE_NAO_INICIADO)
        if not template.is_absolute():
            template = BASE_DIR / template

        if template.exists():
            return template

        print(f"[AVISO] TEMPLATE_NAO_INICIADO informado, mas não encontrado: {template}")

    return MODELOS_DOWNLOAD_POR_RELATORIO.get(relatorio["nome"])


def registrar_modelo_download(relatorio, caminho_download):
    """Guarda o primeiro download real do relatório como template para placeholders."""
    if relatorio["nome"] not in MODELOS_DOWNLOAD_POR_RELATORIO:
        MODELOS_DOWNLOAD_POR_RELATORIO[relatorio["nome"]] = Path(caminho_download)
        print(f"[TEMPLATE] Modelo de colunas definido para {relatorio['nome']}: {caminho_download}")


def criar_placeholder_csv(template_path, destino):
    """Cria CSV vazio preservando o cabeçalho do template, quando possível."""
    header = ""

    if template_path and Path(template_path).exists():
        with open(template_path, "r", encoding="utf-8-sig", errors="ignore") as origem:
            header = origem.readline().rstrip("\n")

    with open(destino, "w", encoding="utf-8-sig", newline="") as saida:
        if header:
            saida.write(header + "\n")


def criar_placeholder_xlsx(template_path, destino):
    """
    Cria XLSX vazio usando o primeiro download real como modelo.
    Mantém cabeçalho/estrutura e apaga linhas de dados a partir da linha 2.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise Exception(
            "Para criar placeholder .xlsx preservando colunas, instale openpyxl: pip install openpyxl"
        ) from e

    wb = load_workbook(template_path)

    for ws in wb.worksheets:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    wb.save(destino)


def criar_download_nao_iniciado(page, origem, numero_pagina, relatorio):
    """
    Cria arquivo placeholder quando o cenário existe, mas não possui link Edit.

    O arquivo usa o mesmo cabeçalho/colunas de um download real quando houver template.
    Se ainda não existir template, gera um CSV vazio para registrar o cenário não iniciado.
    """
    if not GERAR_PLACEHOLDER_SEM_EDIT:
        print("[PLACEHOLDER] Geração de placeholder desativada no .env.")
        return None

    pasta_relatorio = pasta_download_relatorio(relatorio)
    template = resolver_template_nao_iniciado(relatorio)

    responsavel_download = prefixo_responsavel_download(relatorio)

    if responsavel_download:
        nome_base = (
            f"pagina_{numero_pagina}_{responsavel_download}_"
            f"{limpar_nome_arquivo(origem)}_Cenario_e_Execucao_nao_iniciado"
        )
    else:
        nome_base = f"pagina_{numero_pagina}_{limpar_nome_arquivo(origem)}_Cenario_e_Execucao_nao_iniciado"

    if template and Path(template).exists():
        extensao = Path(template).suffix.lower() or ".csv"
    else:
        extensao = ".csv"

    destino = caminho_unico(pasta_relatorio / f"{nome_base}{extensao}")

    try:
        if template and Path(template).exists() and extensao in {".xlsx", ".xlsm"}:
            criar_placeholder_xlsx(template, destino)
        elif template and Path(template).exists() and extensao in {".csv", ".txt"}:
            criar_placeholder_csv(template, destino)
        else:
            # Sem template ainda: cria um marcador CSV vazio.
            # Assim o processo não perde o registro do cenário sem execução iniciada.
            with open(destino, "w", encoding="utf-8-sig", newline="") as saida:
                saida.write("")

        print(f"[PLACEHOLDER] Arquivo de cenário não iniciado criado: {destino}")
        registrar_arquivo_para_consolidacao(relatorio, destino, "placeholder_sem_edit")
        return destino

    except Exception as e:
        print("[ERRO] Não consegui criar o arquivo placeholder de cenário não iniciado.")
        print(f"[DETALHE] {e}")
        return None


def validar_env():
    if not GTN_USUARIO:
        raise Exception("Não encontrei GTN_USUARIO ou GTN_USER no arquivo .env.")

    if not GTN_SENHA:
        raise Exception("Não encontrei GTN_SENHA ou GTN_PASS no arquivo .env.")

    if not URL_LOGIN:
        raise Exception("URL_LOGIN não pode ficar vazia no .env.")

    if not RELATORIOS:
        raise Exception("Nenhum relatório carregado do .env.")


def aguardar_processamento_apex(page):
    seletores = [
        ".u-Processing",
        ".a-Processing",
        ".apex_wait_overlay",
        ".ui-widget-overlay",
    ]

    for seletor in seletores:
        try:
            page.locator(seletor).first.wait_for(state="hidden", timeout=8000)
        except Exception:
            pass


def aguardar_carregamento(page):
    try:
        page.wait_for_load_state("domcontentloaded", timeout=20000)
    except Exception:
        pass

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass

    aguardar_processamento_apex(page)


def aguardar_relatorio_pronto(page):
    aguardar_processamento_apex(page)

    try:
        page.get_by_role("cell", name="Editar").first.wait_for(state="visible", timeout=30000)
        return True
    except Exception:
        pass

    try:
        page.locator("tr > .a-IRR-linkCol").first.wait_for(state="visible", timeout=30000)
        return True
    except Exception:
        pass

    return False


def listar_opcoes_relatorio(select_relatorio):
    try:
        return select_relatorio.evaluate(
            """
            el => Array.from(el.options).map(opt => ({
                value: opt.value,
                text: opt.textContent.trim(),
                selected: opt.selected
            }))
            """
        )
    except Exception as e:
        return [{"erro": str(e)}]


def obter_frame(page, titulo):
    iframe = page.locator(f'iframe[title="{titulo}"]').first
    iframe.wait_for(state="attached", timeout=TIMEOUT_PADRAO)

    try:
        iframe.wait_for(state="visible", timeout=TIMEOUT_PADRAO)
    except Exception:
        pass

    return iframe.content_frame


# ============================================================
# CONSOLIDAÇÃO FINAL DOS DOWNLOADS
# ============================================================
def linha_vazia(linha):
    return not linha or all(valor is None or str(valor).strip() == "" for valor in linha)


def normalizar_nome_coluna(valor, indice):
    nome = "" if valor is None else str(valor).strip()
    nome = re.sub(r"\s+", " ", nome)
    return nome or f"COLUNA_{indice}"


def normalizar_cabecalho(linha):
    cabecalho = []
    usados = {}

    for indice, valor in enumerate(linha or [], start=1):
        nome_base = normalizar_nome_coluna(valor, indice)
        nome = nome_base
        contador = 2

        while nome.upper() in usados:
            nome = f"{nome_base}_{contador}"
            contador += 1

        usados[nome.upper()] = True
        cabecalho.append(nome)

    return cabecalho


def detectar_encoding_texto(caminho):
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            with open(caminho, "r", encoding=encoding) as arquivo:
                arquivo.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin1"


def iterar_linhas_csv(caminho):
    encoding = detectar_encoding_texto(caminho)

    with open(caminho, "r", encoding=encoding, newline="", errors="ignore") as arquivo:
        amostra = arquivo.read(8192)
        arquivo.seek(0)

        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";" if amostra.count(";") >= amostra.count(",") else ","

        for linha in csv.reader(arquivo, dialect):
            yield list(linha)


def iterar_linhas_xlsx(caminho):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise Exception("Para consolidar .xlsx, instale openpyxl: pip install openpyxl") from e

    wb = load_workbook(caminho, read_only=True, data_only=True)

    try:
        ws = wb.worksheets[0]
        for linha in ws.iter_rows(values_only=True):
            yield list(linha)
    finally:
        wb.close()


def iterar_linhas_arquivo(caminho):
    caminho = Path(caminho)
    extensao = caminho.suffix.lower()

    if extensao in {".xlsx", ".xlsm"}:
        yield from iterar_linhas_xlsx(caminho)
        return

    if extensao in {".csv", ".txt"}:
        yield from iterar_linhas_csv(caminho)
        return

    # Alguns downloads APEX podem vir como .xls, mas na prática serem texto/HTML.
    # Primeiro tenta como planilha moderna; se não der, tenta como texto delimitado.
    if extensao == ".xls":
        try:
            yield from iterar_linhas_xlsx(caminho)
            return
        except Exception:
            yield from iterar_linhas_csv(caminho)
            return

    raise Exception(f"Extensão não suportada para consolidação: {extensao}")


def obter_cabecalho_arquivo(caminho):
    for linha in iterar_linhas_arquivo(caminho):
        if not linha_vazia(linha):
            return normalizar_cabecalho(linha)

    return []


def construir_cabecalhos_consolidados(arquivos_por_aba):
    cabecalhos_por_aba = {}

    for aba, arquivos in arquivos_por_aba.items():
        uniao = []
        usados = set()

        for info in arquivos:
            caminho = info["caminho"]

            try:
                cabecalho = obter_cabecalho_arquivo(caminho)
            except Exception as e:
                print(f"[CONSOLIDADO][AVISO] Não consegui ler cabeçalho de {caminho.name}: {e}")
                cabecalho = []

            info["cabecalho"] = cabecalho

            for nome_coluna in cabecalho:
                chave = nome_coluna.upper()
                if chave not in usados:
                    usados.add(chave)
                    uniao.append(nome_coluna)

        cabecalhos_por_aba[aba] = [
            "RESPONSAVEL",
            "RELATORIO_ORIGEM",
            "SAVED_VALUE",
            "ARQUIVO_ORIGEM",
            "TIPO_ORIGEM",
            "STATUS_CONSOLIDACAO",
            "LINHA_ORIGEM",
        ] + uniao

    return cabecalhos_por_aba


def escrever_cabecalho(ws, cabecalho):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fonte = Font(bold=True, color="FFFFFF")
    preenchimento = PatternFill(fill_type="solid", fgColor="1F4E78")
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

    celulas = []
    for valor in cabecalho:
        celula = WriteOnlyCell(ws, value=valor)
        celula.font = fonte
        celula.fill = preenchimento
        celula.alignment = alinhamento
        celulas.append(celula)

    ws.append(celulas)
    ws.freeze_panes = "A2"

    for indice, nome_coluna in enumerate(cabecalho, start=1):
        largura = min(max(len(str(nome_coluna)) + 2, 12), 45)
        ws.column_dimensions[get_column_letter(indice)].width = largura


def criar_celula_resumo(ws, valor, negrito=False, preenchimento=None):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill

    celula = WriteOnlyCell(ws, value=valor)
    celula.alignment = Alignment(vertical="center", wrap_text=True)

    if negrito:
        celula.font = Font(bold=True, color="FFFFFF" if preenchimento else "000000")

    if preenchimento:
        celula.fill = PatternFill(fill_type="solid", fgColor=preenchimento)

    return celula


def escrever_linha_resumo(ws, valores, cabecalho=False):
    preenchimento = "1F4E78" if cabecalho else None
    ws.append([criar_celula_resumo(ws, valor, negrito=cabecalho, preenchimento=preenchimento) for valor in valores])


def contar_arquivos_por_aba(arquivos_por_aba, aba, tipo=None):
    arquivos = arquivos_por_aba.get(aba, [])
    if tipo is None:
        return len(arquivos)
    return sum(1 for info in arquivos if info.get("tipo") == tipo)


def escrever_aba_resumo_execucao(
    wb_saida,
    destino,
    inicio_execucao,
    fim_execucao,
    resumos_execucao,
    arquivos_por_aba,
    totais_por_aba,
    estatisticas_arquivos,
):
    from openpyxl.utils import get_column_letter

    ws = wb_saida.create_sheet(title="RESUMO_EXECUCAO")

    if inicio_execucao is None:
        inicio_execucao = fim_execucao

    duracao_segundos = max((fim_execucao - inicio_execucao).total_seconds(), 0)

    total_paginas = sum((resumo or {}).get("paginas", 0) for resumo in (resumos_execucao or []))
    total_downloads_ok = sum((resumo or {}).get("downloads_ok", 0) for resumo in (resumos_execucao or []))
    total_sem_edit = sum((resumo or {}).get("sem_edit_ou_erro", 0) for resumo in (resumos_execucao or []))
    total_arquivos = len(estatisticas_arquivos)
    total_linhas_consolidadas = sum(info.get("linhas_consolidadas", 0) for info in estatisticas_arquivos)
    total_downloads_arquivo = sum(1 for info in estatisticas_arquivos if info.get("tipo") == "download")
    total_placeholders = sum(1 for info in estatisticas_arquivos if info.get("tipo") == "placeholder_sem_edit")
    total_erros_consolidacao = sum(1 for info in estatisticas_arquivos if str(info.get("status", "")).startswith("ERRO"))

    escrever_linha_resumo(ws, ["RESUMO DA EXECUÇÃO GTN"], cabecalho=True)
    escrever_linha_resumo(ws, [])

    escrever_linha_resumo(ws, ["Indicador", "Valor"], cabecalho=True)
    escrever_linha_resumo(ws, ["Início da execução", inicio_execucao.strftime("%d/%m/%Y %H:%M:%S")])
    escrever_linha_resumo(ws, ["Fim da execução / geração do consolidado", fim_execucao.strftime("%d/%m/%Y %H:%M:%S")])
    escrever_linha_resumo(ws, ["Tempo total", formatar_duracao(duracao_segundos)])
    escrever_linha_resumo(ws, ["Tempo total em segundos", round(duracao_segundos, 2)])
    escrever_linha_resumo(ws, ["Arquivo consolidado", str(destino)])
    escrever_linha_resumo(ws, ["Relatórios configurados", len(RELATORIOS)])
    escrever_linha_resumo(ws, ["Relatórios processados", len(resumos_execucao or [])])
    escrever_linha_resumo(ws, ["Páginas processadas", total_paginas])
    escrever_linha_resumo(ws, ["Downloads OK no fluxo", total_downloads_ok])
    escrever_linha_resumo(ws, ["Sem Edit / erro / pulados no fluxo", total_sem_edit])
    escrever_linha_resumo(ws, ["Arquivos gerados e considerados no consolidado", total_arquivos])
    escrever_linha_resumo(ws, ["Arquivos de download real", total_downloads_arquivo])
    escrever_linha_resumo(ws, ["Arquivos placeholder sem Edit", total_placeholders])
    escrever_linha_resumo(ws, ["Linhas gravadas nas abas consolidadas", total_linhas_consolidadas])
    escrever_linha_resumo(ws, ["Erros durante a consolidação", total_erros_consolidacao])

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Resumo por aba/responsável"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        ["ABA", "ARQUIVOS", "DOWNLOADS", "PLACEHOLDERS", "LINHAS_CONSOLIDADAS", "RELATORIOS_ENVOLVIDOS"],
        cabecalho=True,
    )

    for aba in ["WALCERI", "CAMILA", "LUCASRAMOS"]:
        relatorios_aba = sorted({info.get("relatorio", "") for info in arquivos_por_aba.get(aba, []) if info.get("relatorio")})
        escrever_linha_resumo(
            ws,
            [
                aba,
                contar_arquivos_por_aba(arquivos_por_aba, aba),
                contar_arquivos_por_aba(arquivos_por_aba, aba, "download"),
                contar_arquivos_por_aba(arquivos_por_aba, aba, "placeholder_sem_edit"),
                totais_por_aba.get(aba, 0),
                ", ".join(relatorios_aba),
            ],
        )

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Resumo por relatório"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        ["RELATORIO", "RESPONSAVEL", "STATUS", "PAGINAS", "DOWNLOADS_OK", "SEM_EDIT_OU_ERRO", "ARQUIVOS_GERADOS", "ERRO"],
        cabecalho=True,
    )

    arquivos_por_relatorio = {}
    responsavel_por_relatorio = {}
    for info in estatisticas_arquivos:
        relatorio = info.get("relatorio", "")
        arquivos_por_relatorio[relatorio] = arquivos_por_relatorio.get(relatorio, 0) + 1
        responsavel_por_relatorio[relatorio] = info.get("responsavel", "")

    for resumo in resumos_execucao or []:
        relatorio = resumo.get("relatorio", "")
        escrever_linha_resumo(
            ws,
            [
                relatorio,
                resumo.get("responsavel", "") or responsavel_por_relatorio.get(relatorio, ""),
                resumo.get("status", "OK"),
                resumo.get("paginas", 0),
                resumo.get("downloads_ok", 0),
                resumo.get("sem_edit_ou_erro", 0),
                arquivos_por_relatorio.get(relatorio, 0),
                resumo.get("erro", ""),
            ],
        )

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Detalhe dos arquivos consolidados"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        [
            "ABA",
            "RESPONSAVEL",
            "RELATORIO",
            "SAVED_VALUE",
            "ARQUIVO",
            "TIPO",
            "STATUS",
            "LINHAS_CONSOLIDADAS",
            "LINHAS_DADOS_REAIS",
        ],
        cabecalho=True,
    )

    for info in estatisticas_arquivos:
        escrever_linha_resumo(
            ws,
            [
                info.get("aba", ""),
                info.get("responsavel", ""),
                info.get("relatorio", ""),
                info.get("saved_value", ""),
                info.get("arquivo", ""),
                info.get("tipo", ""),
                info.get("status", ""),
                info.get("linhas_consolidadas", 0),
                info.get("linhas_dados_reais", 0),
            ],
        )

    ws.freeze_panes = "A4"
    larguras = {
        1: 24,
        2: 28,
        3: 22,
        4: 18,
        5: 58,
        6: 22,
        7: 24,
        8: 20,
        9: 20,
        10: 80,
    }

    for indice, largura in larguras.items():
        ws.column_dimensions[get_column_letter(indice)].width = largura


def consolidar_arquivos_download_execucao(resumos_execucao=None, inicio_execucao=None):
    """
    Consolida os arquivos gerados na execução atual em um XLSX único.

    O arquivo final fica na pasta raiz de downloads com nome:
    consolidado_HH-MM-SS.xlsx

    Os dois-pontos foram substituídos por hífen para compatibilidade com Windows.
    """
    if not GERAR_CONSOLIDADO_FINAL:
        print("[CONSOLIDADO] Consolidação final desativada no .env.")
        return None

    if not ARQUIVOS_DOWNLOAD_EXECUCAO:
        print("[CONSOLIDADO] Nenhum arquivo da execução atual para consolidar. Mesmo assim vou gerar o consolidado com RESUMO_EXECUCAO.")

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise Exception("Para gerar o consolidado final, instale openpyxl: pip install openpyxl") from e

    DOWNLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    nome_consolidado = datetime.now().strftime("consolidado_%H-%M-%S.xlsx")
    destino = caminho_unico(DOWNLOAD_ROOT / nome_consolidado)

    abas_obrigatorias = ["WALCERI", "CAMILA", "LUCASRAMOS"]
    arquivos_por_aba = {aba: [] for aba in abas_obrigatorias}

    for info in ARQUIVOS_DOWNLOAD_EXECUCAO:
        aba = info.get("aba")
        if aba in arquivos_por_aba:
            arquivos_por_aba[aba].append(info)

    cabecalhos_por_aba = construir_cabecalhos_consolidados(arquivos_por_aba)

    wb_saida = Workbook(write_only=True)
    totais_por_aba = {}
    estatisticas_arquivos = []

    for aba in abas_obrigatorias:
        ws = wb_saida.create_sheet(title=aba)
        cabecalho_saida = cabecalhos_por_aba[aba]
        escrever_cabecalho(ws, cabecalho_saida)

        mapa_colunas_saida = {nome.upper(): idx for idx, nome in enumerate(cabecalho_saida)}
        total_linhas_dados = 0

        for info in arquivos_por_aba[aba]:
            caminho = info["caminho"]
            cabecalho_arquivo = info.get("cabecalho") or []
            estatistica = {
                "aba": aba,
                "responsavel": info.get("responsavel", ""),
                "relatorio": info.get("relatorio", ""),
                "saved_value": info.get("saved_value", ""),
                "arquivo": caminho.name,
                "tipo": info.get("tipo", ""),
                "status": "OK",
                "linhas_consolidadas": 0,
                "linhas_dados_reais": 0,
            }

            if not cabecalho_arquivo:
                ws.append([
                    info.get("responsavel", ""),
                    info.get("relatorio", ""),
                    info.get("saved_value", ""),
                    caminho.name,
                    info.get("tipo", ""),
                    "SEM_DADOS",
                    "",
                ] + [None] * (len(cabecalho_saida) - 7))
                total_linhas_dados += 1
                estatistica["status"] = "SEM_DADOS"
                estatistica["linhas_consolidadas"] = 1
                estatisticas_arquivos.append(estatistica)
                continue

            primeira_linha_util = True
            linhas_dados_arquivo = 0
            linha_origem = 0

            try:
                for linha in iterar_linhas_arquivo(caminho):
                    if linha_vazia(linha):
                        continue

                    linha_origem += 1

                    if primeira_linha_util:
                        primeira_linha_util = False
                        # Primeira linha útil é tratada como cabeçalho.
                        continue

                    saida = [None] * len(cabecalho_saida)
                    saida[0] = info.get("responsavel", "")
                    saida[1] = info.get("relatorio", "")
                    saida[2] = info.get("saved_value", "")
                    saida[3] = caminho.name
                    saida[4] = info.get("tipo", "")
                    saida[5] = "OK"
                    saida[6] = linha_origem

                    for indice_coluna, valor in enumerate(linha):
                        if indice_coluna >= len(cabecalho_arquivo):
                            continue

                        nome_coluna = cabecalho_arquivo[indice_coluna]
                        indice_saida = mapa_colunas_saida.get(nome_coluna.upper())

                        if indice_saida is not None:
                            saida[indice_saida] = valor

                    ws.append(saida)
                    total_linhas_dados += 1
                    linhas_dados_arquivo += 1

                if linhas_dados_arquivo == 0:
                    ws.append([
                        info.get("responsavel", ""),
                        info.get("relatorio", ""),
                        info.get("saved_value", ""),
                        caminho.name,
                        info.get("tipo", ""),
                        "SEM_DADOS",
                        "",
                    ] + [None] * (len(cabecalho_saida) - 7))
                    total_linhas_dados += 1
                    estatistica["status"] = "SEM_DADOS"
                    estatistica["linhas_consolidadas"] = 1
                    estatistica["linhas_dados_reais"] = 0
                else:
                    estatistica["status"] = "OK"
                    estatistica["linhas_consolidadas"] = linhas_dados_arquivo
                    estatistica["linhas_dados_reais"] = linhas_dados_arquivo

            except Exception as e:
                print(f"[CONSOLIDADO][ERRO] Falha ao consolidar {caminho.name}: {e}")
                ws.append([
                    info.get("responsavel", ""),
                    info.get("relatorio", ""),
                    info.get("saved_value", ""),
                    caminho.name,
                    info.get("tipo", ""),
                    f"ERRO: {e}",
                    "",
                ] + [None] * (len(cabecalho_saida) - 7))
                total_linhas_dados += 1
                estatistica["status"] = f"ERRO: {e}"
                estatistica["linhas_consolidadas"] = 1
                estatistica["linhas_dados_reais"] = 0

            estatisticas_arquivos.append(estatistica)

        if total_linhas_dados > 0:
            ultima_coluna = get_column_letter(len(cabecalho_saida))
            ws.auto_filter.ref = f"A1:{ultima_coluna}{total_linhas_dados + 1}"

        totais_por_aba[aba] = total_linhas_dados

    fim_execucao = datetime.now()
    escrever_aba_resumo_execucao(
        wb_saida=wb_saida,
        destino=destino,
        inicio_execucao=inicio_execucao,
        fim_execucao=fim_execucao,
        resumos_execucao=resumos_execucao,
        arquivos_por_aba=arquivos_por_aba,
        totais_por_aba=totais_por_aba,
        estatisticas_arquivos=estatisticas_arquivos,
    )

    wb_saida.save(destino)

    print("=" * 100)
    print(f"[CONSOLIDADO] Arquivo final criado: {destino}")
    print("[CONSOLIDADO] Abas geradas:")
    for aba in abas_obrigatorias:
        print(f"- {aba}: {totais_por_aba.get(aba, 0)} linha(s)")
    print(f"- RESUMO_EXECUCAO: tempo={formatar_duracao((fim_execucao - (inicio_execucao or fim_execucao)).total_seconds())}")
    print("=" * 100)

    return destino


# ============================================================
# LOGIN E NAVEGAÇÃO
# ============================================================
def login_e_abrir_execucao_testes(page):
    print("[1] Abrindo tela de login...")
    page.goto(URL_LOGIN, wait_until="domcontentloaded")
    page.wait_for_timeout(2000)

    print("[2] Preenchendo usuário...")
    campo_usuario = page.get_by_role("textbox", name="Usuário")
    campo_usuario.wait_for(state="visible")
    campo_usuario.fill(GTN_USUARIO)

    page.wait_for_timeout(800)

    print("[3] Indo para senha...")
    campo_usuario.press("Tab")
    page.wait_for_timeout(800)

    print("[4] Preenchendo senha...")
    campo_senha = page.get_by_role("textbox", name="Senha")
    campo_senha.wait_for(state="visible")
    campo_senha.fill(GTN_SENHA)

    page.wait_for_timeout(1000)

    print("[5] Clicando em Acessar...")
    botao_acessar = page.get_by_role("button", name="Acessar")
    botao_acessar.wait_for(state="visible")
    botao_acessar.click()

    print("[6] Aguardando pós-login/renderização...")
    page.wait_for_timeout(5000)
    aguardar_carregamento(page)
    page.wait_for_timeout(3000)

    print(f"[INFO] URL atual depois do login: {page.url}")

    print("[7] Aguardando botão Navegação Principal...")
    botao_nav = page.get_by_role("button", name="Navegação Principal")
    botao_nav.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1500)

    print("[8] Clicando em Navegação Principal...")
    botao_nav.click()
    page.wait_for_timeout(3000)

    print("[9] Aguardando toggle da árvore...")
    toggle_arvore = page.locator(".a-TreeView-toggle:visible").first
    toggle_arvore.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print("[10] Clicando no toggle da árvore...")
    toggle_arvore.click()
    page.wait_for_timeout(3000)

    print("[11] Aguardando item Execução de Testes...")
    item_execucao = page.get_by_role("treeitem", name="Execução de Testes")
    item_execucao.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print("[12] Clicando em Execução de Testes...")
    item_execucao.click()
    page.wait_for_timeout(7000)
    aguardar_carregamento(page)

    print("[OK] Clique em Execução de Testes executado.")
    print(f"[INFO] URL atual após clique em Execução de Testes: {page.url}")

    print("[13] Mantendo a página atual de Execução de Testes.")
    print("[INFO] Não vou usar page.goto com session/cs fixos, porque isso pode voltar para o login.")
    print(f"[INFO] URL atual válida: {page.url}")

    page.wait_for_timeout(7000)


# ============================================================
# SELEÇÃO DO RELATÓRIO
# ============================================================
def selecionar_relatorio_apex(page, relatorio):
    select_id = seletor_select_relatorio(relatorio)
    botao_ir_id = seletor_botao_ir_relatorio(relatorio)

    print("=" * 100)
    print(f"[RELATÓRIO] Selecionando: {relatorio['nome']}")
    print(f"[RELATÓRIO] APEX ID: {relatorio['apex_id']}")
    print(f"[RELATÓRIO] VALUE salvo: {relatorio['saved_value']}")
    print(f"[RELATÓRIO] Responsável no download: {relatorio.get('responsavel') or 'não definido'}")
    print("=" * 100)

    print(f"[14] Aguardando select do relatório APEX {relatorio['apex_id']}...")
    print(f"[INFO] Select esperado: {select_id}")

    select_relatorio = page.locator(select_id)
    select_relatorio.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print(f"[15] Selecionando relatório salvo pelo VALUE real: {relatorio['saved_value']}")

    try:
        select_relatorio.select_option(value=relatorio["saved_value"], timeout=TIMEOUT_PADRAO)
    except Exception:
        print("[ERRO] Não consegui selecionar o relatório salvo pelo VALUE informado.")
        print("[DIAGNÓSTICO] Opções encontradas no combo:")
        for opcao in listar_opcoes_relatorio(select_relatorio):
            print(opcao)
        raise

    page.wait_for_timeout(3000)

    print(f"[16] Aguardando botão Ir do relatório APEX {relatorio['apex_id']}...")
    print(f"[INFO] Botão Ir esperado: {botao_ir_id}")

    botao_ir = page.locator(botao_ir_id)

    try:
        botao_ir.wait_for(state="visible", timeout=15000)
    except Exception:
        print("[AVISO] Não encontrei botão Ir pelo ID exato. Tentando pelo botão Ir visível da tela.")
        botao_ir = page.get_by_role("button", name="Ir", exact=True)
        botao_ir.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1500)

    print(f"[17] Clicando no botão Ir do relatório APEX {relatorio['apex_id']}...")
    botao_ir.click()

    page.wait_for_timeout(7000)
    aguardar_carregamento(page)
    aguardar_relatorio_pronto(page)

    print("[OK] Relatório selecionado e botão Ir executado.")
    print(f"[INFO] URL atual: {page.url}")


# ============================================================
# MODAIS / JANELAS
# ============================================================
def fechar_modal_se_existir(page):
    tentativas = [
        lambda: page.get_by_role("dialog", name="Execução de Teste").get_by_label("Fechar"),
        lambda: page.get_by_role("dialog", name="Cenário e Execuções").get_by_label("Fechar"),
        lambda: page.get_by_role("button", name="Fechar"),
        lambda: page.locator("button[title='Fechar']"),
        lambda: page.locator("button[aria-label='Fechar']"),
        lambda: page.locator(".ui-dialog-titlebar-close"),
    ]

    for criar_locator in tentativas:
        try:
            locator = criar_locator().first
            if locator.count() > 0:
                locator.click(timeout=3000)
                page.wait_for_timeout(1000)
        except Exception:
            pass

    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(500)
    except Exception:
        pass

    aguardar_processamento_apex(page)


# ============================================================
# DOWNLOAD DE UMA EXECUÇÃO
# ============================================================
def localizar_links_edit(frame_cenario):
    """
    Localiza links Edit de forma mais tolerante.

    Motivo: em alguns cenários o modal "Cenário e Execuções" demora mais
    para renderizar o link Edit, ou o APEX expõe o link por seletor diferente.
    O script antigo esperava apenas 5 segundos e poderia registrar como SEM EDIT
    mesmo quando a execução existia.
    """
    tentativas = [
        ("role link name=Edit", lambda: frame_cenario.get_by_role("link", name="Edit")),
        ("role link regex Edit", lambda: frame_cenario.get_by_role("link", name=re.compile(r"^\s*Edit\s*$"))),
        ("anchor text Edit", lambda: frame_cenario.locator("a").filter(has_text=re.compile(r"^\s*Edit\s*$"))),
        ("anchor title/aria Edit", lambda: frame_cenario.locator("a[title*='Edit'], a[aria-label*='Edit']")),
    ]

    ultimo_erro = None

    for descricao, criar_locator in tentativas:
        try:
            locator = criar_locator()
            locator.first.wait_for(state="visible", timeout=TIMEOUT_EDIT_EXECUCAO)
            qtd = locator.count()

            if qtd > 0:
                return locator, qtd, descricao

        except Exception as e:
            ultimo_erro = e

    return None, 0, str(ultimo_erro) if ultimo_erro else "nenhum seletor encontrou Edit"


def baixar_execucao_aberta(page, origem, numero_pagina, relatorio):
    frame_cenario = obter_frame(page, "Cenário e Execuções")

    print(f"[VALIDAÇÃO] Verificando link Edit para {origem}...")
    links_edit, qtd_edit, estrategia_edit = localizar_links_edit(frame_cenario)

    if not links_edit or qtd_edit <= 0:
        print(f"[SEM EDIT] {origem} não possui link Edit visível após {TIMEOUT_EDIT_EXECUCAO}ms.")
        print(f"[DIAGNÓSTICO] Última estratégia/erro: {estrategia_edit}")
        criar_download_nao_iniciado(page, origem, numero_pagina, relatorio)
        fechar_modal_se_existir(page)
        return False

    print(
        f"[VALIDAÇÃO] {origem} possui {qtd_edit} link(s) Edit. "
        f"Estratégia: {estrategia_edit}. Usando sempre o primeiro visível: first."
    )

    print(f"[DOWNLOAD] Clicando no primeiro Edit para {origem}...")
    links_edit.first.click(timeout=TIMEOUT_PADRAO)
    page.wait_for_timeout(2500)
    aguardar_processamento_apex(page)

    frame_execucao = obter_frame(page, "Execução de Teste")

    print(f"[DOWNLOAD] Abrindo Ações para {origem}...")
    frame_execucao.get_by_role("button", name="Ações", exact=True).click(timeout=TIMEOUT_PADRAO)
    page.wait_for_timeout(1200)

    print(f"[DOWNLOAD] Clicando em Fazer Download no menu para {origem}...")
    try:
        frame_execucao.get_by_role("menuitem", name="Fazer Download").click(timeout=5000)
    except Exception:
        frame_execucao.locator("span").filter(has_text="Fazer Download").click(timeout=5000)

    page.wait_for_timeout(1200)

    print(f"[DOWNLOAD] Confirmando Fazer Download para {origem}...")
    with page.expect_download(timeout=TIMEOUT_DOWNLOAD) as download_info:
        frame_execucao.get_by_role("button", name="Fazer Download").click(timeout=TIMEOUT_PADRAO)

    download = download_info.value

    pasta_relatorio = pasta_download_relatorio(relatorio)
    nome_original = limpar_nome_arquivo(download.suggested_filename)
    responsavel_download = prefixo_responsavel_download(relatorio)

    if responsavel_download:
        nome_final = f"pagina_{numero_pagina}_{responsavel_download}_{limpar_nome_arquivo(origem)}_{nome_original}"
    else:
        nome_final = f"pagina_{numero_pagina}_{limpar_nome_arquivo(origem)}_{nome_original}"

    destino = caminho_unico(pasta_relatorio / nome_final)

    download.save_as(str(destino))
    print(f"[OK] Download salvo em: {destino}")

    registrar_modelo_download(relatorio, destino)
    registrar_arquivo_para_consolidacao(relatorio, destino, "download")

    fechar_modal_se_existir(page)
    page.wait_for_timeout(1500)
    return True

def baixar_execucao_por_indice(page, indice, numero_pagina, relatorio):
    origem = f"editar_indice_{indice}"

    try:
        print(f"[PÁGINA {numero_pagina}] Processando {origem}...")

        celulas_editar = page.get_by_role("cell", name="Editar")
        locator = celulas_editar.first if indice == 0 else celulas_editar.nth(indice)

        locator.wait_for(state="visible", timeout=TIMEOUT_LINHA_CURTO)
        locator.click(timeout=TIMEOUT_PADRAO)

        page.wait_for_timeout(2500)
        aguardar_processamento_apex(page)

        return baixar_execucao_aberta(page, origem, numero_pagina, relatorio)

    except Exception as e:
        print(f"[ERRO] Falha em {origem}. Fechando janelas e seguindo.")
        print(f"[DETALHE] {e}")
        fechar_modal_se_existir(page)
        page.wait_for_timeout(1000)
        return False


def baixar_execucao_por_tr_child(page, linha_child, numero_pagina, relatorio):
    origem = f"tr_child_{linha_child}"

    try:
        print(f"[PÁGINA {numero_pagina}] Processando tr:nth-child({linha_child})...")

        linha_link = page.locator(f"tr:nth-child({linha_child}) > .a-IRR-linkCol").first
        linha_link.wait_for(state="visible", timeout=TIMEOUT_LINHA_CURTO)

        linha_link.click(timeout=TIMEOUT_PADRAO)
        page.wait_for_timeout(2500)
        aguardar_processamento_apex(page)

        return baixar_execucao_aberta(page, origem, numero_pagina, relatorio)

    except PlaywrightTimeoutError:
        print(f"[SEM LINHA] tr:nth-child({linha_child}) não encontrado/visível.")
        return None

    except Exception as e:
        print(f"[ERRO] Falha no tr:nth-child({linha_child}). Fechando janelas e seguindo.")
        print(f"[DETALHE] {e}")
        fechar_modal_se_existir(page)
        page.wait_for_timeout(1000)
        return False


# ============================================================
# PROCESSAMENTO DE PÁGINA / PAGINAÇÃO
# ============================================================
def processar_pagina_atual(page, numero_pagina, relatorio):
    print("=" * 100)
    print(f"[RELATÓRIO: {relatorio['nome']}] [PÁGINA {numero_pagina}] INICIANDO PROCESSAMENTO")
    print("=" * 100)

    aguardar_relatorio_pronto(page)

    downloads_ok = 0
    sem_edit_ou_erro = 0
    linhas_inexistentes = 0

    editar_inicio = relatorio["editar_inicio"]
    editar_fim = relatorio["editar_fim"]
    tr_inicio = relatorio["tr_inicio"]
    tr_fim = relatorio["tr_fim"]

    if editar_inicio >= 0 and editar_fim >= editar_inicio:
        print(f"[PÁGINA {numero_pagina}] Processando Editar índice {editar_inicio} até {editar_fim}...")

        for indice in range(editar_inicio, editar_fim + 1):
            resultado = baixar_execucao_por_indice(page, indice, numero_pagina, relatorio)

            if resultado is True:
                downloads_ok += 1
            else:
                sem_edit_ou_erro += 1

        print(
            f"[PÁGINA {numero_pagina}] Bloco Editar finalizado. "
            f"Downloads OK: {downloads_ok}. Sem Edit/erro: {sem_edit_ou_erro}."
        )
    else:
        print(f"[PÁGINA {numero_pagina}] Bloco Editar ignorado por configuração do .env.")

    if tr_inicio >= 0 and tr_fim >= tr_inicio:
        print(f"[PÁGINA {numero_pagina}] Processando tr:nth-child({tr_inicio}) até tr:nth-child({tr_fim})...")

        for linha_child in range(tr_inicio, tr_fim + 1):
            resultado = baixar_execucao_por_tr_child(page, linha_child, numero_pagina, relatorio)

            if resultado is True:
                downloads_ok += 1
                linhas_inexistentes = 0
            elif resultado is None:
                linhas_inexistentes += 1

                if linhas_inexistentes >= MAX_LINHAS_INEXISTENTES_SEGUIDAS:
                    print(
                        f"[PÁGINA {numero_pagina}] {linhas_inexistentes} linhas seguidas não encontradas. "
                        "Encerrando varredura desta página."
                    )
                    break
            else:
                sem_edit_ou_erro += 1
                linhas_inexistentes = 0

        print(
            f"[PÁGINA {numero_pagina}] Bloco tr:nth-child finalizado. "
            f"Downloads OK: {downloads_ok}. Sem Edit/erro: {sem_edit_ou_erro}."
        )
    else:
        print(f"[PÁGINA {numero_pagina}] Bloco tr:nth-child ignorado por configuração do .env.")

    return {
        "downloads_ok": downloads_ok,
        "sem_edit_ou_erro": sem_edit_ou_erro,
    }


def clicar_proximo_se_existir(page, numero_pagina):
    print(f"[PÁGINA {numero_pagina}] Procurando botão Próximo...")

    candidatos = [
        lambda: page.get_by_role("button", name="Próximo").first,
        lambda: page.get_by_role("link", name="Próximo").first,
        lambda: page.locator("a[aria-label='Próximo']").first,
        lambda: page.locator("button[aria-label='Próximo']").first,
    ]

    for criar_locator in candidatos:
        try:
            botao_proximo = criar_locator()
            botao_proximo.wait_for(state="visible", timeout=5000)

            try:
                if not botao_proximo.is_enabled(timeout=2000):
                    print(f"[PÁGINA {numero_pagina}] Próximo encontrado, mas desabilitado. Fim.")
                    return False
            except Exception:
                pass

            print(f"[PÁGINA {numero_pagina}] Clicando em Próximo...")
            botao_proximo.click(timeout=TIMEOUT_PADRAO)

            print(f"[PÁGINA {numero_pagina}] Aguardando próxima página carregar...")
            page.wait_for_timeout(8000)
            aguardar_carregamento(page)
            aguardar_relatorio_pronto(page)
            page.wait_for_timeout(3000)

            print(f"[PÁGINA {numero_pagina}] Próxima página carregada. URL atual: {page.url}")
            return True

        except Exception:
            continue

    print(f"[PÁGINA {numero_pagina}] Não encontrei botão Próximo disponível. Processo encerrado para este relatório.")
    return False


def processar_todas_as_paginas(page, relatorio):
    print("=" * 100)
    print(f"[INÍCIO] Processamento do relatório: {relatorio['nome']}")
    print("=" * 100)

    total_paginas = 0
    total_downloads_ok = 0
    total_sem_edit_ou_erro = 0

    numero_pagina = 1

    while True:
        resumo = processar_pagina_atual(page, numero_pagina, relatorio)

        total_paginas += 1
        total_downloads_ok += resumo["downloads_ok"]
        total_sem_edit_ou_erro += resumo["sem_edit_ou_erro"]

        conseguiu_ir_proxima = clicar_proximo_se_existir(page, numero_pagina)

        if not conseguiu_ir_proxima:
            print(f"[FIM] Relatório {relatorio['nome']} finalizado.")
            break

        numero_pagina += 1

    print("=" * 100)
    print(f"[RESUMO DO RELATÓRIO] {relatorio['nome']}")
    print(f"Total de páginas processadas: {total_paginas}")
    print(f"Total de downloads OK: {total_downloads_ok}")
    print(f"Total sem Edit/erro/pulados: {total_sem_edit_ou_erro}")
    print(f"Pasta de downloads: {pasta_download_relatorio(relatorio)}")
    print("=" * 100)

    return {
        "relatorio": relatorio["nome"],
        "paginas": total_paginas,
        "downloads_ok": total_downloads_ok,
        "sem_edit_ou_erro": total_sem_edit_ou_erro,
    }


# ============================================================
# FLUXO PRINCIPAL
# ============================================================
def run():
    validar_env()

    inicio_execucao = datetime.now()
    cronometro = CronometroExecucao(inicio_execucao)
    cronometro.iniciar()

    print("=" * 100)
    print(f"[TEMPO] Início da execução: {inicio_execucao.strftime('%d/%m/%Y %H:%M:%S')}")
    print("[CONFIGURAÇÃO] Relatórios carregados do .env")
    for idx, relatorio in enumerate(RELATORIOS, start=1):
        print(
            f"{idx}. {relatorio['nome']} | "
            f"APEX={relatorio['apex_id']} | "
            f"VALUE={relatorio['saved_value']} | "
            f"Responsável={relatorio.get('responsavel') or 'não definido'} | "
            f"Editar={relatorio['editar_inicio']}..{relatorio['editar_fim']} | "
            f"TR={relatorio['tr_inicio']}..{relatorio['tr_fim']}"
        )
    print("=" * 100)

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=False,
            slow_mo=300,
        )

        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(TIMEOUT_PADRAO)

        resumos = []
        consolidado_final = None

        try:
            cronometro.atualizar_status("Login e abertura da tela de Execução de Testes")
            login_e_abrir_execucao_testes(page)

            for indice, relatorio in enumerate(RELATORIOS, start=1):
                cronometro.atualizar_status(
                    f"Processando {relatorio['nome']} ({indice}/{len(RELATORIOS)})"
                )

                print("=" * 100)
                print(f"[LOOP] Iniciando relatório {indice}/{len(RELATORIOS)}: {relatorio['nome']}")
                print("=" * 100)

                try:
                    selecionar_relatorio_apex(page, relatorio)
                    resumo = processar_todas_as_paginas(page, relatorio)
                    resumo["status"] = "OK"
                    resumo["erro"] = ""
                    resumo["responsavel"] = relatorio.get("responsavel", "")
                    resumos.append(resumo)

                    print(f"[LOOP] Relatório finalizado: {relatorio['nome']}")

                except Exception as e:
                    erro = str(e)
                    print("=" * 100)
                    print(f"[RELATÓRIO][ERRO] Falha no relatório {relatorio['nome']}. Vou registrar no resumo e continuar para o próximo.")
                    print(f"[DETALHE] {erro}")
                    print("=" * 100)

                    resumos.append(
                        {
                            "relatorio": relatorio.get("nome", ""),
                            "responsavel": relatorio.get("responsavel", ""),
                            "paginas": 0,
                            "downloads_ok": 0,
                            "sem_edit_ou_erro": 0,
                            "status": "ERRO",
                            "erro": erro,
                        }
                    )

                    try:
                        fechar_modal_se_existir(page)
                    except Exception:
                        pass

                page.wait_for_timeout(3000)
                aguardar_carregamento(page)

            print("=" * 100)
            print("[RESUMO GERAL]")
            for resumo in resumos:
                print(
                    f"- {resumo['relatorio']}: "
                    f"status={resumo.get('status', 'OK')}, "
                    f"páginas={resumo.get('paginas', 0)}, "
                    f"downloads_ok={resumo.get('downloads_ok', 0)}, "
                    f"sem_edit_ou_erro={resumo.get('sem_edit_ou_erro', 0)}"
                )
                if resumo.get("erro"):
                    print(f"  erro={resumo.get('erro')}")
            print(f"Pasta raiz dos downloads: {DOWNLOAD_ROOT}")
            print("=" * 100)

            try:
                cronometro.atualizar_status("Gerando consolidado final e aba RESUMO_EXECUCAO")
                consolidado_final = consolidar_arquivos_download_execucao(
                    resumos_execucao=resumos,
                    inicio_execucao=inicio_execucao,
                )
            except Exception as e:
                print("[CONSOLIDADO][ERRO] Não consegui gerar o consolidado final.")
                print(f"[DETALHE] {e}")

            fim_execucao = datetime.now()
            print("=" * 100)
            print(f"[TEMPO] Fim da execução: {fim_execucao.strftime('%d/%m/%Y %H:%M:%S')}")
            print(f"[TEMPO] Duração total: {formatar_duracao((fim_execucao - inicio_execucao).total_seconds())}")
            if consolidado_final:
                print(f"[TEMPO] Resumo gravado no consolidado: {consolidado_final}")
            print("=" * 100)

            cronometro.atualizar_status("Finalizado. Consolidado gerado.")
            cronometro.finalizar()

            print("[INFO] O navegador ficará aberto para validação visual.")
            input("Pressione ENTER aqui no terminal para fechar o navegador...")

        finally:
            cronometro.finalizar()
            context.close()
            browser.close()


if __name__ == "__main__":
    try:
        run()
    except PlaywrightTimeoutError as e:
        print("[ERRO] Timeout no Playwright.")
        print(e)
        raise
    except Exception as e:
        print("[ERRO] Falha geral.")
        print(e)
        raise
