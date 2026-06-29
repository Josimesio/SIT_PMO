import csv
import os
import re
import threading
import time
import unicodedata
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

VERSAO_AJUSTE_PASTA_LIDER_CENARIO = "2026-06-27"
VERSAO_AJUSTE_ANTI_TRAVAMENTO_EDIT = "2026-06-27_TEMPO_EQUILIBRADO"
VERSAO_AJUSTE_METRICAS_TEMPO = "2026-06-27"


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")


def env_int(nome, padrao):
    """Lê inteiro do .env aceitando valor em branco como padrão.

    Evita quebra quando o .env tem algo como TIMEOUT_EDIT_EXECUCAO=
    ou quando existe variável duplicada e a última ficou vazia.
    """
    valor = os.getenv(nome)

    if valor is None or str(valor).strip() == "":
        return int(padrao)

    try:
        return int(str(valor).strip())
    except ValueError as e:
        raise Exception(
            f"Valor inválido no .env para {nome}: {valor!r}. Informe um número inteiro, exemplo: {nome}={padrao}"
        ) from e


GTN_USUARIO = os.getenv("GTN_USUARIO") or os.getenv("GTN_USER")
GTN_SENHA = os.getenv("GTN_SENHA") or os.getenv("GTN_PASS")

URL_LOGIN = os.getenv(
    "URL_LOGIN",
    "https://gtn.ninecon.com.br/ords/r/gtn/gtn/login?session=601689329474445&tz=-3:00",
)

TIMEOUT_PADRAO = env_int("TIMEOUT_PADRAO", "60000")
TIMEOUT_DOWNLOAD = env_int("TIMEOUT_DOWNLOAD", "180000")
TIMEOUT_LINHA_CURTO = env_int("TIMEOUT_LINHA_CURTO", "5000")
TIMEOUT_EDIT_EXECUCAO = env_int("TIMEOUT_EDIT_EXECUCAO", "15000")
# Limite real de espera por Edit. Evita que cenário sem execução deixe o robô parado tempo demais,
# mas dá tempo suficiente para o modal subir em telas lentas.
TIMEOUT_EDIT_EXECUCAO_MAXIMO_REAL = env_int("TIMEOUT_EDIT_EXECUCAO_MAXIMO_REAL", "25000")
INTERVALO_LOG_EDIT_SEGUNDOS = env_int("INTERVALO_LOG_EDIT_SEGUNDOS", "3")
MAX_LINHAS_INEXISTENTES_SEGUIDAS = env_int("MAX_LINHAS_INEXISTENTES_SEGUIDAS", "10")

DOWNLOAD_ROOT = BASE_DIR / os.getenv("DOWNLOAD_DIR", "downloads_gtn")

GERAR_PLACEHOLDER_SEM_EDIT = os.getenv("GERAR_PLACEHOLDER_SEM_EDIT", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}
TEMPLATE_NAO_INICIADO = os.getenv("TEMPLATE_NAO_INICIADO", "").strip()

GERAR_CONSOLIDADO_FINAL = os.getenv("GERAR_CONSOLIDADO_FINAL", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}
EXIBIR_TELA_TEMPO = os.getenv("EXIBIR_TELA_TEMPO", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}
MANTER_NAVEGADOR_ABERTO = os.getenv("MANTER_NAVEGADOR_ABERTO", "N").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}

# Arquivos gerados na execução atual.
# Usado para consolidar somente o que acabou de ser baixado/criado, sem misturar execuções antigas.
ARQUIVOS_DOWNLOAD_EXECUCAO = []

# Métricas de performance da execução atual.
# A ideia é descobrir onde o robô perde tempo: login, APEX, página, Edit, download ou placeholder.
METRICAS_EXECUCAO = []

# Guarda o primeiro arquivo baixado com sucesso por relatório para usar como modelo
# quando algum cenário não tiver link Edit.
MODELOS_DOWNLOAD_POR_RELATORIO = {}


# Líder de cenário/responsável que será usado no nome dos downloads e na pasta do relatório.
# Regra oficial solicitada:
# RELATORIO_01 a RELATORIO_04 = Walceir
# RELATORIO_05 a RELATORIO_08 = LucasRamos
# RELATORIO_09 a RELATORIO_12 = Camila
RESPONSAVEL_POR_RELATORIO_NOME = {
    "RELATORIO_01": "Walceir",
    "RELATORIO_02": "Walceir",
    "RELATORIO_03": "Walceir",
    "RELATORIO_04": "Walceir",
    "RELATORIO_05": "LucasRamos",
    "RELATORIO_06": "LucasRamos",
    "RELATORIO_07": "LucasRamos",
    "RELATORIO_08": "LucasRamos",
    "RELATORIO_09": "Camila",
    "RELATORIO_10": "Camila",
    "RELATORIO_11": "Camila",
    "RELATORIO_12": "Camila",
}

# Mantém a regra mesmo se o nome do relatório mudar, usando o VALUE real salvo do APEX.
RESPONSAVEL_POR_SAVED_VALUE = {
    "96443782140269930": "Walceir",
    "96445783233272755": "Walceir",
    "96447734078275977": "Walceir",
    "96461147945312616": "Walceir",
    "96466495992333195": "LucasRamos",
    "96468946665336064": "LucasRamos",
    "96470998136338958": "LucasRamos",
    "96473161397342454": "LucasRamos",
    "96454201749302012": "Camila",
    "96451968465288102": "Camila",
    "96449983914285918": "Camila",
    "96456284157307243": "Camila",
}


def resolver_responsavel_relatorio(nome, saved_value):
    """Resolve o nome que será concatenado no download do relatório."""
    nome = (nome or "").strip().upper()
    saved_value = (saved_value or "").strip()

    return (
        RESPONSAVEL_POR_RELATORIO_NOME.get(nome)
        or RESPONSAVEL_POR_SAVED_VALUE.get(saved_value)
        or os.getenv("RELATORIO_RESPONSAVEL", "").strip()
    )


def prefixo_responsavel_download(relatorio):
    """Retorna o responsável já higienizado para compor o nome do arquivo."""
    responsavel = (relatorio.get("responsavel") or "").strip()
    return limpar_nome_arquivo(responsavel) if responsavel else ""


def aba_consolidado_responsavel(responsavel):
    """
    Define em qual aba o arquivo deve entrar no consolidado.

    Observação: a aba WALCERI foi mantida exatamente conforme solicitado.
    """
    chave = re.sub(r"[^A-Z0-9]", "", (responsavel or "").upper())

    if chave in {"WALCEIR", "WALCERI"}:
        return "WALCERI"

    if chave == "CAMILA":
        return "CAMILA"

    if chave in {"LUCASRAMOS", "LUCAS"}:
        return "LUCASRAMOS"

    return "SEM_RESPONSAVEL"


def registrar_arquivo_para_consolidacao(relatorio, caminho_arquivo, tipo_origem):
    """Registra arquivo gerado na execução atual para consolidação final."""
    if not caminho_arquivo:
        return

    caminho = Path(caminho_arquivo)

    if not caminho.exists():
        return

    responsavel = relatorio.get("responsavel") or ""
    aba_destino = aba_consolidado_responsavel(responsavel)

    if aba_destino == "SEM_RESPONSAVEL":
        print(
            f"[CONSOLIDADO] Responsável não mapeado para {caminho.name}. "
            "Arquivo não entrará nas abas WALCERI/CAMILA/LUCASRAMOS."
        )
        return

    ARQUIVOS_DOWNLOAD_EXECUCAO.append(
        {
            "caminho": caminho,
            "relatorio": relatorio.get("nome", ""),
            "saved_value": relatorio.get("saved_value", ""),
            "responsavel": responsavel,
            "aba": aba_destino,
            "tipo": tipo_origem,
        }
    )


# ============================================================
# CRONÔMETRO / AUDITORIA DE EXECUÇÃO
# ============================================================
def formatar_duracao(segundos):
    segundos = int(max(segundos or 0, 0))
    horas, resto = divmod(segundos, 3600)
    minutos, segundos = divmod(resto, 60)
    return f"{horas:02d}:{minutos:02d}:{segundos:02d}"


def segundos_decorridos(inicio_perf):
    """Retorna o tempo decorrido em segundos, com duas casas."""
    try:
        return round(max(time.perf_counter() - inicio_perf, 0), 2)
    except Exception:
        return 0


def duracao_media(lista):
    valores = [float(v or 0) for v in lista]
    return round(sum(valores) / len(valores), 2) if valores else 0


def registrar_metrica_execucao(
    etapa,
    relatorio=None,
    numero_pagina=None,
    origem="",
    status="OK",
    duracao_segundos=0,
    detalhes="",
    arquivo="",
    inicio_texto=None,
    fim_texto=None,
):
    """Guarda uma medição de tempo para entrar na aba RESUMO_EXECUCAO do consolidado."""
    relatorio = relatorio or {}
    fim_dt = datetime.now()

    if fim_texto is None:
        fim_texto = fim_dt.strftime("%d/%m/%Y %H:%M:%S")

    if inicio_texto is None:
        try:
            inicio_dt = fim_dt.timestamp() - float(duracao_segundos or 0)
            inicio_texto = datetime.fromtimestamp(inicio_dt).strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            inicio_texto = ""

    duracao_segundos = round(float(duracao_segundos or 0), 2)

    registro = {
        "sequencia": len(METRICAS_EXECUCAO) + 1,
        "etapa": etapa or "",
        "relatorio": relatorio.get("nome", "") if isinstance(relatorio, dict) else "",
        "saved_value": relatorio.get("saved_value", "") if isinstance(relatorio, dict) else "",
        "lider_cenario": relatorio.get("responsavel", "") if isinstance(relatorio, dict) else "",
        "pagina": numero_pagina if numero_pagina is not None else "",
        "origem": origem or "",
        "status": status or "",
        "duracao_segundos": duracao_segundos,
        "duracao_formatada": formatar_duracao(duracao_segundos),
        "inicio": inicio_texto or "",
        "fim": fim_texto or "",
        "arquivo": str(arquivo or ""),
        "detalhes": str(detalhes or ""),
    }

    METRICAS_EXECUCAO.append(registro)

    print(
        f"[MÉTRICA] {registro['etapa']} | "
        f"relatorio={registro['relatorio'] or '-'} | "
        f"pagina={registro['pagina'] or '-'} | "
        f"origem={registro['origem'] or '-'} | "
        f"status={registro['status']} | "
        f"tempo={registro['duracao_segundos']}s"
    )


class CronometroExecucao:
    """
    Janela simples para acompanhar o tempo de execução.

    A tela é opcional e pode ser desativada no .env com:
    EXIBIR_TELA_TEMPO=N
    """

    def __init__(self, inicio_execucao):
        self.inicio_execucao = inicio_execucao
        self._inicio_perf = time.perf_counter()
        self._parar = False
        self._status = "Iniciando execução..."
        self._lock = threading.Lock()
        self._thread = None

    def iniciar(self):
        if not EXIBIR_TELA_TEMPO:
            return

        self._thread = threading.Thread(target=self._executar_janela, daemon=True)
        self._thread.start()

    def atualizar_status(self, status):
        with self._lock:
            self._status = status or "Executando..."

    def finalizar(self):
        with self._lock:
            self._status = "Finalizado. Consolidado gerado."
            self._parar = True

    def _executar_janela(self):
        try:
            import tkinter as tk
        except Exception as e:
            print(f"[TEMPO][AVISO] Não consegui abrir tela de tempo. Seguindo apenas pelo terminal. Detalhe: {e}")
            return

        try:
            root = tk.Tk()
            root.title("GTN - Tempo de Execução")
            root.geometry("460x185")
            root.resizable(False, False)

            try:
                root.attributes("-topmost", True)
            except Exception:
                pass

            frame = tk.Frame(root, padx=18, pady=16)
            frame.pack(fill="both", expand=True)

            titulo = tk.Label(frame, text="Automação GTN em execução", font=("Segoe UI", 13, "bold"))
            titulo.pack(anchor="w")

            lbl_inicio = tk.Label(
                frame,
                text=f"Início: {self.inicio_execucao.strftime('%d/%m/%Y %H:%M:%S')}",
                font=("Segoe UI", 10),
            )
            lbl_inicio.pack(anchor="w", pady=(8, 0))

            lbl_tempo = tk.Label(frame, text="Tempo decorrido: 00:00:00", font=("Segoe UI", 18, "bold"))
            lbl_tempo.pack(anchor="w", pady=(10, 4))

            lbl_status = tk.Label(frame, text="Status: Iniciando execução...", font=("Segoe UI", 10), wraplength=420, justify="left")
            lbl_status.pack(anchor="w")

            def atualizar():
                with self._lock:
                    parar = self._parar
                    status = self._status

                segundos = time.perf_counter() - self._inicio_perf
                lbl_tempo.config(text=f"Tempo decorrido: {formatar_duracao(segundos)}")
                lbl_status.config(text=f"Status: {status}")

                if parar:
                    root.after(1200, root.destroy)
                    return

                root.after(1000, atualizar)

            atualizar()
            root.mainloop()

        except Exception as e:
            print(f"[TEMPO][AVISO] A tela de tempo falhou, mas o script seguirá normalmente. Detalhe: {e}")


# ============================================================
# RELATÓRIOS PELO .ENV
#
# Formato da variável RELATORIOS:
# nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim
#
# Exemplo:
# RELATORIOS=PMO_EXECUCAO_TESTES|35932200234408468|96461147945312616|0|5|9|1002;OUTRO_RELATORIO|111|222|0|5|9|1002
# ============================================================
def montar_relatorio_config(item, origem):
    """
    Converte uma configuração textual de relatório em dict.

    Formato esperado:
    nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim
    """
    item = (item or "").strip()

    if not item:
        return None

    partes = [parte.strip() for parte in item.split("|")]

    if len(partes) != 7:
        raise Exception(
            f"Formato inválido em {origem}. "
            "Use: nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim"
        )

    nome, apex_id, saved_value, editar_inicio, editar_fim, tr_inicio, tr_fim = partes

    if not nome:
        raise Exception(f"Nome vazio em {origem}.")

    if not apex_id:
        raise Exception(f"apex_id vazio em {origem}.")

    if not saved_value:
        raise Exception(f"saved_value vazio em {origem}.")

    try:
        editar_inicio_int = int(editar_inicio)
        editar_fim_int = int(editar_fim)
        tr_inicio_int = int(tr_inicio)
        tr_fim_int = int(tr_fim)
    except Exception as e:
        raise Exception(
            f"Erro convertendo índices numéricos em {origem}. "
            f"Valores recebidos: editar_inicio={editar_inicio}, editar_fim={editar_fim}, "
            f"tr_inicio={tr_inicio}, tr_fim={tr_fim}."
        ) from e

    if tr_inicio_int >= 0 and tr_fim_int >= 0 and tr_fim_int < tr_inicio_int:
        print(
            f"[AVISO][CONFIG] {nome}: tr_fim ({tr_fim_int}) é menor que tr_inicio ({tr_inicio_int}). "
            "O bloco tr:nth-child será ignorado para este relatório."
        )

    return {
        "nome": nome,
        "apex_id": apex_id,
        "saved_value": saved_value,
        "responsavel": resolver_responsavel_relatorio(nome, saved_value),
        "editar_inicio": editar_inicio_int,
        "editar_fim": editar_fim_int,
        "tr_inicio": tr_inicio_int,
        "tr_fim": tr_fim_int,
    }


def carregar_relatorios_individuais_env():
    """
    Lê relatórios declarados um por linha no .env.

    Exemplo:
    RELATORIO_01=RELATORIO_01|35932200234408468|96443782140269930|0|5|9|45
    RELATORIO_02=RELATORIO_02|35932200234408468|96445783233272755|0|5|9|38

    Esse modelo evita o problema clássico do .env carregar somente o primeiro item
    quando a lista RELATORIOS é quebrada em várias linhas sem aspas.
    """
    encontrados = []

    for chave, valor in os.environ.items():
        match = re.fullmatch(r"RELATORIO_(\d{1,3})", chave.strip().upper())

        if not match:
            continue

        valor = (valor or "").strip()

        if not valor:
            continue

        encontrados.append((int(match.group(1)), chave, valor))

    relatorios = []

    for _, chave, valor in sorted(encontrados, key=lambda x: x[0]):
        relatorio = montar_relatorio_config(valor, chave)

        if relatorio:
            relatorios.append(relatorio)

    return relatorios


# ============================================================
# RELATÓRIOS PELO .ENV
#
# Modelo preferencial, um por linha:
# RELATORIO_01=nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim
# RELATORIO_02=nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim
#
# Modelo compatível legado:
# RELATORIOS=nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim;OUTRO|...
# ============================================================
def carregar_relatorios_env():
    # Preferencial: RELATORIO_01, RELATORIO_02, ..., um por linha.
    relatorios_individuais = carregar_relatorios_individuais_env()

    if relatorios_individuais:
        return relatorios_individuais

    # Compatibilidade: RELATORIOS em uma linha só, separando os relatórios por ponto e vírgula.
    relatorios_raw = os.getenv("RELATORIOS", "").strip()

    if relatorios_raw:
        relatorios = []

        for posicao, item in enumerate(relatorios_raw.split(";"), start=1):
            item = item.strip()

            if not item:
                continue

            relatorio = montar_relatorio_config(item, f"RELATORIOS item {posicao}")

            if relatorio:
                relatorios.append(relatorio)

        if not relatorios:
            raise Exception("A variável RELATORIOS está vazia ou inválida no .env.")

        return relatorios

    # Compatibilidade com o modelo antigo, caso ainda use variáveis separadas no .env.
    apex_id = os.getenv("RELATORIO_APEX_ID", "").strip()
    saved_value = os.getenv("RELATORIO_SALVO_VALUE", "").strip()

    if apex_id and saved_value:
        nome = os.getenv("RELATORIO_NOME", "PMO_EXECUCAO_TESTES").strip()

        return [
            {
                "nome": nome,
                "apex_id": apex_id,
                "saved_value": saved_value,
                "responsavel": resolver_responsavel_relatorio(nome, saved_value),
                "editar_inicio": int(os.getenv("EDITAR_INDICE_INICIO", "0")),
                "editar_fim": int(os.getenv("EDITAR_INDICE_FIM", "5")),
                "tr_inicio": int(os.getenv("TR_CHILD_INICIO", "9")),
                "tr_fim": int(os.getenv("TR_CHILD_FIM", "1002")),
            }
        ]

    raise Exception(
        "Nenhum relatório configurado. Informe RELATORIO_01, RELATORIO_02, etc. no .env, "
        "ou RELATORIOS em uma única linha no formato: "
        "nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim"
    )

RELATORIOS = carregar_relatorios_env()


# ============================================================
# UTILITÁRIOS
# ============================================================
def limpar_nome_arquivo(nome):
    nome = nome or "download"
    nome = re.sub(r'[\\/:*?"<>|]+', "_", nome)
    nome = re.sub(r"\s+", "_", nome).strip("._ ")
    return nome or "download"


def seletor_select_relatorio(relatorio):
    return f"#R{relatorio['apex_id']}_saved_reports"


def seletor_botao_ir_relatorio(relatorio):
    return f"#R{relatorio['apex_id']}_saved_reports_go"


def nome_pasta_lider_cenario(relatorio):
    """Resolve o nome da pasta pelo líder de cenário/responsável do relatório."""
    lider = (relatorio.get("responsavel") or "").strip()

    if not lider:
        lider = resolver_responsavel_relatorio(
            relatorio.get("nome", ""),
            relatorio.get("saved_value", ""),
        )

    return limpar_nome_arquivo(lider or relatorio.get("nome") or "SEM_LIDER_CENARIO")


def pasta_download_relatorio(relatorio):
    # A pasta agora é agrupada pelo líder de cenário, não pelo nome do relatório.
    # Exemplo: downloads_gtn/Walceir, downloads_gtn/LucasRamos, downloads_gtn/Camila.
    pasta = DOWNLOAD_ROOT / nome_pasta_lider_cenario(relatorio)
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta


def caminho_unico(caminho):
    if not caminho.exists():
        return caminho

    base = caminho.stem
    ext = caminho.suffix
    pasta = caminho.parent

    contador = 2
    while True:
        novo = pasta / f"{base}_{contador}{ext}"
        if not novo.exists():
            return novo
        contador += 1


def resolver_template_nao_iniciado(relatorio):
    """
    Retorna um arquivo modelo para criar o placeholder de cenário sem Edit.

    Prioridade:
    1) TEMPLATE_NAO_INICIADO informado no .env
    2) Primeiro download real já salvo para o relatório atual
    3) Nenhum modelo encontrado
    """
    if TEMPLATE_NAO_INICIADO:
        template = Path(TEMPLATE_NAO_INICIADO)
        if not template.is_absolute():
            template = BASE_DIR / template

        if template.exists():
            return template

        print(f"[AVISO] TEMPLATE_NAO_INICIADO informado, mas não encontrado: {template}")

    return MODELOS_DOWNLOAD_POR_RELATORIO.get(relatorio["nome"])


def registrar_modelo_download(relatorio, caminho_download):
    """Guarda o primeiro download real do relatório como template para placeholders."""
    if relatorio["nome"] not in MODELOS_DOWNLOAD_POR_RELATORIO:
        MODELOS_DOWNLOAD_POR_RELATORIO[relatorio["nome"]] = Path(caminho_download)
        print(f"[TEMPLATE] Modelo de colunas definido para {relatorio['nome']}: {caminho_download}")


def criar_placeholder_csv(template_path, destino):
    """Cria CSV vazio preservando o cabeçalho do template, quando possível."""
    header = ""

    if template_path and Path(template_path).exists():
        with open(template_path, "r", encoding="utf-8-sig", errors="ignore") as origem:
            header = origem.readline().rstrip("\n")

    with open(destino, "w", encoding="utf-8-sig", newline="") as saida:
        if header:
            saida.write(header + "\n")


def criar_placeholder_xlsx(template_path, destino):
    """
    Cria XLSX vazio usando o primeiro download real como modelo.
    Mantém cabeçalho/estrutura e apaga linhas de dados a partir da linha 2.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise Exception(
            "Para criar placeholder .xlsx preservando colunas, instale openpyxl: pip install openpyxl"
        ) from e

    wb = load_workbook(template_path)

    for ws in wb.worksheets:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    wb.save(destino)


# ============================================================
# METADADOS DO CENÁRIO DENTRO DO DOWNLOAD
# ============================================================
COLUNA_LIDER_CENARIO = "LIDER_CENARIO"
COLUNA_IDENTIFICADOR_CENARIO = "IDENTIFICADOR_CENARIO"
COLUNA_NOME_CENARIO = "NOME_CENARIO"


def remover_acentos(texto):
    texto = "" if texto is None else str(texto)
    return "".join(
        caractere
        for caractere in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(caractere)
    )


def normalizar_texto_busca(texto):
    texto = remover_acentos(texto).upper()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def texto_util(valor):
    valor = "" if valor is None else str(valor)
    return re.sub(r"\s+", " ", valor).strip()


def parece_texto_de_acao(valor):
    chave = normalizar_texto_busca(valor)
    return chave in {"EDITAR", "EDIT", "ACOES", "ACAO", "AÇÕES", "AÇÃO"}


def extrair_identificador_do_texto(texto):
    """Tenta achar identificadores como 0017_02 dentro do texto do cenário."""
    texto = texto_util(texto)

    padroes = [
        r"\b\d{3,}[_\-.]\d{1,}\b",          # 0017_02, 0017-02, 0017.02
        r"\b[A-Z]{1,5}\d{2,}[_\-.]\d{1,}\b", # CEN0017_02, CT0017_02
        r"\bID\s*[:\-]?\s*([A-Z0-9_.\-]{3,})\b",
        r"\bCEN[ÁA]RIO\s*[:\-]?\s*([A-Z0-9_.\-]{3,})\b",
    ]

    for padrao in padroes:
        match = re.search(padrao, texto, flags=re.IGNORECASE)
        if match:
            return match.group(1) if match.lastindex else match.group(0)

    return ""


def escolher_dados_cenario(dados_linha, origem):
    """Escolhe nome e identificador do cenário a partir das células da linha do relatório."""
    dados_linha = dados_linha or {}
    celulas = dados_linha.get("cells") or []
    headers = dados_linha.get("headers") or []
    linha_texto = texto_util(dados_linha.get("rowText") or "")

    itens = []
    for indice, celula in enumerate(celulas):
        texto_celula = texto_util((celula or {}).get("text") or "")
        if not texto_celula or parece_texto_de_acao(texto_celula):
            continue

        header = ""
        if indice < len(headers):
            header = headers[indice]

        possiveis_headers = [
            header,
            (celula or {}).get("dataLabel"),
            (celula or {}).get("headers"),
            (celula or {}).get("aria"),
            (celula or {}).get("title"),
        ]
        header_final = " ".join(texto_util(h) for h in possiveis_headers if texto_util(h))
        chave_header = normalizar_texto_busca(header_final)
        itens.append({"header": header_final, "chave_header": chave_header, "texto": texto_celula})

    identificador = ""
    nome = ""

    # 1) Identificador por coluna explícita.
    for item in itens:
        chave = item["chave_header"]
        if (
            "CENARIO" in chave
            and any(token in chave for token in ["ID", "IDENT", "COD", "CODIGO", "NUM", "NUMERO"])
        ):
            identificador = item["texto"]
            break

    # 2) Identificador por coluna genérica de ID/código, evitando status/responsável.
    if not identificador:
        for item in itens:
            chave = item["chave_header"]
            if (
                any(token in chave.split() for token in ["ID", "COD", "CODIGO", "IDENTIFICADOR"])
                and not any(token in chave for token in ["STATUS", "RESPONSAVEL", "RESPONSÁVEL", "DATA"])
            ):
                identificador = item["texto"]
                break

    # 3) Nome por coluna de cenário / descrição.
    for item in itens:
        chave = item["chave_header"]
        if "CENARIO" in chave and not any(token in chave for token in ["ID", "IDENT", "COD", "CODIGO", "STATUS"]):
            nome = item["texto"]
            break

    if not nome:
        for item in itens:
            chave = item["chave_header"]
            if any(token in chave for token in ["NOME", "DESCRICAO", "DESCRIÇÃO", "TITULO", "TÍTULO"]):
                nome = item["texto"]
                break

    # 4) Fallback: maior texto útil da linha geralmente é o nome do cenário.
    if not nome and itens:
        nome = max((item["texto"] for item in itens), key=len, default="")

    if not identificador:
        identificador = extrair_identificador_do_texto(nome) or extrair_identificador_do_texto(linha_texto)

    # 5) Último fallback para não deixar vazio.
    if not identificador:
        identificador = origem

    if not nome:
        nome = linha_texto or origem

    return {
        "identificador": texto_util(identificador),
        "nome": texto_util(nome),
        "linha_texto": linha_texto,
    }


def extrair_dados_cenario_da_linha(locator_linha, origem):
    """Captura os dados da linha antes de abrir o modal de Cenário e Execuções."""
    try:
        dados_linha = locator_linha.evaluate(
            """
            el => {
                const tr = el.closest('tr');
                if (!tr) {
                    return { erro: 'Não encontrei a linha TR mais próxima.' };
                }

                const normalizar = txt => (txt || '').replace(/\\s+/g, ' ').trim();

                const cells = Array.from(tr.querySelectorAll('th,td')).map(td => ({
                    text: normalizar(td.innerText || td.textContent || ''),
                    headers: normalizar(td.getAttribute('headers') || ''),
                    aria: normalizar(td.getAttribute('aria-label') || ''),
                    title: normalizar(td.getAttribute('title') || ''),
                    dataLabel: normalizar(td.getAttribute('data-label') || '')
                }));

                let headers = [];
                const table = tr.closest('table');
                if (table) {
                    headers = Array.from(table.querySelectorAll('thead th')).map(th =>
                        normalizar(th.innerText || th.textContent || '')
                    );
                }

                return {
                    cells,
                    headers,
                    rowText: cells.map(c => c.text).filter(Boolean).join(' | ')
                };
            }
            """
        )

        dados = escolher_dados_cenario(dados_linha, origem)
        print(
            f"[CENÁRIO] origem={origem} | "
            f"identificador={dados.get('identificador')} | nome={dados.get('nome')}"
        )
        return dados

    except Exception as e:
        print(f"[CENÁRIO][AVISO] Não consegui capturar nome/identificador da linha {origem}: {e}")
        return {
            "identificador": origem,
            "nome": origem,
            "linha_texto": "",
        }


def montar_metadados_cenario_para_arquivo(dados_cenario, origem):
    dados_cenario = dados_cenario or {}
    nome = texto_util(dados_cenario.get("nome") or dados_cenario.get("linha_texto") or origem)
    identificador = texto_util(dados_cenario.get("identificador") or extrair_identificador_do_texto(nome) or origem)

    return identificador, nome


def detectar_dialeto_csv_para_escrita(caminho):
    encoding = detectar_encoding_texto(caminho)

    with open(caminho, "r", encoding=encoding, newline="", errors="ignore") as arquivo:
        amostra = arquivo.read(8192)
        arquivo.seek(0)

        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";" if amostra.count(";") >= amostra.count(",") else ","

    return encoding, dialect


def inserir_coluna_metadado_csv(linhas, nome_coluna, valor):
    """Insere uma coluna de metadado no início do CSV, se ela ainda não existir."""
    if not linhas:
        linhas.append([])

    cabecalho = [texto_util(v) for v in linhas[0]]
    cabecalho_normalizado = [normalizar_texto_busca(v) for v in cabecalho]

    if normalizar_texto_busca(nome_coluna) in cabecalho_normalizado:
        return

    linhas[0].insert(0, nome_coluna)

    for indice in range(1, len(linhas)):
        linhas[indice].insert(0, valor)


def enriquecer_csv_com_cenario(caminho, identificador, nome, lider_cenario):
    encoding, dialect = detectar_dialeto_csv_para_escrita(caminho)

    with open(caminho, "r", encoding=encoding, newline="", errors="ignore") as entrada:
        linhas = list(csv.reader(entrada, dialect))

    if not linhas:
        linhas = [[]]

    # A ordem final das colunas fica:
    # LIDER_CENARIO | IDENTIFICADOR_CENARIO | NOME_CENARIO | demais colunas do download
    inserir_coluna_metadado_csv(linhas, COLUNA_NOME_CENARIO, nome)
    inserir_coluna_metadado_csv(linhas, COLUNA_IDENTIFICADOR_CENARIO, identificador)
    inserir_coluna_metadado_csv(linhas, COLUNA_LIDER_CENARIO, lider_cenario)

    with open(caminho, "w", encoding="utf-8-sig", newline="") as saida:
        escritor = csv.writer(saida, delimiter=getattr(dialect, "delimiter", ";"), lineterminator="\n")
        escritor.writerows(linhas)


def inserir_coluna_metadado_xlsx(ws, nome_coluna, valor):
    """Insere uma coluna de metadado no início da planilha, se ela ainda não existir."""
    cabecalho = [texto_util(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    cabecalho_normalizado = [normalizar_texto_busca(v) for v in cabecalho]

    if normalizar_texto_busca(nome_coluna) in cabecalho_normalizado:
        return

    ws.insert_cols(1)
    ws.cell(row=1, column=1).value = nome_coluna

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).value = valor


def enriquecer_xlsx_com_cenario(caminho, identificador, nome, lider_cenario):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise Exception("Para enriquecer .xlsx, instale openpyxl: pip install openpyxl") from e

    wb = load_workbook(caminho)

    for ws in wb.worksheets:
        # A ordem final das colunas fica:
        # LIDER_CENARIO | IDENTIFICADOR_CENARIO | NOME_CENARIO | demais colunas do download
        inserir_coluna_metadado_xlsx(ws, COLUNA_NOME_CENARIO, nome)
        inserir_coluna_metadado_xlsx(ws, COLUNA_IDENTIFICADOR_CENARIO, identificador)
        inserir_coluna_metadado_xlsx(ws, COLUNA_LIDER_CENARIO, lider_cenario)

    wb.save(caminho)


def resolver_lider_cenario_para_arquivo(relatorio):
    """Resolve o líder do cenário que será gravado dentro do arquivo baixado."""
    if not relatorio:
        return ""

    return texto_util(
        relatorio.get("responsavel")
        or resolver_responsavel_relatorio(relatorio.get("nome", ""), relatorio.get("saved_value", ""))
    )


def enriquecer_arquivo_download_com_cenario(caminho, dados_cenario, origem, relatorio=None):
    """
    Inclui LIDER_CENARIO, IDENTIFICADOR_CENARIO e NOME_CENARIO dentro do arquivo baixado/criado.

    Para CSV/TXT: regrava o arquivo com as colunas no início.
    Para XLSX/XLSM: insere as colunas no início da primeira linha e preenche as linhas de dados.
    """
    caminho = Path(caminho)

    if not caminho.exists():
        return

    identificador, nome = montar_metadados_cenario_para_arquivo(dados_cenario, origem)
    lider_cenario = resolver_lider_cenario_para_arquivo(relatorio)
    extensao = caminho.suffix.lower()

    try:
        if extensao in {".csv", ".txt"}:
            enriquecer_csv_com_cenario(caminho, identificador, nome, lider_cenario)
        elif extensao in {".xlsx", ".xlsm"}:
            enriquecer_xlsx_com_cenario(caminho, identificador, nome, lider_cenario)
        elif extensao == ".xls":
            # Muitos downloads APEX com extensão .xls são texto delimitado.
            # Se for HTML real, evita corromper o arquivo.
            encoding = detectar_encoding_texto(caminho)
            amostra = Path(caminho).read_text(encoding=encoding, errors="ignore")[:500].lower()
            if "<html" in amostra or "<table" in amostra:
                print(
                    f"[CENÁRIO][AVISO] {caminho.name} parece ser HTML com extensão .xls. "
                    "Não alterei o conteúdo para evitar corromper o arquivo."
                )
                return
            enriquecer_csv_com_cenario(caminho, identificador, nome, lider_cenario)
        else:
            print(f"[CENÁRIO][AVISO] Extensão não suportada para enriquecer metadados: {caminho.name}")
            return

        print(
            f"[CENÁRIO] Metadados gravados no arquivo: {caminho.name} | "
            f"{COLUNA_LIDER_CENARIO}={lider_cenario} | "
            f"{COLUNA_IDENTIFICADOR_CENARIO}={identificador} | "
            f"{COLUNA_NOME_CENARIO}={nome}"
        )

    except Exception as e:
        print(f"[CENÁRIO][ERRO] Não consegui gravar líder/nome/identificador no arquivo {caminho.name}: {e}")

def criar_download_nao_iniciado(page, origem, numero_pagina, relatorio, dados_cenario=None):
    """
    Cria arquivo placeholder quando o cenário existe, mas não possui link Edit.

    O arquivo usa o mesmo cabeçalho/colunas de um download real quando houver template.
    Se ainda não existir template, gera um CSV vazio para registrar o cenário não iniciado.
    """
    if not GERAR_PLACEHOLDER_SEM_EDIT:
        print("[PLACEHOLDER] Geração de placeholder desativada no .env.")
        return None

    pasta_relatorio = pasta_download_relatorio(relatorio)
    template = resolver_template_nao_iniciado(relatorio)

    responsavel_download = prefixo_responsavel_download(relatorio)

    if responsavel_download:
        nome_base = (
            f"pagina_{numero_pagina}_{responsavel_download}_"
            f"{limpar_nome_arquivo(origem)}_Cenario_e_Execucao_nao_iniciado"
        )
    else:
        nome_base = f"pagina_{numero_pagina}_{limpar_nome_arquivo(origem)}_Cenario_e_Execucao_nao_iniciado"

    if template and Path(template).exists():
        extensao = Path(template).suffix.lower() or ".csv"
    else:
        extensao = ".csv"

    destino = caminho_unico(pasta_relatorio / f"{nome_base}{extensao}")

    try:
        if template and Path(template).exists() and extensao in {".xlsx", ".xlsm"}:
            criar_placeholder_xlsx(template, destino)
        elif template and Path(template).exists() and extensao in {".csv", ".txt"}:
            criar_placeholder_csv(template, destino)
        else:
            # Sem template ainda: cria um marcador CSV vazio.
            # Assim o processo não perde o registro do cenário sem execução iniciada.
            with open(destino, "w", encoding="utf-8-sig", newline="") as saida:
                saida.write("")

        enriquecer_arquivo_download_com_cenario(destino, dados_cenario, origem, relatorio)
        print(f"[PLACEHOLDER] Arquivo de cenário não iniciado criado: {destino}")
        registrar_arquivo_para_consolidacao(relatorio, destino, "placeholder_sem_edit")
        return destino

    except Exception as e:
        print("[ERRO] Não consegui criar o arquivo placeholder de cenário não iniciado.")
        print(f"[DETALHE] {e}")
        return None


def validar_env():
    if not GTN_USUARIO:
        raise Exception("Não encontrei GTN_USUARIO ou GTN_USER no arquivo .env.")

    if not GTN_SENHA:
        raise Exception("Não encontrei GTN_SENHA ou GTN_PASS no arquivo .env.")

    if not URL_LOGIN:
        raise Exception("URL_LOGIN não pode ficar vazia no .env.")

    if not RELATORIOS:
        raise Exception("Nenhum relatório carregado do .env.")

    if GERAR_CONSOLIDADO_FINAL:
        try:
            import openpyxl  # noqa: F401
        except Exception as e:
            raise Exception(
                "GERAR_CONSOLIDADO_FINAL está ativo, mas o pacote openpyxl não está instalado. "
                "Execute: pip install openpyxl  ou  pip install -r requirements.txt"
            ) from e


def aguardar_processamento_apex(page):
    seletores = [
        ".u-Processing",
        ".a-Processing",
        ".apex_wait_overlay",
        ".ui-widget-overlay",
    ]

    for seletor in seletores:
        try:
            page.locator(seletor).first.wait_for(state="hidden", timeout=8000)
        except Exception:
            pass


def aguardar_carregamento(page):
    try:
        page.wait_for_load_state("domcontentloaded", timeout=20000)
    except Exception:
        pass

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass

    aguardar_processamento_apex(page)


def aguardar_relatorio_pronto(page):
    aguardar_processamento_apex(page)

    try:
        page.get_by_role("cell", name="Editar").first.wait_for(state="visible", timeout=30000)
        return True
    except Exception:
        pass

    try:
        page.locator("tr > .a-IRR-linkCol").first.wait_for(state="visible", timeout=30000)
        return True
    except Exception:
        pass

    return False


def listar_opcoes_relatorio(select_relatorio):
    try:
        return select_relatorio.evaluate(
            """
            el => Array.from(el.options).map(opt => ({
                value: opt.value,
                text: opt.textContent.trim(),
                selected: opt.selected
            }))
            """
        )
    except Exception as e:
        return [{"erro": str(e)}]


def obter_frame(page, titulo):
    iframe = page.locator(f'iframe[title="{titulo}"]').first
    iframe.wait_for(state="attached", timeout=TIMEOUT_PADRAO)

    try:
        iframe.wait_for(state="visible", timeout=TIMEOUT_PADRAO)
    except Exception:
        pass

    return iframe.content_frame


# ============================================================
# CONSOLIDAÇÃO FINAL DOS DOWNLOADS
# ============================================================
def linha_vazia(linha):
    return not linha or all(valor is None or str(valor).strip() == "" for valor in linha)


def normalizar_nome_coluna(valor, indice):
    nome = "" if valor is None else str(valor).strip()
    nome = re.sub(r"\s+", " ", nome)
    return nome or f"COLUNA_{indice}"


def normalizar_cabecalho(linha):
    cabecalho = []
    usados = {}

    for indice, valor in enumerate(linha or [], start=1):
        nome_base = normalizar_nome_coluna(valor, indice)
        nome = nome_base
        contador = 2

        while nome.upper() in usados:
            nome = f"{nome_base}_{contador}"
            contador += 1

        usados[nome.upper()] = True
        cabecalho.append(nome)

    return cabecalho


def detectar_encoding_texto(caminho):
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            with open(caminho, "r", encoding=encoding) as arquivo:
                arquivo.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin1"


def iterar_linhas_csv(caminho):
    encoding = detectar_encoding_texto(caminho)

    with open(caminho, "r", encoding=encoding, newline="", errors="ignore") as arquivo:
        amostra = arquivo.read(8192)
        arquivo.seek(0)

        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";" if amostra.count(";") >= amostra.count(",") else ","

        for linha in csv.reader(arquivo, dialect):
            yield list(linha)


def iterar_linhas_xlsx(caminho):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise Exception("Para consolidar .xlsx, instale openpyxl: pip install openpyxl") from e

    wb = load_workbook(caminho, read_only=True, data_only=True)

    try:
        ws = wb.worksheets[0]
        for linha in ws.iter_rows(values_only=True):
            yield list(linha)
    finally:
        wb.close()


def iterar_linhas_arquivo(caminho):
    caminho = Path(caminho)
    extensao = caminho.suffix.lower()

    if extensao in {".xlsx", ".xlsm"}:
        yield from iterar_linhas_xlsx(caminho)
        return

    if extensao in {".csv", ".txt"}:
        yield from iterar_linhas_csv(caminho)
        return

    # Alguns downloads APEX podem vir como .xls, mas na prática serem texto/HTML.
    # Primeiro tenta como planilha moderna; se não der, tenta como texto delimitado.
    if extensao == ".xls":
        try:
            yield from iterar_linhas_xlsx(caminho)
            return
        except Exception:
            yield from iterar_linhas_csv(caminho)
            return

    raise Exception(f"Extensão não suportada para consolidação: {extensao}")


def obter_cabecalho_arquivo(caminho):
    for linha in iterar_linhas_arquivo(caminho):
        if not linha_vazia(linha):
            return normalizar_cabecalho(linha)

    return []


def construir_cabecalhos_consolidados(arquivos_por_aba):
    cabecalhos_por_aba = {}

    for aba, arquivos in arquivos_por_aba.items():
        uniao = []
        usados = set()

        for info in arquivos:
            caminho = info["caminho"]

            try:
                cabecalho = obter_cabecalho_arquivo(caminho)
            except Exception as e:
                print(f"[CONSOLIDADO][AVISO] Não consegui ler cabeçalho de {caminho.name}: {e}")
                cabecalho = []

            info["cabecalho"] = cabecalho

            for nome_coluna in cabecalho:
                chave = nome_coluna.upper()
                if chave not in usados:
                    usados.add(chave)
                    uniao.append(nome_coluna)

        cabecalhos_por_aba[aba] = [
            "RESPONSAVEL",
            "RELATORIO_ORIGEM",
            "SAVED_VALUE",
            "ARQUIVO_ORIGEM",
            "TIPO_ORIGEM",
            "STATUS_CONSOLIDACAO",
            "LINHA_ORIGEM",
        ] + uniao

    return cabecalhos_por_aba


def escrever_cabecalho(ws, cabecalho):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fonte = Font(bold=True, color="FFFFFF")
    preenchimento = PatternFill(fill_type="solid", fgColor="1F4E78")
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

    celulas = []
    for valor in cabecalho:
        celula = WriteOnlyCell(ws, value=valor)
        celula.font = fonte
        celula.fill = preenchimento
        celula.alignment = alinhamento
        celulas.append(celula)

    ws.append(celulas)
    ws.freeze_panes = "A2"

    for indice, nome_coluna in enumerate(cabecalho, start=1):
        largura = min(max(len(str(nome_coluna)) + 2, 12), 45)
        ws.column_dimensions[get_column_letter(indice)].width = largura


def criar_celula_resumo(ws, valor, negrito=False, preenchimento=None):
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill

    celula = WriteOnlyCell(ws, value=valor)
    celula.alignment = Alignment(vertical="center", wrap_text=True)

    if negrito:
        celula.font = Font(bold=True, color="FFFFFF" if preenchimento else "000000")

    if preenchimento:
        celula.fill = PatternFill(fill_type="solid", fgColor=preenchimento)

    return celula


def escrever_linha_resumo(ws, valores, cabecalho=False):
    preenchimento = "1F4E78" if cabecalho else None
    ws.append([criar_celula_resumo(ws, valor, negrito=cabecalho, preenchimento=preenchimento) for valor in valores])


def contar_arquivos_por_aba(arquivos_por_aba, aba, tipo=None):
    arquivos = arquivos_por_aba.get(aba, [])
    if tipo is None:
        return len(arquivos)
    return sum(1 for info in arquivos if info.get("tipo") == tipo)


def escrever_aba_resumo_execucao(
    wb_saida,
    destino,
    inicio_execucao,
    fim_execucao,
    resumos_execucao,
    arquivos_por_aba,
    totais_por_aba,
    estatisticas_arquivos,
):
    from openpyxl.utils import get_column_letter

    ws = wb_saida.create_sheet(title="RESUMO_EXECUCAO")

    if inicio_execucao is None:
        inicio_execucao = fim_execucao

    duracao_segundos = max((fim_execucao - inicio_execucao).total_seconds(), 0)

    total_paginas = sum((resumo or {}).get("paginas", 0) for resumo in (resumos_execucao or []))
    total_downloads_ok = sum((resumo or {}).get("downloads_ok", 0) for resumo in (resumos_execucao or []))
    total_sem_edit = sum((resumo or {}).get("sem_edit_ou_erro", 0) for resumo in (resumos_execucao or []))
    total_arquivos = len(estatisticas_arquivos)
    total_linhas_consolidadas = sum(info.get("linhas_consolidadas", 0) for info in estatisticas_arquivos)
    total_downloads_arquivo = sum(1 for info in estatisticas_arquivos if info.get("tipo") == "download")
    total_placeholders = sum(1 for info in estatisticas_arquivos if info.get("tipo") == "placeholder_sem_edit")
    total_erros_consolidacao = sum(1 for info in estatisticas_arquivos if str(info.get("status", "")).startswith("ERRO"))

    escrever_linha_resumo(ws, ["RESUMO DA EXECUÇÃO GTN"], cabecalho=True)
    escrever_linha_resumo(ws, [])

    escrever_linha_resumo(ws, ["Indicador", "Valor"], cabecalho=True)
    escrever_linha_resumo(ws, ["Início da execução", inicio_execucao.strftime("%d/%m/%Y %H:%M:%S")])
    escrever_linha_resumo(ws, ["Fim da execução / geração do consolidado", fim_execucao.strftime("%d/%m/%Y %H:%M:%S")])
    escrever_linha_resumo(ws, ["Tempo total", formatar_duracao(duracao_segundos)])
    escrever_linha_resumo(ws, ["Tempo total em segundos", round(duracao_segundos, 2)])
    escrever_linha_resumo(ws, ["Arquivo consolidado", str(destino)])
    escrever_linha_resumo(ws, ["Relatórios configurados", len(RELATORIOS)])
    escrever_linha_resumo(ws, ["Relatórios processados", len(resumos_execucao or [])])
    escrever_linha_resumo(ws, ["Páginas processadas", total_paginas])
    escrever_linha_resumo(ws, ["Downloads OK no fluxo", total_downloads_ok])
    escrever_linha_resumo(ws, ["Sem Edit / erro / pulados no fluxo", total_sem_edit])
    escrever_linha_resumo(ws, ["Arquivos gerados e considerados no consolidado", total_arquivos])
    escrever_linha_resumo(ws, ["Arquivos de download real", total_downloads_arquivo])
    escrever_linha_resumo(ws, ["Arquivos placeholder sem Edit", total_placeholders])
    escrever_linha_resumo(ws, ["Linhas gravadas nas abas consolidadas", total_linhas_consolidadas])
    escrever_linha_resumo(ws, ["Erros durante a consolidação", total_erros_consolidacao])
    escrever_linha_resumo(ws, ["Medições de tempo capturadas", len(METRICAS_EXECUCAO)])

    metricas_download = [m for m in METRICAS_EXECUCAO if m.get("etapa") == "LINHA_COM_EDIT_DOWNLOAD"]
    metricas_placeholder = [m for m in METRICAS_EXECUCAO if m.get("etapa") == "LINHA_SEM_EDIT_PLACEHOLDER"]
    metricas_login = [m for m in METRICAS_EXECUCAO if m.get("etapa") == "LOGIN_E_ABERTURA_EXECUCAO_TESTES"]
    metricas_carga_relatorio = [m for m in METRICAS_EXECUCAO if m.get("etapa") == "CARGA_RELATORIO_APEX"]
    metricas_proxima_pagina = [m for m in METRICAS_EXECUCAO if m.get("etapa") == "CARGA_PROXIMA_PAGINA_APEX"]

    escrever_linha_resumo(ws, ["Tempo login/abertura Execução de Testes - segundos", metricas_login[-1].get("duracao_segundos", 0) if metricas_login else 0])
    escrever_linha_resumo(ws, ["Tempo médio carga relatório APEX - segundos", duracao_media([m.get("duracao_segundos") for m in metricas_carga_relatorio])])
    escrever_linha_resumo(ws, ["Tempo médio próxima página APEX - segundos", duracao_media([m.get("duracao_segundos") for m in metricas_proxima_pagina])])
    escrever_linha_resumo(ws, ["Tempo médio linha com Edit/download - segundos", duracao_media([m.get("duracao_segundos") for m in metricas_download])])
    escrever_linha_resumo(ws, ["Tempo médio linha sem Edit/placeholder - segundos", duracao_media([m.get("duracao_segundos") for m in metricas_placeholder])])
    escrever_linha_resumo(ws, ["Maior tempo linha com Edit/download - segundos", max([m.get("duracao_segundos", 0) for m in metricas_download], default=0)])
    escrever_linha_resumo(ws, ["Maior tempo linha sem Edit/placeholder - segundos", max([m.get("duracao_segundos", 0) for m in metricas_placeholder], default=0)])

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Resumo por aba/responsável"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        ["ABA", "ARQUIVOS", "DOWNLOADS", "PLACEHOLDERS", "LINHAS_CONSOLIDADAS", "RELATORIOS_ENVOLVIDOS"],
        cabecalho=True,
    )

    for aba in ["WALCERI", "CAMILA", "LUCASRAMOS"]:
        relatorios_aba = sorted({info.get("relatorio", "") for info in arquivos_por_aba.get(aba, []) if info.get("relatorio")})
        escrever_linha_resumo(
            ws,
            [
                aba,
                contar_arquivos_por_aba(arquivos_por_aba, aba),
                contar_arquivos_por_aba(arquivos_por_aba, aba, "download"),
                contar_arquivos_por_aba(arquivos_por_aba, aba, "placeholder_sem_edit"),
                totais_por_aba.get(aba, 0),
                ", ".join(relatorios_aba),
            ],
        )

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Resumo por relatório"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        ["RELATORIO", "RESPONSAVEL", "STATUS", "PAGINAS", "DOWNLOADS_OK", "SEM_EDIT_OU_ERRO", "ARQUIVOS_GERADOS", "ERRO"],
        cabecalho=True,
    )

    arquivos_por_relatorio = {}
    responsavel_por_relatorio = {}
    for info in estatisticas_arquivos:
        relatorio = info.get("relatorio", "")
        arquivos_por_relatorio[relatorio] = arquivos_por_relatorio.get(relatorio, 0) + 1
        responsavel_por_relatorio[relatorio] = info.get("responsavel", "")

    for resumo in resumos_execucao or []:
        relatorio = resumo.get("relatorio", "")
        escrever_linha_resumo(
            ws,
            [
                relatorio,
                resumo.get("responsavel", "") or responsavel_por_relatorio.get(relatorio, ""),
                resumo.get("status", "OK"),
                resumo.get("paginas", 0),
                resumo.get("downloads_ok", 0),
                resumo.get("sem_edit_ou_erro", 0),
                arquivos_por_relatorio.get(relatorio, 0),
                resumo.get("erro", ""),
            ],
        )

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Detalhe dos arquivos consolidados"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        [
            "ABA",
            "RESPONSAVEL",
            "RELATORIO",
            "SAVED_VALUE",
            "ARQUIVO",
            "TIPO",
            "STATUS",
            "LINHAS_CONSOLIDADAS",
            "LINHAS_DADOS_REAIS",
        ],
        cabecalho=True,
    )

    for info in estatisticas_arquivos:
        escrever_linha_resumo(
            ws,
            [
                info.get("aba", ""),
                info.get("responsavel", ""),
                info.get("relatorio", ""),
                info.get("saved_value", ""),
                info.get("arquivo", ""),
                info.get("tipo", ""),
                info.get("status", ""),
                info.get("linhas_consolidadas", 0),
                info.get("linhas_dados_reais", 0),
            ],
        )

    escrever_linha_resumo(ws, [])
    escrever_linha_resumo(ws, ["Relatório de tempos por etapa"], cabecalho=True)
    escrever_linha_resumo(
        ws,
        [
            "SEQ",
            "ETAPA",
            "RELATORIO",
            "LIDER_CENARIO",
            "PAGINA",
            "ORIGEM_LINHA",
            "STATUS",
            "DURACAO_SEGUNDOS",
            "DURACAO_FORMATADA",
            "INICIO",
            "FIM",
            "ARQUIVO",
            "DETALHES",
        ],
        cabecalho=True,
    )

    for metrica in METRICAS_EXECUCAO:
        escrever_linha_resumo(
            ws,
            [
                metrica.get("sequencia", ""),
                metrica.get("etapa", ""),
                metrica.get("relatorio", ""),
                metrica.get("lider_cenario", ""),
                metrica.get("pagina", ""),
                metrica.get("origem", ""),
                metrica.get("status", ""),
                metrica.get("duracao_segundos", 0),
                metrica.get("duracao_formatada", ""),
                metrica.get("inicio", ""),
                metrica.get("fim", ""),
                metrica.get("arquivo", ""),
                metrica.get("detalhes", ""),
            ],
        )

    ws.freeze_panes = "A4"
    larguras = {
        1: 24,
        2: 32,
        3: 26,
        4: 22,
        5: 58,
        6: 24,
        7: 24,
        8: 22,
        9: 22,
        10: 24,
        11: 24,
        12: 70,
        13: 100,
    }

    for indice, largura in larguras.items():
        ws.column_dimensions[get_column_letter(indice)].width = largura


def consolidar_arquivos_download_execucao(resumos_execucao=None, inicio_execucao=None):
    """
    Consolida os arquivos gerados na execução atual em um XLSX único.

    O arquivo final fica na pasta raiz de downloads com nome:
    consolidado_HH-MM-SS.xlsx

    Os dois-pontos foram substituídos por hífen para compatibilidade com Windows.
    """
    if not GERAR_CONSOLIDADO_FINAL:
        print("[CONSOLIDADO] Consolidação final desativada no .env.")
        return None

    if not ARQUIVOS_DOWNLOAD_EXECUCAO:
        print("[CONSOLIDADO] Nenhum arquivo da execução atual para consolidar. Mesmo assim vou gerar o consolidado com RESUMO_EXECUCAO.")

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise Exception("Para gerar o consolidado final, instale openpyxl: pip install openpyxl") from e

    DOWNLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    nome_consolidado = datetime.now().strftime("consolidado_%H-%M-%S.xlsx")
    destino = caminho_unico(DOWNLOAD_ROOT / nome_consolidado)

    abas_obrigatorias = ["WALCERI", "CAMILA", "LUCASRAMOS"]
    arquivos_por_aba = {aba: [] for aba in abas_obrigatorias}

    for info in ARQUIVOS_DOWNLOAD_EXECUCAO:
        aba = info.get("aba")
        if aba in arquivos_por_aba:
            arquivos_por_aba[aba].append(info)

    cabecalhos_por_aba = construir_cabecalhos_consolidados(arquivos_por_aba)

    wb_saida = Workbook(write_only=True)
    totais_por_aba = {}
    estatisticas_arquivos = []

    for aba in abas_obrigatorias:
        ws = wb_saida.create_sheet(title=aba)
        cabecalho_saida = cabecalhos_por_aba[aba]
        escrever_cabecalho(ws, cabecalho_saida)

        mapa_colunas_saida = {nome.upper(): idx for idx, nome in enumerate(cabecalho_saida)}
        total_linhas_dados = 0

        for info in arquivos_por_aba[aba]:
            caminho = info["caminho"]
            cabecalho_arquivo = info.get("cabecalho") or []
            estatistica = {
                "aba": aba,
                "responsavel": info.get("responsavel", ""),
                "relatorio": info.get("relatorio", ""),
                "saved_value": info.get("saved_value", ""),
                "arquivo": caminho.name,
                "tipo": info.get("tipo", ""),
                "status": "OK",
                "linhas_consolidadas": 0,
                "linhas_dados_reais": 0,
            }

            if not cabecalho_arquivo:
                ws.append([
                    info.get("responsavel", ""),
                    info.get("relatorio", ""),
                    info.get("saved_value", ""),
                    caminho.name,
                    info.get("tipo", ""),
                    "SEM_DADOS",
                    "",
                ] + [None] * (len(cabecalho_saida) - 7))
                total_linhas_dados += 1
                estatistica["status"] = "SEM_DADOS"
                estatistica["linhas_consolidadas"] = 1
                estatisticas_arquivos.append(estatistica)
                continue

            primeira_linha_util = True
            linhas_dados_arquivo = 0
            linha_origem = 0

            try:
                for linha in iterar_linhas_arquivo(caminho):
                    if linha_vazia(linha):
                        continue

                    linha_origem += 1

                    if primeira_linha_util:
                        primeira_linha_util = False
                        # Primeira linha útil é tratada como cabeçalho.
                        continue

                    saida = [None] * len(cabecalho_saida)
                    saida[0] = info.get("responsavel", "")
                    saida[1] = info.get("relatorio", "")
                    saida[2] = info.get("saved_value", "")
                    saida[3] = caminho.name
                    saida[4] = info.get("tipo", "")
                    saida[5] = "OK"
                    saida[6] = linha_origem

                    for indice_coluna, valor in enumerate(linha):
                        if indice_coluna >= len(cabecalho_arquivo):
                            continue

                        nome_coluna = cabecalho_arquivo[indice_coluna]
                        indice_saida = mapa_colunas_saida.get(nome_coluna.upper())

                        if indice_saida is not None:
                            saida[indice_saida] = valor

                    ws.append(saida)
                    total_linhas_dados += 1
                    linhas_dados_arquivo += 1

                if linhas_dados_arquivo == 0:
                    ws.append([
                        info.get("responsavel", ""),
                        info.get("relatorio", ""),
                        info.get("saved_value", ""),
                        caminho.name,
                        info.get("tipo", ""),
                        "SEM_DADOS",
                        "",
                    ] + [None] * (len(cabecalho_saida) - 7))
                    total_linhas_dados += 1
                    estatistica["status"] = "SEM_DADOS"
                    estatistica["linhas_consolidadas"] = 1
                    estatistica["linhas_dados_reais"] = 0
                else:
                    estatistica["status"] = "OK"
                    estatistica["linhas_consolidadas"] = linhas_dados_arquivo
                    estatistica["linhas_dados_reais"] = linhas_dados_arquivo

            except Exception as e:
                print(f"[CONSOLIDADO][ERRO] Falha ao consolidar {caminho.name}: {e}")
                ws.append([
                    info.get("responsavel", ""),
                    info.get("relatorio", ""),
                    info.get("saved_value", ""),
                    caminho.name,
                    info.get("tipo", ""),
                    f"ERRO: {e}",
                    "",
                ] + [None] * (len(cabecalho_saida) - 7))
                total_linhas_dados += 1
                estatistica["status"] = f"ERRO: {e}"
                estatistica["linhas_consolidadas"] = 1
                estatistica["linhas_dados_reais"] = 0

            estatisticas_arquivos.append(estatistica)

        if total_linhas_dados > 0:
            ultima_coluna = get_column_letter(len(cabecalho_saida))
            ws.auto_filter.ref = f"A1:{ultima_coluna}{total_linhas_dados + 1}"

        totais_por_aba[aba] = total_linhas_dados

    fim_execucao = datetime.now()
    escrever_aba_resumo_execucao(
        wb_saida=wb_saida,
        destino=destino,
        inicio_execucao=inicio_execucao,
        fim_execucao=fim_execucao,
        resumos_execucao=resumos_execucao,
        arquivos_por_aba=arquivos_por_aba,
        totais_por_aba=totais_por_aba,
        estatisticas_arquivos=estatisticas_arquivos,
    )

    wb_saida.save(destino)

    print("=" * 100)
    print(f"[CONSOLIDADO] Arquivo final criado: {destino}")
    print("[CONSOLIDADO] Abas geradas:")
    for aba in abas_obrigatorias:
        print(f"- {aba}: {totais_por_aba.get(aba, 0)} linha(s)")
    print(f"- RESUMO_EXECUCAO: tempo={formatar_duracao((fim_execucao - (inicio_execucao or fim_execucao)).total_seconds())}")
    print("=" * 100)

    return destino


# ============================================================
# LOGIN E NAVEGAÇÃO
# ============================================================
def login_e_abrir_execucao_testes(page):
    inicio_login_perf = time.perf_counter()

    print("[1] Abrindo tela de login...")
    page.goto(URL_LOGIN, wait_until="domcontentloaded")
    page.wait_for_timeout(2000)

    print("[2] Preenchendo usuário...")
    campo_usuario = page.get_by_role("textbox", name="Usuário")
    campo_usuario.wait_for(state="visible")
    campo_usuario.fill(GTN_USUARIO)

    page.wait_for_timeout(800)

    print("[3] Indo para senha...")
    campo_usuario.press("Tab")
    page.wait_for_timeout(800)

    print("[4] Preenchendo senha...")
    campo_senha = page.get_by_role("textbox", name="Senha")
    campo_senha.wait_for(state="visible")
    campo_senha.fill(GTN_SENHA)

    page.wait_for_timeout(1000)

    print("[5] Clicando em Acessar...")
    botao_acessar = page.get_by_role("button", name="Acessar")
    botao_acessar.wait_for(state="visible")
    botao_acessar.click()

    print("[6] Aguardando pós-login/renderização...")
    page.wait_for_timeout(5000)
    aguardar_carregamento(page)
    page.wait_for_timeout(3000)

    print(f"[INFO] URL atual depois do login: {page.url}")

    print("[7] Aguardando botão Navegação Principal...")
    botao_nav = page.get_by_role("button", name="Navegação Principal")
    botao_nav.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1500)

    print("[8] Clicando em Navegação Principal...")
    botao_nav.click()
    page.wait_for_timeout(3000)

    print("[9] Aguardando toggle da árvore...")
    toggle_arvore = page.locator(".a-TreeView-toggle:visible").first
    toggle_arvore.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print("[10] Clicando no toggle da árvore...")
    toggle_arvore.click()
    page.wait_for_timeout(3000)

    print("[11] Aguardando item Execução de Testes...")
    item_execucao = page.get_by_role("treeitem", name="Execução de Testes")
    item_execucao.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print("[12] Clicando em Execução de Testes...")
    item_execucao.click()
    page.wait_for_timeout(7000)
    aguardar_carregamento(page)

    print("[OK] Clique em Execução de Testes executado.")
    print(f"[INFO] URL atual após clique em Execução de Testes: {page.url}")

    print("[13] Mantendo a página atual de Execução de Testes.")
    print("[INFO] Não vou usar page.goto com session/cs fixos, porque isso pode voltar para o login.")
    print(f"[INFO] URL atual válida: {page.url}")

    page.wait_for_timeout(7000)

    registrar_metrica_execucao(
        etapa="LOGIN_E_ABERTURA_EXECUCAO_TESTES",
        status="OK",
        duracao_segundos=segundos_decorridos(inicio_login_perf),
        detalhes="Tempo desde abrir URL de login até a tela de Execução de Testes ficar carregada.",
    )


# ============================================================
# SELEÇÃO DO RELATÓRIO
# ============================================================
def selecionar_relatorio_apex(page, relatorio):
    inicio_selecao_perf = time.perf_counter()

    select_id = seletor_select_relatorio(relatorio)
    botao_ir_id = seletor_botao_ir_relatorio(relatorio)

    print("=" * 100)
    print(f"[RELATÓRIO] Selecionando: {relatorio['nome']}")
    print(f"[RELATÓRIO] APEX ID: {relatorio['apex_id']}")
    print(f"[RELATÓRIO] VALUE salvo: {relatorio['saved_value']}")
    print(f"[RELATÓRIO] Responsável no download: {relatorio.get('responsavel') or 'não definido'}")
    print("=" * 100)

    print(f"[14] Aguardando select do relatório APEX {relatorio['apex_id']}...")
    print(f"[INFO] Select esperado: {select_id}")

    select_relatorio = page.locator(select_id)
    select_relatorio.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print(f"[15] Selecionando relatório salvo pelo VALUE real: {relatorio['saved_value']}")

    try:
        select_relatorio.select_option(value=relatorio["saved_value"], timeout=TIMEOUT_PADRAO)
    except Exception:
        print("[ERRO] Não consegui selecionar o relatório salvo pelo VALUE informado.")
        print("[DIAGNÓSTICO] Opções encontradas no combo:")
        for opcao in listar_opcoes_relatorio(select_relatorio):
            print(opcao)
        raise

    page.wait_for_timeout(3000)

    print(f"[16] Aguardando botão Ir do relatório APEX {relatorio['apex_id']}...")
    print(f"[INFO] Botão Ir esperado: {botao_ir_id}")

    botao_ir = page.locator(botao_ir_id)

    try:
        botao_ir.wait_for(state="visible", timeout=15000)
    except Exception:
        print("[AVISO] Não encontrei botão Ir pelo ID exato. Tentando pelo botão Ir visível da tela.")
        botao_ir = page.get_by_role("button", name="Ir", exact=True)
        botao_ir.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1500)

    print(f"[17] Clicando no botão Ir do relatório APEX {relatorio['apex_id']}...")
    inicio_carga_apex_perf = time.perf_counter()
    botao_ir.click()

    page.wait_for_timeout(7000)
    aguardar_carregamento(page)
    aguardar_relatorio_pronto(page)

    tempo_carga_apex = segundos_decorridos(inicio_carga_apex_perf)
    registrar_metrica_execucao(
        etapa="CARGA_RELATORIO_APEX",
        relatorio=relatorio,
        status="OK",
        duracao_segundos=tempo_carga_apex,
        detalhes=f"Tempo após clicar em Ir até o relatório ficar pronto. Tempo total da seleção: {segundos_decorridos(inicio_selecao_perf)}s.",
    )

    print("[OK] Relatório selecionado e botão Ir executado.")
    print(f"[INFO] URL atual: {page.url}")


# ============================================================
# MODAIS / JANELAS
# ============================================================
def fechar_modal_se_existir(page):
    tentativas = [
        lambda: page.get_by_role("dialog", name="Execução de Teste").get_by_label("Fechar"),
        lambda: page.get_by_role("dialog", name="Cenário e Execuções").get_by_label("Fechar"),
        lambda: page.get_by_role("button", name="Fechar"),
        lambda: page.locator("button[title='Fechar']"),
        lambda: page.locator("button[aria-label='Fechar']"),
        lambda: page.locator(".ui-dialog-titlebar-close"),
    ]

    for criar_locator in tentativas:
        try:
            locator = criar_locator().first
            if locator.count() > 0:
                locator.click(timeout=3000)
                page.wait_for_timeout(1000)
        except Exception:
            pass

    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(500)
    except Exception:
        pass

    aguardar_processamento_apex(page)


# ============================================================
# DOWNLOAD DE UMA EXECUÇÃO
# ============================================================
def localizar_link_edit(frame_cenario, origem=""):
    """
    Localiza um link Edit visível com timeout total controlado.

    Correção importante:
    - A versão anterior com busca DOM convertia índice 0 para -1 quando o Edit era o primeiro link.
      Isso gerava falso "sem Edit" e criava placeholder indevido.
    - Esta versão volta a usar os seletores Playwright originais como busca principal.
    - A busca DOM fica apenas como fallback e preserva corretamente o índice 0.
    """
    tentativas = [
        ("role link name=Edit", lambda: frame_cenario.get_by_role("link", name="Edit")),
        ("role link regex Edit", lambda: frame_cenario.get_by_role("link", name=re.compile(r"^\s*Edit\s*$", re.I))),
        ("anchor text Edit", lambda: frame_cenario.locator("a").filter(has_text=re.compile(r"^\s*Edit\s*$", re.I))),
        ("anchor title/aria Edit", lambda: frame_cenario.locator("a[title*='Edit'], a[aria-label*='Edit']")),
        ("anchor href javascript", lambda: frame_cenario.locator("a[href*='javascript']").filter(has_text=re.compile(r"Edit", re.I))),
    ]

    limite_config_ms = max(int(TIMEOUT_EDIT_EXECUCAO or 20000), 1000)
    limite_maximo_ms = max(int(globals().get("TIMEOUT_EDIT_EXECUCAO_MAXIMO_REAL", 30000) or 30000), 1000)
    limite_ms = min(limite_config_ms, limite_maximo_ms)
    limite_segundos = max(limite_ms / 1000, 1)

    inicio = time.perf_counter()
    proximo_log = max(int(INTERVALO_LOG_EDIT_SEGUNDOS or 5), 1)
    ultimo_erro = None

    while True:
        decorrido = time.perf_counter() - inicio

        if decorrido >= limite_segundos:
            detalhe = f"timeout total de {limite_ms}ms atingido"
            if ultimo_erro:
                detalhe += f" | último erro: {ultimo_erro}"
            return None, 0, detalhe

        if decorrido >= proximo_log:
            print(
                f"[AGUARDANDO EDIT] {origem}: ainda procurando link Edit "
                f"({int(decorrido)}s/{int(limite_segundos)}s)..."
            )
            proximo_log += max(int(INTERVALO_LOG_EDIT_SEGUNDOS or 5), 1)

        # 1) Busca principal: seletores Playwright originais, mais confiáveis para iframe/modal APEX.
        for descricao, criar_locator in tentativas:
            try:
                locator = criar_locator()
                qtd = locator.count()

                if qtd <= 0:
                    continue

                for indice in range(qtd):
                    candidato = locator.nth(indice)
                    try:
                        if candidato.is_visible(timeout=300):
                            return candidato, qtd, descricao
                    except Exception as e:
                        ultimo_erro = e
                        continue

            except Exception as e:
                ultimo_erro = e
                continue

        # 2) Fallback DOM: procura anchors cujo texto/title/aria/href indiquem Edit.
        # Preserva corretamente o índice 0. Não usar "or -1", porque 0 é índice válido.
        try:
            resultado = frame_cenario.evaluate("""
                () => {
                    const limpar = txt => (txt || '').replace(/\s+/g, ' ').trim();
                    const visivel = el => {
                        if (!el) return false;
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        return style && style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
                    };

                    const anchors = Array.from(document.querySelectorAll('a'));
                    const candidatos = [];

                    anchors.forEach((a, indice) => {
                        const texto = limpar(a.innerText || a.textContent || '');
                        const title = limpar(a.getAttribute('title') || '');
                        const aria = limpar(a.getAttribute('aria-label') || '');
                        const href = limpar(a.getAttribute('href') || '');
                        const combinado = `${texto} ${title} ${aria} ${href}`;

                        const ehEdit =
                            /^Edit$/i.test(texto) ||
                            /^Edit$/i.test(title) ||
                            /^Edit$/i.test(aria) ||
                            /Edit/i.test(combinado);

                        if (ehEdit && visivel(a)) {
                            candidatos.push({ indice, texto, title, aria });
                        }
                    });

                    return {
                        quantidade: candidatos.length,
                        primeiroIndice: candidatos.length ? candidatos[0].indice : -1,
                        primeiroTexto: candidatos.length ? (candidatos[0].texto || candidatos[0].title || candidatos[0].aria || '') : ''
                    };
                }
            """)

            qtd_dom = int((resultado or {}).get("quantidade") or 0)
            primeiro_indice = (resultado or {}).get("primeiroIndice")
            indice_dom = int(primeiro_indice) if primeiro_indice is not None else -1

            if qtd_dom > 0 and indice_dom >= 0:
                candidato = frame_cenario.locator("a").nth(indice_dom)
                try:
                    if candidato.is_visible(timeout=300):
                        return candidato, qtd_dom, "fallback DOM por link Edit"
                except Exception as e:
                    ultimo_erro = e

        except Exception as e:
            ultimo_erro = e

        time.sleep(0.35)


def baixar_execucao_aberta(page, origem, numero_pagina, relatorio, dados_cenario=None, inicio_linha_perf=None):
    if inicio_linha_perf is None:
        inicio_linha_perf = time.perf_counter()

    frame_cenario = obter_frame(page, "Cenário e Execuções")

    # Respiro curto para o conteúdo interno do modal terminar de renderizar antes de procurar o Edit.
    # Isso reduz falso "sem Edit" em telas lentas.
    try:
        page.wait_for_timeout(1200)
        aguardar_processamento_apex(page)
    except Exception:
        pass

    print(f"[VALIDAÇÃO] Verificando link Edit para {origem}...")
    inicio_busca_edit_perf = time.perf_counter()
    link_edit, qtd_edit, estrategia_edit = localizar_link_edit(frame_cenario, origem)
    tempo_busca_edit = segundos_decorridos(inicio_busca_edit_perf)

    if not link_edit or qtd_edit <= 0:
        print(f"[SEM EDIT] {origem} não possui link Edit visível dentro do limite configurado.")
        print(f"[DIAGNÓSTICO] Última estratégia/erro: {estrategia_edit}")
        inicio_placeholder_perf = time.perf_counter()
        arquivo_placeholder = criar_download_nao_iniciado(page, origem, numero_pagina, relatorio, dados_cenario)
        tempo_placeholder = segundos_decorridos(inicio_placeholder_perf)
        tempo_total_linha = segundos_decorridos(inicio_linha_perf)
        registrar_metrica_execucao(
            etapa="LINHA_SEM_EDIT_PLACEHOLDER",
            relatorio=relatorio,
            numero_pagina=numero_pagina,
            origem=origem,
            status="PLACEHOLDER_GERADO" if arquivo_placeholder else "SEM_EDIT_SEM_ARQUIVO",
            duracao_segundos=tempo_total_linha,
            detalhes=f"Busca Edit: {tempo_busca_edit}s. Geração placeholder: {tempo_placeholder}s. Estratégia/diagnóstico: {estrategia_edit}.",
            arquivo=arquivo_placeholder or "",
        )
        fechar_modal_se_existir(page)
        return False

    print(
        f"[VALIDAÇÃO] {origem} possui {qtd_edit} link(s) Edit. "
        f"Estratégia: {estrategia_edit}. Usando o primeiro Edit visível encontrado."
    )

    print(f"[DOWNLOAD] Clicando no Edit visível para {origem}...")
    link_edit.click(timeout=TIMEOUT_PADRAO)
    page.wait_for_timeout(2500)
    aguardar_processamento_apex(page)

    frame_execucao = obter_frame(page, "Execução de Teste")

    print(f"[DOWNLOAD] Abrindo Ações para {origem}...")
    frame_execucao.get_by_role("button", name="Ações", exact=True).click(timeout=TIMEOUT_PADRAO)
    page.wait_for_timeout(1200)

    print(f"[DOWNLOAD] Clicando em Fazer Download no menu para {origem}...")
    try:
        frame_execucao.get_by_role("menuitem", name="Fazer Download").click(timeout=5000)
    except Exception:
        frame_execucao.locator("span").filter(has_text="Fazer Download").click(timeout=5000)

    page.wait_for_timeout(1200)

    print(f"[DOWNLOAD] Confirmando Fazer Download para {origem}...")
    inicio_download_perf = time.perf_counter()
    with page.expect_download(timeout=TIMEOUT_DOWNLOAD) as download_info:
        frame_execucao.get_by_role("button", name="Fazer Download").click(timeout=TIMEOUT_PADRAO)

    download = download_info.value
    tempo_download_bruto = segundos_decorridos(inicio_download_perf)

    pasta_relatorio = pasta_download_relatorio(relatorio)
    nome_original = limpar_nome_arquivo(download.suggested_filename)
    responsavel_download = prefixo_responsavel_download(relatorio)

    if responsavel_download:
        nome_final = f"pagina_{numero_pagina}_{responsavel_download}_{limpar_nome_arquivo(origem)}_{nome_original}"
    else:
        nome_final = f"pagina_{numero_pagina}_{limpar_nome_arquivo(origem)}_{nome_original}"

    destino = caminho_unico(pasta_relatorio / nome_final)

    download.save_as(str(destino))
    print(f"[OK] Download salvo em: {destino}")

    enriquecer_arquivo_download_com_cenario(destino, dados_cenario, origem, relatorio)

    registrar_modelo_download(relatorio, destino)
    registrar_arquivo_para_consolidacao(relatorio, destino, "download")

    tempo_total_linha = segundos_decorridos(inicio_linha_perf)
    registrar_metrica_execucao(
        etapa="LINHA_COM_EDIT_DOWNLOAD",
        relatorio=relatorio,
        numero_pagina=numero_pagina,
        origem=origem,
        status="DOWNLOAD_OK",
        duracao_segundos=tempo_total_linha,
        detalhes=f"Busca Edit: {tempo_busca_edit}s. Download bruto: {tempo_download_bruto}s. Estratégia Edit: {estrategia_edit}. Links Edit encontrados: {qtd_edit}.",
        arquivo=destino,
    )

    fechar_modal_se_existir(page)
    page.wait_for_timeout(1500)
    return True

def baixar_execucao_por_indice(page, indice, numero_pagina, relatorio):
    origem = f"editar_indice_{indice}"
    inicio_linha_perf = time.perf_counter()

    try:
        print(f"[PÁGINA {numero_pagina}] Processando {origem}...")

        celulas_editar = page.get_by_role("cell", name="Editar")
        locator = celulas_editar.first if indice == 0 else celulas_editar.nth(indice)

        locator.wait_for(state="visible", timeout=TIMEOUT_LINHA_CURTO)
        dados_cenario = extrair_dados_cenario_da_linha(locator, origem)
        locator.click(timeout=TIMEOUT_PADRAO)

        page.wait_for_timeout(2500)
        aguardar_processamento_apex(page)

        return baixar_execucao_aberta(page, origem, numero_pagina, relatorio, dados_cenario, inicio_linha_perf)

    except Exception as e:
        registrar_metrica_execucao(
            etapa="LINHA_EDITAR_ERRO",
            relatorio=relatorio,
            numero_pagina=numero_pagina,
            origem=origem,
            status="ERRO",
            duracao_segundos=segundos_decorridos(inicio_linha_perf),
            detalhes=str(e),
        )
        print(f"[ERRO] Falha em {origem}. Fechando janelas e seguindo.")
        print(f"[DETALHE] {e}")
        fechar_modal_se_existir(page)
        page.wait_for_timeout(1000)
        return False


def baixar_execucao_por_tr_child(page, linha_child, numero_pagina, relatorio):
    origem = f"tr_child_{linha_child}"
    inicio_linha_perf = time.perf_counter()

    try:
        print(f"[PÁGINA {numero_pagina}] Processando tr:nth-child({linha_child})...")

        linha_link = page.locator(f"tr:nth-child({linha_child}) > .a-IRR-linkCol").first
        linha_link.wait_for(state="visible", timeout=TIMEOUT_LINHA_CURTO)
        dados_cenario = extrair_dados_cenario_da_linha(linha_link, origem)

        linha_link.click(timeout=TIMEOUT_PADRAO)
        page.wait_for_timeout(2500)
        aguardar_processamento_apex(page)

        return baixar_execucao_aberta(page, origem, numero_pagina, relatorio, dados_cenario, inicio_linha_perf)

    except PlaywrightTimeoutError:
        registrar_metrica_execucao(
            etapa="LINHA_TR_NAO_ENCONTRADA",
            relatorio=relatorio,
            numero_pagina=numero_pagina,
            origem=origem,
            status="NAO_ENCONTRADA",
            duracao_segundos=segundos_decorridos(inicio_linha_perf),
            detalhes=f"tr:nth-child({linha_child}) não encontrado/visível.",
        )
        print(f"[SEM LINHA] tr:nth-child({linha_child}) não encontrado/visível.")
        return None

    except Exception as e:
        registrar_metrica_execucao(
            etapa="LINHA_TR_ERRO",
            relatorio=relatorio,
            numero_pagina=numero_pagina,
            origem=origem,
            status="ERRO",
            duracao_segundos=segundos_decorridos(inicio_linha_perf),
            detalhes=str(e),
        )
        print(f"[ERRO] Falha no tr:nth-child({linha_child}). Fechando janelas e seguindo.")
        print(f"[DETALHE] {e}")
        fechar_modal_se_existir(page)
        page.wait_for_timeout(1000)
        return False


# ============================================================
# PROCESSAMENTO DE PÁGINA / PAGINAÇÃO
# ============================================================
def processar_pagina_atual(page, numero_pagina, relatorio):
    print("=" * 100)
    print(f"[RELATÓRIO: {relatorio['nome']}] [PÁGINA {numero_pagina}] INICIANDO PROCESSAMENTO")
    print("=" * 100)

    inicio_pagina_pronta_perf = time.perf_counter()
    aguardar_relatorio_pronto(page)
    registrar_metrica_execucao(
        etapa="APEX_PAGINA_PRONTA",
        relatorio=relatorio,
        numero_pagina=numero_pagina,
        origem=f"pagina_{numero_pagina}",
        status="OK",
        duracao_segundos=segundos_decorridos(inicio_pagina_pronta_perf),
        detalhes="Tempo para confirmar que a página atual do relatório está pronta para processar linhas.",
    )

    downloads_ok = 0
    sem_edit_ou_erro = 0
    linhas_inexistentes = 0

    editar_inicio = relatorio["editar_inicio"]
    editar_fim = relatorio["editar_fim"]
    tr_inicio = relatorio["tr_inicio"]
    tr_fim = relatorio["tr_fim"]

    if editar_inicio >= 0 and editar_fim >= editar_inicio:
        print(f"[PÁGINA {numero_pagina}] Processando Editar índice {editar_inicio} até {editar_fim}...")

        for indice in range(editar_inicio, editar_fim + 1):
            resultado = baixar_execucao_por_indice(page, indice, numero_pagina, relatorio)

            if resultado is True:
                downloads_ok += 1
            else:
                sem_edit_ou_erro += 1

        print(
            f"[PÁGINA {numero_pagina}] Bloco Editar finalizado. "
            f"Downloads OK: {downloads_ok}. Sem Edit/erro: {sem_edit_ou_erro}."
        )
    else:
        print(f"[PÁGINA {numero_pagina}] Bloco Editar ignorado por configuração do .env.")

    if tr_inicio >= 0 and tr_fim >= tr_inicio:
        print(f"[PÁGINA {numero_pagina}] Processando tr:nth-child({tr_inicio}) até tr:nth-child({tr_fim})...")

        for linha_child in range(tr_inicio, tr_fim + 1):
            resultado = baixar_execucao_por_tr_child(page, linha_child, numero_pagina, relatorio)

            if resultado is True:
                downloads_ok += 1
                linhas_inexistentes = 0
            elif resultado is None:
                linhas_inexistentes += 1

                if linhas_inexistentes >= MAX_LINHAS_INEXISTENTES_SEGUIDAS:
                    print(
                        f"[PÁGINA {numero_pagina}] {linhas_inexistentes} linhas seguidas não encontradas. "
                        "Encerrando varredura desta página."
                    )
                    break
            else:
                sem_edit_ou_erro += 1
                linhas_inexistentes = 0

        print(
            f"[PÁGINA {numero_pagina}] Bloco tr:nth-child finalizado. "
            f"Downloads OK: {downloads_ok}. Sem Edit/erro: {sem_edit_ou_erro}."
        )
    else:
        print(f"[PÁGINA {numero_pagina}] Bloco tr:nth-child ignorado por configuração do .env.")

    return {
        "downloads_ok": downloads_ok,
        "sem_edit_ou_erro": sem_edit_ou_erro,
    }


def clicar_proximo_se_existir(page, numero_pagina, relatorio=None):
    print(f"[PÁGINA {numero_pagina}] Procurando botão Próximo...")

    candidatos = [
        lambda: page.get_by_role("button", name="Próximo").first,
        lambda: page.get_by_role("link", name="Próximo").first,
        lambda: page.locator("a[aria-label='Próximo']").first,
        lambda: page.locator("button[aria-label='Próximo']").first,
    ]

    for criar_locator in candidatos:
        try:
            botao_proximo = criar_locator()
            botao_proximo.wait_for(state="visible", timeout=5000)

            try:
                if not botao_proximo.is_enabled(timeout=2000):
                    print(f"[PÁGINA {numero_pagina}] Próximo encontrado, mas desabilitado. Fim.")
                    return False
            except Exception:
                pass

            print(f"[PÁGINA {numero_pagina}] Clicando em Próximo...")
            inicio_proxima_pagina_perf = time.perf_counter()
            botao_proximo.click(timeout=TIMEOUT_PADRAO)

            print(f"[PÁGINA {numero_pagina}] Aguardando próxima página carregar...")
            page.wait_for_timeout(8000)
            aguardar_carregamento(page)
            aguardar_relatorio_pronto(page)
            page.wait_for_timeout(3000)

            registrar_metrica_execucao(
                etapa="CARGA_PROXIMA_PAGINA_APEX",
                relatorio=relatorio,
                numero_pagina=numero_pagina + 1,
                origem=f"pagina_{numero_pagina}_para_{numero_pagina + 1}",
                status="OK",
                duracao_segundos=segundos_decorridos(inicio_proxima_pagina_perf),
                detalhes="Tempo após clicar em Próximo até a próxima página ficar pronta.",
            )

            print(f"[PÁGINA {numero_pagina}] Próxima página carregada. URL atual: {page.url}")
            return True

        except Exception:
            continue

    print(f"[PÁGINA {numero_pagina}] Não encontrei botão Próximo disponível. Processo encerrado para este relatório.")
    return False


def processar_todas_as_paginas(page, relatorio):
    print("=" * 100)
    print(f"[INÍCIO] Processamento do relatório: {relatorio['nome']}")
    print("=" * 100)

    total_paginas = 0
    total_downloads_ok = 0
    total_sem_edit_ou_erro = 0

    numero_pagina = 1

    while True:
        resumo = processar_pagina_atual(page, numero_pagina, relatorio)

        total_paginas += 1
        total_downloads_ok += resumo["downloads_ok"]
        total_sem_edit_ou_erro += resumo["sem_edit_ou_erro"]

        conseguiu_ir_proxima = clicar_proximo_se_existir(page, numero_pagina, relatorio)

        if not conseguiu_ir_proxima:
            print(f"[FIM] Relatório {relatorio['nome']} finalizado.")
            break

        numero_pagina += 1

    print("=" * 100)
    print(f"[RESUMO DO RELATÓRIO] {relatorio['nome']}")
    print(f"Total de páginas processadas: {total_paginas}")
    print(f"Total de downloads OK: {total_downloads_ok}")
    print(f"Total sem Edit/erro/pulados: {total_sem_edit_ou_erro}")
    print(f"Pasta de downloads do líder de cenário: {pasta_download_relatorio(relatorio)}")
    print("=" * 100)

    return {
        "relatorio": relatorio["nome"],
        "paginas": total_paginas,
        "downloads_ok": total_downloads_ok,
        "sem_edit_ou_erro": total_sem_edit_ou_erro,
    }


# ============================================================
# FLUXO PRINCIPAL
# ============================================================
def run():
    validar_env()

    inicio_execucao = datetime.now()
    cronometro = CronometroExecucao(inicio_execucao)
    cronometro.iniciar()

    print("=" * 100)
    print(f"[TEMPO] Início da execução: {inicio_execucao.strftime('%d/%m/%Y %H:%M:%S')}")
    print("[CONFIGURAÇÃO] Relatórios carregados do .env")
    print(f"[CONFIGURAÇÃO] Total de relatórios carregados: {len(RELATORIOS)}")
    if len(RELATORIOS) <= 1:
        print("[AVISO][CONFIGURAÇÃO] Apenas 1 relatório foi carregado. Verifique se o .env usado é o correto e se RELATORIO_02, RELATORIO_03... existem nele.")

    for idx, relatorio in enumerate(RELATORIOS, start=1):
        print(
            f"{idx}. {relatorio['nome']} | "
            f"APEX={relatorio['apex_id']} | "
            f"VALUE={relatorio['saved_value']} | "
            f"Responsável={relatorio.get('responsavel') or 'não definido'} | "
            f"Editar={relatorio['editar_inicio']}..{relatorio['editar_fim']} | "
            f"TR={relatorio['tr_inicio']}..{relatorio['tr_fim']}"
        )
    print("=" * 100)

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=False,
            slow_mo=300,
        )

        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(TIMEOUT_PADRAO)

        resumos = []
        consolidado_final = None

        try:
            cronometro.atualizar_status("Login e abertura da tela de Execução de Testes")
            login_e_abrir_execucao_testes(page)

            for indice, relatorio in enumerate(RELATORIOS, start=1):
                cronometro.atualizar_status(
                    f"Processando {relatorio['nome']} ({indice}/{len(RELATORIOS)})"
                )

                print("=" * 100)
                print(f"[LOOP] Iniciando relatório {indice}/{len(RELATORIOS)}: {relatorio['nome']}")
                print("=" * 100)

                try:
                    selecionar_relatorio_apex(page, relatorio)
                    resumo = processar_todas_as_paginas(page, relatorio)
                    resumo["status"] = "OK"
                    resumo["erro"] = ""
                    resumo["responsavel"] = relatorio.get("responsavel", "")
                    resumos.append(resumo)

                    print(f"[LOOP] Relatório finalizado: {relatorio['nome']}")
                    if indice < len(RELATORIOS):
                        proximo = RELATORIOS[indice]
                        print(f"[LOOP] Próximo relatório na fila: {proximo['nome']}")

                except Exception as e:
                    erro = str(e)
                    print("=" * 100)
                    print(f"[RELATÓRIO][ERRO] Falha no relatório {relatorio['nome']}. Vou registrar no resumo e continuar para o próximo.")
                    print(f"[DETALHE] {erro}")
                    print("=" * 100)

                    resumos.append(
                        {
                            "relatorio": relatorio.get("nome", ""),
                            "responsavel": relatorio.get("responsavel", ""),
                            "paginas": 0,
                            "downloads_ok": 0,
                            "sem_edit_ou_erro": 0,
                            "status": "ERRO",
                            "erro": erro,
                        }
                    )

                    try:
                        fechar_modal_se_existir(page)
                    except Exception:
                        pass

                try:
                    fechar_modal_se_existir(page)
                except Exception:
                    pass

                page.wait_for_timeout(3000)
                aguardar_carregamento(page)

            print("=" * 100)
            print("[RESUMO GERAL]")
            for resumo in resumos:
                print(
                    f"- {resumo['relatorio']}: "
                    f"status={resumo.get('status', 'OK')}, "
                    f"páginas={resumo.get('paginas', 0)}, "
                    f"downloads_ok={resumo.get('downloads_ok', 0)}, "
                    f"sem_edit_ou_erro={resumo.get('sem_edit_ou_erro', 0)}"
                )
                if resumo.get("erro"):
                    print(f"  erro={resumo.get('erro')}")
            print(f"Pasta raiz dos downloads: {DOWNLOAD_ROOT}")
            print(f"Métricas de tempo capturadas: {len(METRICAS_EXECUCAO)}")
            print("=" * 100)

            try:
                cronometro.atualizar_status("Gerando consolidado final e aba RESUMO_EXECUCAO")
                consolidado_final = consolidar_arquivos_download_execucao(
                    resumos_execucao=resumos,
                    inicio_execucao=inicio_execucao,
                )
            except Exception as e:
                print("[CONSOLIDADO][ERRO] Não consegui gerar o consolidado final.")
                print(f"[DETALHE] {e}")

            fim_execucao = datetime.now()
            print("=" * 100)
            print(f"[TEMPO] Fim da execução: {fim_execucao.strftime('%d/%m/%Y %H:%M:%S')}")
            print(f"[TEMPO] Duração total: {formatar_duracao((fim_execucao - inicio_execucao).total_seconds())}")
            if consolidado_final:
                print(f"[TEMPO] Resumo gravado no consolidado: {consolidado_final}")
            print("=" * 100)

            cronometro.atualizar_status("Finalizado. Consolidado gerado.")
            cronometro.finalizar()

            if MANTER_NAVEGADOR_ABERTO:
                print("[INFO] O navegador ficará aberto para validação visual.")
                input("Pressione ENTER aqui no terminal para fechar o navegador...")
            else:
                print("[INFO] Navegador será fechado automaticamente. Processo concluído.")

        finally:
            cronometro.finalizar()
            context.close()
            browser.close()


if __name__ == "__main__":
    try:
        run()
    except PlaywrightTimeoutError as e:
        print("[ERRO] Timeout no Playwright.")
        print(e)
        raise
    except Exception as e:
        print("[ERRO] Falha geral.")
        print(e)
        raise
