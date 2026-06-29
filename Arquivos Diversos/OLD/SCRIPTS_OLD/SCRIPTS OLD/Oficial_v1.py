import os
import re
from pathlib import Path

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

GTN_USUARIO = os.getenv("GTN_USUARIO") or os.getenv("GTN_USER")
GTN_SENHA = os.getenv("GTN_SENHA") or os.getenv("GTN_PASS")

URL_LOGIN = os.getenv(
    "URL_LOGIN",
    "https://gtn.ninecon.com.br/ords/r/gtn/gtn/login?session=601689329474445&tz=-3:00",
)

TIMEOUT_PADRAO = int(os.getenv("TIMEOUT_PADRAO", "60000"))
TIMEOUT_DOWNLOAD = int(os.getenv("TIMEOUT_DOWNLOAD", "180000"))
TIMEOUT_LINHA_CURTO = int(os.getenv("TIMEOUT_LINHA_CURTO", "5000"))
MAX_LINHAS_INEXISTENTES_SEGUIDAS = int(os.getenv("MAX_LINHAS_INEXISTENTES_SEGUIDAS", "10"))

DOWNLOAD_ROOT = BASE_DIR / os.getenv("DOWNLOAD_DIR", "downloads_gtn")

GERAR_PLACEHOLDER_SEM_EDIT = os.getenv("GERAR_PLACEHOLDER_SEM_EDIT", "S").strip().upper() in {"S", "SIM", "TRUE", "1", "YES"}
TEMPLATE_NAO_INICIADO = os.getenv("TEMPLATE_NAO_INICIADO", "").strip()

# Guarda o primeiro arquivo baixado com sucesso por relatório para usar como modelo
# quando algum cenário não tiver link Edit.
MODELOS_DOWNLOAD_POR_RELATORIO = {}


# ============================================================
# RELATÓRIOS PELO .ENV
#
# Formato da variável RELATORIOS:
# nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim
#
# Exemplo:
# RELATORIOS=PMO_EXECUCAO_TESTES|35932200234408468|96461147945312616|0|5|9|1002;OUTRO_RELATORIO|111|222|0|5|9|1002
# ============================================================
def carregar_relatorios_env():
    relatorios_raw = os.getenv("RELATORIOS", "").strip()

    # Compatibilidade com o modelo antigo, caso você ainda deixe as variáveis separadas no .env.
    if not relatorios_raw:
        apex_id = os.getenv("RELATORIO_APEX_ID", "").strip()
        saved_value = os.getenv("RELATORIO_SALVO_VALUE", "").strip()

        if apex_id and saved_value:
            return [
                {
                    "nome": os.getenv("RELATORIO_NOME", "PMO_EXECUCAO_TESTES").strip(),
                    "apex_id": apex_id,
                    "saved_value": saved_value,
                    "editar_inicio": int(os.getenv("EDITAR_INDICE_INICIO", "0")),
                    "editar_fim": int(os.getenv("EDITAR_INDICE_FIM", "5")),
                    "tr_inicio": int(os.getenv("TR_CHILD_INICIO", "9")),
                    "tr_fim": int(os.getenv("TR_CHILD_FIM", "1002")),
                }
            ]

        raise Exception(
            "Nenhum relatório configurado. Informe RELATORIOS no .env no formato: "
            "nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim"
        )

    relatorios = []

    for posicao, item in enumerate(relatorios_raw.split(";"), start=1):
        item = item.strip()

        if not item:
            continue

        partes = [parte.strip() for parte in item.split("|")]

        if len(partes) != 7:
            raise Exception(
                f"Formato inválido no relatório {posicao} da variável RELATORIOS. "
                "Use: nome|apex_id|saved_value|editar_inicio|editar_fim|tr_inicio|tr_fim"
            )

        nome, apex_id, saved_value, editar_inicio, editar_fim, tr_inicio, tr_fim = partes

        if not nome:
            raise Exception(f"Nome vazio no relatório {posicao} da variável RELATORIOS.")

        if not apex_id:
            raise Exception(f"apex_id vazio no relatório {posicao} da variável RELATORIOS.")

        if not saved_value:
            raise Exception(f"saved_value vazio no relatório {posicao} da variável RELATORIOS.")

        relatorios.append(
            {
                "nome": nome,
                "apex_id": apex_id,
                "saved_value": saved_value,
                "editar_inicio": int(editar_inicio),
                "editar_fim": int(editar_fim),
                "tr_inicio": int(tr_inicio),
                "tr_fim": int(tr_fim),
            }
        )

    if not relatorios:
        raise Exception("A variável RELATORIOS está vazia ou inválida no .env.")

    return relatorios


RELATORIOS = carregar_relatorios_env()


# ============================================================
# UTILITÁRIOS
# ============================================================
def limpar_nome_arquivo(nome):
    nome = nome or "download"
    nome = re.sub(r'[\\/:*?"<>|]+', "_", nome)
    nome = re.sub(r"\s+", "_", nome).strip("._ ")
    return nome or "download"


def seletor_select_relatorio(relatorio):
    return f"#R{relatorio['apex_id']}_saved_reports"


def seletor_botao_ir_relatorio(relatorio):
    return f"#R{relatorio['apex_id']}_saved_reports_go"


def pasta_download_relatorio(relatorio):
    pasta = DOWNLOAD_ROOT / limpar_nome_arquivo(relatorio["nome"])
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta


def caminho_unico(caminho):
    if not caminho.exists():
        return caminho

    base = caminho.stem
    ext = caminho.suffix
    pasta = caminho.parent

    contador = 2
    while True:
        novo = pasta / f"{base}_{contador}{ext}"
        if not novo.exists():
            return novo
        contador += 1


def resolver_template_nao_iniciado(relatorio):
    """
    Retorna um arquivo modelo para criar o placeholder de cenário sem Edit.

    Prioridade:
    1) TEMPLATE_NAO_INICIADO informado no .env
    2) Primeiro download real já salvo para o relatório atual
    3) Nenhum modelo encontrado
    """
    if TEMPLATE_NAO_INICIADO:
        template = Path(TEMPLATE_NAO_INICIADO)
        if not template.is_absolute():
            template = BASE_DIR / template

        if template.exists():
            return template

        print(f"[AVISO] TEMPLATE_NAO_INICIADO informado, mas não encontrado: {template}")

    return MODELOS_DOWNLOAD_POR_RELATORIO.get(relatorio["nome"])


def registrar_modelo_download(relatorio, caminho_download):
    """Guarda o primeiro download real do relatório como template para placeholders."""
    if relatorio["nome"] not in MODELOS_DOWNLOAD_POR_RELATORIO:
        MODELOS_DOWNLOAD_POR_RELATORIO[relatorio["nome"]] = Path(caminho_download)
        print(f"[TEMPLATE] Modelo de colunas definido para {relatorio['nome']}: {caminho_download}")


def criar_placeholder_csv(template_path, destino):
    """Cria CSV vazio preservando o cabeçalho do template, quando possível."""
    header = ""

    if template_path and Path(template_path).exists():
        with open(template_path, "r", encoding="utf-8-sig", errors="ignore") as origem:
            header = origem.readline().rstrip("\n")

    with open(destino, "w", encoding="utf-8-sig", newline="") as saida:
        if header:
            saida.write(header + "\n")


def criar_placeholder_xlsx(template_path, destino):
    """
    Cria XLSX vazio usando o primeiro download real como modelo.
    Mantém cabeçalho/estrutura e apaga linhas de dados a partir da linha 2.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise Exception(
            "Para criar placeholder .xlsx preservando colunas, instale openpyxl: pip install openpyxl"
        ) from e

    wb = load_workbook(template_path)

    for ws in wb.worksheets:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    wb.save(destino)


def criar_download_nao_iniciado(page, origem, numero_pagina, relatorio):
    """
    Cria arquivo placeholder quando o cenário existe, mas não possui link Edit.

    O arquivo usa o mesmo cabeçalho/colunas de um download real quando houver template.
    Se ainda não existir template, gera um CSV vazio para registrar o cenário não iniciado.
    """
    if not GERAR_PLACEHOLDER_SEM_EDIT:
        print("[PLACEHOLDER] Geração de placeholder desativada no .env.")
        return None

    pasta_relatorio = pasta_download_relatorio(relatorio)
    template = resolver_template_nao_iniciado(relatorio)

    nome_base = f"pagina_{numero_pagina}_{limpar_nome_arquivo(origem)}_Cenario_e_Execucao_nao_iniciado"

    if template and Path(template).exists():
        extensao = Path(template).suffix.lower() or ".csv"
    else:
        extensao = ".csv"

    destino = caminho_unico(pasta_relatorio / f"{nome_base}{extensao}")

    try:
        if template and Path(template).exists() and extensao in {".xlsx", ".xlsm"}:
            criar_placeholder_xlsx(template, destino)
        elif template and Path(template).exists() and extensao in {".csv", ".txt"}:
            criar_placeholder_csv(template, destino)
        else:
            # Sem template ainda: cria um marcador CSV vazio.
            # Assim o processo não perde o registro do cenário sem execução iniciada.
            with open(destino, "w", encoding="utf-8-sig", newline="") as saida:
                saida.write("")

        print(f"[PLACEHOLDER] Arquivo de cenário não iniciado criado: {destino}")
        return destino

    except Exception as e:
        print("[ERRO] Não consegui criar o arquivo placeholder de cenário não iniciado.")
        print(f"[DETALHE] {e}")
        return None


def validar_env():
    if not GTN_USUARIO:
        raise Exception("Não encontrei GTN_USUARIO ou GTN_USER no arquivo .env.")

    if not GTN_SENHA:
        raise Exception("Não encontrei GTN_SENHA ou GTN_PASS no arquivo .env.")

    if not URL_LOGIN:
        raise Exception("URL_LOGIN não pode ficar vazia no .env.")

    if not RELATORIOS:
        raise Exception("Nenhum relatório carregado do .env.")


def aguardar_processamento_apex(page):
    seletores = [
        ".u-Processing",
        ".a-Processing",
        ".apex_wait_overlay",
        ".ui-widget-overlay",
    ]

    for seletor in seletores:
        try:
            page.locator(seletor).first.wait_for(state="hidden", timeout=8000)
        except Exception:
            pass


def aguardar_carregamento(page):
    try:
        page.wait_for_load_state("domcontentloaded", timeout=20000)
    except Exception:
        pass

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass

    aguardar_processamento_apex(page)


def aguardar_relatorio_pronto(page):
    aguardar_processamento_apex(page)

    try:
        page.get_by_role("cell", name="Editar").first.wait_for(state="visible", timeout=30000)
        return True
    except Exception:
        pass

    try:
        page.locator("tr > .a-IRR-linkCol").first.wait_for(state="visible", timeout=30000)
        return True
    except Exception:
        pass

    return False


def listar_opcoes_relatorio(select_relatorio):
    try:
        return select_relatorio.evaluate(
            """
            el => Array.from(el.options).map(opt => ({
                value: opt.value,
                text: opt.textContent.trim(),
                selected: opt.selected
            }))
            """
        )
    except Exception as e:
        return [{"erro": str(e)}]


def obter_frame(page, titulo):
    iframe = page.locator(f'iframe[title="{titulo}"]').first
    iframe.wait_for(state="attached", timeout=TIMEOUT_PADRAO)

    try:
        iframe.wait_for(state="visible", timeout=TIMEOUT_PADRAO)
    except Exception:
        pass

    return iframe.content_frame


# ============================================================
# LOGIN E NAVEGAÇÃO
# ============================================================
def login_e_abrir_execucao_testes(page):
    print("[1] Abrindo tela de login...")
    page.goto(URL_LOGIN, wait_until="domcontentloaded")
    page.wait_for_timeout(2000)

    print("[2] Preenchendo usuário...")
    campo_usuario = page.get_by_role("textbox", name="Usuário")
    campo_usuario.wait_for(state="visible")
    campo_usuario.fill(GTN_USUARIO)

    page.wait_for_timeout(800)

    print("[3] Indo para senha...")
    campo_usuario.press("Tab")
    page.wait_for_timeout(800)

    print("[4] Preenchendo senha...")
    campo_senha = page.get_by_role("textbox", name="Senha")
    campo_senha.wait_for(state="visible")
    campo_senha.fill(GTN_SENHA)

    page.wait_for_timeout(1000)

    print("[5] Clicando em Acessar...")
    botao_acessar = page.get_by_role("button", name="Acessar")
    botao_acessar.wait_for(state="visible")
    botao_acessar.click()

    print("[6] Aguardando pós-login/renderização...")
    page.wait_for_timeout(5000)
    aguardar_carregamento(page)
    page.wait_for_timeout(3000)

    print(f"[INFO] URL atual depois do login: {page.url}")

    print("[7] Aguardando botão Navegação Principal...")
    botao_nav = page.get_by_role("button", name="Navegação Principal")
    botao_nav.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1500)

    print("[8] Clicando em Navegação Principal...")
    botao_nav.click()
    page.wait_for_timeout(3000)

    print("[9] Aguardando toggle da árvore...")
    toggle_arvore = page.locator(".a-TreeView-toggle:visible").first
    toggle_arvore.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print("[10] Clicando no toggle da árvore...")
    toggle_arvore.click()
    page.wait_for_timeout(3000)

    print("[11] Aguardando item Execução de Testes...")
    item_execucao = page.get_by_role("treeitem", name="Execução de Testes")
    item_execucao.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print("[12] Clicando em Execução de Testes...")
    item_execucao.click()
    page.wait_for_timeout(7000)
    aguardar_carregamento(page)

    print("[OK] Clique em Execução de Testes executado.")
    print(f"[INFO] URL atual após clique em Execução de Testes: {page.url}")

    print("[13] Mantendo a página atual de Execução de Testes.")
    print("[INFO] Não vou usar page.goto com session/cs fixos, porque isso pode voltar para o login.")
    print(f"[INFO] URL atual válida: {page.url}")

    page.wait_for_timeout(7000)


# ============================================================
# SELEÇÃO DO RELATÓRIO
# ============================================================
def selecionar_relatorio_apex(page, relatorio):
    select_id = seletor_select_relatorio(relatorio)
    botao_ir_id = seletor_botao_ir_relatorio(relatorio)

    print("=" * 100)
    print(f"[RELATÓRIO] Selecionando: {relatorio['nome']}")
    print(f"[RELATÓRIO] APEX ID: {relatorio['apex_id']}")
    print(f"[RELATÓRIO] VALUE salvo: {relatorio['saved_value']}")
    print("=" * 100)

    print(f"[14] Aguardando select do relatório APEX {relatorio['apex_id']}...")
    print(f"[INFO] Select esperado: {select_id}")

    select_relatorio = page.locator(select_id)
    select_relatorio.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1000)

    print(f"[15] Selecionando relatório salvo pelo VALUE real: {relatorio['saved_value']}")

    try:
        select_relatorio.select_option(value=relatorio["saved_value"], timeout=TIMEOUT_PADRAO)
    except Exception:
        print("[ERRO] Não consegui selecionar o relatório salvo pelo VALUE informado.")
        print("[DIAGNÓSTICO] Opções encontradas no combo:")
        for opcao in listar_opcoes_relatorio(select_relatorio):
            print(opcao)
        raise

    page.wait_for_timeout(3000)

    print(f"[16] Aguardando botão Ir do relatório APEX {relatorio['apex_id']}...")
    print(f"[INFO] Botão Ir esperado: {botao_ir_id}")

    botao_ir = page.locator(botao_ir_id)

    try:
        botao_ir.wait_for(state="visible", timeout=15000)
    except Exception:
        print("[AVISO] Não encontrei botão Ir pelo ID exato. Tentando pelo botão Ir visível da tela.")
        botao_ir = page.get_by_role("button", name="Ir", exact=True)
        botao_ir.wait_for(state="visible", timeout=TIMEOUT_PADRAO)

    page.wait_for_timeout(1500)

    print(f"[17] Clicando no botão Ir do relatório APEX {relatorio['apex_id']}...")
    botao_ir.click()

    page.wait_for_timeout(7000)
    aguardar_carregamento(page)
    aguardar_relatorio_pronto(page)

    print("[OK] Relatório selecionado e botão Ir executado.")
    print(f"[INFO] URL atual: {page.url}")


# ============================================================
# MODAIS / JANELAS
# ============================================================
def fechar_modal_se_existir(page):
    tentativas = [
        lambda: page.get_by_role("dialog", name="Execução de Teste").get_by_label("Fechar"),
        lambda: page.get_by_role("dialog", name="Cenário e Execuções").get_by_label("Fechar"),
        lambda: page.get_by_role("button", name="Fechar"),
        lambda: page.locator("button[title='Fechar']"),
        lambda: page.locator("button[aria-label='Fechar']"),
        lambda: page.locator(".ui-dialog-titlebar-close"),
    ]

    for criar_locator in tentativas:
        try:
            locator = criar_locator().first
            if locator.count() > 0:
                locator.click(timeout=3000)
                page.wait_for_timeout(1000)
        except Exception:
            pass

    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(500)
    except Exception:
        pass

    aguardar_processamento_apex(page)


# ============================================================
# DOWNLOAD DE UMA EXECUÇÃO
# ============================================================
def baixar_execucao_aberta(page, origem, numero_pagina, relatorio):
    frame_cenario = obter_frame(page, "Cenário e Execuções")

    print(f"[VALIDAÇÃO] Verificando link Edit para {origem}...")
    links_edit = frame_cenario.get_by_role("link", name="Edit")

    try:
        links_edit.first.wait_for(state="visible", timeout=5000)
    except Exception:
        print(f"[SEM EDIT] {origem} não possui link Edit.")
        criar_download_nao_iniciado(page, origem, numero_pagina, relatorio)
        fechar_modal_se_existir(page)
        return False

    qtd_edit = links_edit.count()
    print(f"[VALIDAÇÃO] {origem} possui {qtd_edit} link(s) Edit. Usando sempre o primeiro: first.")

    print(f"[DOWNLOAD] Clicando no primeiro Edit para {origem}...")
    links_edit.first.click(timeout=TIMEOUT_PADRAO)
    page.wait_for_timeout(2500)
    aguardar_processamento_apex(page)

    frame_execucao = obter_frame(page, "Execução de Teste")

    print(f"[DOWNLOAD] Abrindo Ações para {origem}...")
    frame_execucao.get_by_role("button", name="Ações", exact=True).click(timeout=TIMEOUT_PADRAO)
    page.wait_for_timeout(1200)

    print(f"[DOWNLOAD] Clicando em Fazer Download no menu para {origem}...")
    try:
        frame_execucao.get_by_role("menuitem", name="Fazer Download").click(timeout=5000)
    except Exception:
        frame_execucao.locator("span").filter(has_text="Fazer Download").click(timeout=5000)

    page.wait_for_timeout(1200)

    print(f"[DOWNLOAD] Confirmando Fazer Download para {origem}...")
    with page.expect_download(timeout=TIMEOUT_DOWNLOAD) as download_info:
        frame_execucao.get_by_role("button", name="Fazer Download").click(timeout=TIMEOUT_PADRAO)

    download = download_info.value

    pasta_relatorio = pasta_download_relatorio(relatorio)
    nome_original = limpar_nome_arquivo(download.suggested_filename)
    nome_final = f"pagina_{numero_pagina}_{limpar_nome_arquivo(origem)}_{nome_original}"
    destino = caminho_unico(pasta_relatorio / nome_final)

    download.save_as(str(destino))
    print(f"[OK] Download salvo em: {destino}")

    registrar_modelo_download(relatorio, destino)

    fechar_modal_se_existir(page)
    page.wait_for_timeout(1500)
    return True

def baixar_execucao_por_indice(page, indice, numero_pagina, relatorio):
    origem = f"editar_indice_{indice}"

    try:
        print(f"[PÁGINA {numero_pagina}] Processando {origem}...")

        celulas_editar = page.get_by_role("cell", name="Editar")
        locator = celulas_editar.first if indice == 0 else celulas_editar.nth(indice)

        locator.wait_for(state="visible", timeout=TIMEOUT_LINHA_CURTO)
        locator.click(timeout=TIMEOUT_PADRAO)

        page.wait_for_timeout(2500)
        aguardar_processamento_apex(page)

        return baixar_execucao_aberta(page, origem, numero_pagina, relatorio)

    except Exception as e:
        print(f"[ERRO] Falha em {origem}. Fechando janelas e seguindo.")
        print(f"[DETALHE] {e}")
        fechar_modal_se_existir(page)
        page.wait_for_timeout(1000)
        return False


def baixar_execucao_por_tr_child(page, linha_child, numero_pagina, relatorio):
    origem = f"tr_child_{linha_child}"

    try:
        print(f"[PÁGINA {numero_pagina}] Processando tr:nth-child({linha_child})...")

        linha_link = page.locator(f"tr:nth-child({linha_child}) > .a-IRR-linkCol").first
        linha_link.wait_for(state="visible", timeout=TIMEOUT_LINHA_CURTO)

        linha_link.click(timeout=TIMEOUT_PADRAO)
        page.wait_for_timeout(2500)
        aguardar_processamento_apex(page)

        return baixar_execucao_aberta(page, origem, numero_pagina, relatorio)

    except PlaywrightTimeoutError:
        print(f"[SEM LINHA] tr:nth-child({linha_child}) não encontrado/visível.")
        return None

    except Exception as e:
        print(f"[ERRO] Falha no tr:nth-child({linha_child}). Fechando janelas e seguindo.")
        print(f"[DETALHE] {e}")
        fechar_modal_se_existir(page)
        page.wait_for_timeout(1000)
        return False


# ============================================================
# PROCESSAMENTO DE PÁGINA / PAGINAÇÃO
# ============================================================
def processar_pagina_atual(page, numero_pagina, relatorio):
    print("=" * 100)
    print(f"[RELATÓRIO: {relatorio['nome']}] [PÁGINA {numero_pagina}] INICIANDO PROCESSAMENTO")
    print("=" * 100)

    aguardar_relatorio_pronto(page)

    downloads_ok = 0
    sem_edit_ou_erro = 0
    linhas_inexistentes = 0

    editar_inicio = relatorio["editar_inicio"]
    editar_fim = relatorio["editar_fim"]
    tr_inicio = relatorio["tr_inicio"]
    tr_fim = relatorio["tr_fim"]

    if editar_inicio >= 0 and editar_fim >= editar_inicio:
        print(f"[PÁGINA {numero_pagina}] Processando Editar índice {editar_inicio} até {editar_fim}...")

        for indice in range(editar_inicio, editar_fim + 1):
            resultado = baixar_execucao_por_indice(page, indice, numero_pagina, relatorio)

            if resultado is True:
                downloads_ok += 1
            else:
                sem_edit_ou_erro += 1

        print(
            f"[PÁGINA {numero_pagina}] Bloco Editar finalizado. "
            f"Downloads OK: {downloads_ok}. Sem Edit/erro: {sem_edit_ou_erro}."
        )
    else:
        print(f"[PÁGINA {numero_pagina}] Bloco Editar ignorado por configuração do .env.")

    if tr_inicio >= 0 and tr_fim >= tr_inicio:
        print(f"[PÁGINA {numero_pagina}] Processando tr:nth-child({tr_inicio}) até tr:nth-child({tr_fim})...")

        for linha_child in range(tr_inicio, tr_fim + 1):
            resultado = baixar_execucao_por_tr_child(page, linha_child, numero_pagina, relatorio)

            if resultado is True:
                downloads_ok += 1
                linhas_inexistentes = 0
            elif resultado is None:
                linhas_inexistentes += 1

                if linhas_inexistentes >= MAX_LINHAS_INEXISTENTES_SEGUIDAS:
                    print(
                        f"[PÁGINA {numero_pagina}] {linhas_inexistentes} linhas seguidas não encontradas. "
                        "Encerrando varredura desta página."
                    )
                    break
            else:
                sem_edit_ou_erro += 1
                linhas_inexistentes = 0

        print(
            f"[PÁGINA {numero_pagina}] Bloco tr:nth-child finalizado. "
            f"Downloads OK: {downloads_ok}. Sem Edit/erro: {sem_edit_ou_erro}."
        )
    else:
        print(f"[PÁGINA {numero_pagina}] Bloco tr:nth-child ignorado por configuração do .env.")

    return {
        "downloads_ok": downloads_ok,
        "sem_edit_ou_erro": sem_edit_ou_erro,
    }


def clicar_proximo_se_existir(page, numero_pagina):
    print(f"[PÁGINA {numero_pagina}] Procurando botão Próximo...")

    candidatos = [
        lambda: page.get_by_role("button", name="Próximo").first,
        lambda: page.get_by_role("link", name="Próximo").first,
        lambda: page.locator("a[aria-label='Próximo']").first,
        lambda: page.locator("button[aria-label='Próximo']").first,
    ]

    for criar_locator in candidatos:
        try:
            botao_proximo = criar_locator()
            botao_proximo.wait_for(state="visible", timeout=5000)

            try:
                if not botao_proximo.is_enabled(timeout=2000):
                    print(f"[PÁGINA {numero_pagina}] Próximo encontrado, mas desabilitado. Fim.")
                    return False
            except Exception:
                pass

            print(f"[PÁGINA {numero_pagina}] Clicando em Próximo...")
            botao_proximo.click(timeout=TIMEOUT_PADRAO)

            print(f"[PÁGINA {numero_pagina}] Aguardando próxima página carregar...")
            page.wait_for_timeout(8000)
            aguardar_carregamento(page)
            aguardar_relatorio_pronto(page)
            page.wait_for_timeout(3000)

            print(f"[PÁGINA {numero_pagina}] Próxima página carregada. URL atual: {page.url}")
            return True

        except Exception:
            continue

    print(f"[PÁGINA {numero_pagina}] Não encontrei botão Próximo disponível. Processo encerrado para este relatório.")
    return False


def processar_todas_as_paginas(page, relatorio):
    print("=" * 100)
    print(f"[INÍCIO] Processamento do relatório: {relatorio['nome']}")
    print("=" * 100)

    total_paginas = 0
    total_downloads_ok = 0
    total_sem_edit_ou_erro = 0

    numero_pagina = 1

    while True:
        resumo = processar_pagina_atual(page, numero_pagina, relatorio)

        total_paginas += 1
        total_downloads_ok += resumo["downloads_ok"]
        total_sem_edit_ou_erro += resumo["sem_edit_ou_erro"]

        conseguiu_ir_proxima = clicar_proximo_se_existir(page, numero_pagina)

        if not conseguiu_ir_proxima:
            print(f"[FIM] Relatório {relatorio['nome']} finalizado.")
            break

        numero_pagina += 1

    print("=" * 100)
    print(f"[RESUMO DO RELATÓRIO] {relatorio['nome']}")
    print(f"Total de páginas processadas: {total_paginas}")
    print(f"Total de downloads OK: {total_downloads_ok}")
    print(f"Total sem Edit/erro/pulados: {total_sem_edit_ou_erro}")
    print(f"Pasta de downloads: {pasta_download_relatorio(relatorio)}")
    print("=" * 100)

    return {
        "relatorio": relatorio["nome"],
        "paginas": total_paginas,
        "downloads_ok": total_downloads_ok,
        "sem_edit_ou_erro": total_sem_edit_ou_erro,
    }


# ============================================================
# FLUXO PRINCIPAL
# ============================================================
def run():
    validar_env()

    print("=" * 100)
    print("[CONFIGURAÇÃO] Relatórios carregados do .env")
    for idx, relatorio in enumerate(RELATORIOS, start=1):
        print(
            f"{idx}. {relatorio['nome']} | "
            f"APEX={relatorio['apex_id']} | "
            f"VALUE={relatorio['saved_value']} | "
            f"Editar={relatorio['editar_inicio']}..{relatorio['editar_fim']} | "
            f"TR={relatorio['tr_inicio']}..{relatorio['tr_fim']}"
        )
    print("=" * 100)

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=False,
            slow_mo=300,
        )

        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(TIMEOUT_PADRAO)

        resumos = []

        try:
            login_e_abrir_execucao_testes(page)

            for indice, relatorio in enumerate(RELATORIOS, start=1):
                print("=" * 100)
                print(f"[LOOP] Iniciando relatório {indice}/{len(RELATORIOS)}: {relatorio['nome']}")
                print("=" * 100)

                selecionar_relatorio_apex(page, relatorio)
                resumo = processar_todas_as_paginas(page, relatorio)
                resumos.append(resumo)

                print(f"[LOOP] Relatório finalizado: {relatorio['nome']}")
                page.wait_for_timeout(3000)
                aguardar_carregamento(page)

            print("=" * 100)
            print("[RESUMO GERAL]")
            for resumo in resumos:
                print(
                    f"- {resumo['relatorio']}: "
                    f"páginas={resumo['paginas']}, "
                    f"downloads_ok={resumo['downloads_ok']}, "
                    f"sem_edit_ou_erro={resumo['sem_edit_ou_erro']}"
                )
            print(f"Pasta raiz dos downloads: {DOWNLOAD_ROOT}")
            print("=" * 100)

            print("[INFO] O navegador ficará aberto para validação visual.")
            input("Pressione ENTER aqui no terminal para fechar o navegador...")

        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    try:
        run()
    except PlaywrightTimeoutError as e:
        print("[ERRO] Timeout no Playwright.")
        print(e)
        raise
    except Exception as e:
        print("[ERRO] Falha geral.")
        print(e)
        raise
